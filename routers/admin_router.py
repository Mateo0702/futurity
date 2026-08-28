from flask import Blueprint, render_template, request, jsonify, session, send_file, redirect, url_for, flash
from datetime import datetime, timedelta, date
from db_config import get_db_connection
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER

admin_bp = Blueprint('admin', __name__)

from urllib.parse import urlencode

@admin_bp.route('/admin/control_calidad')
def dashboard_calidad():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session.get('user_role') not in ['ADMIN', 'ASESOR']:
        flash('No tienes permiso para acceder al control de calidad.', 'danger')
        return redirect(url_for('dashboard'))

    # Redireccionar al dashboard de la app con el parámetro tab=control-calidad y pasar los filtros
    args = request.args.to_dict()
    args['tab'] = 'control-calidad'
    query_string = urlencode(args)
    return redirect(f"/?{query_string}")


@admin_bp.route('/api/admin/control_calidad/datos')
def api_dashboard_calidad():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'id_usuario': session['user_id'], 'role': session.get('user_role'), 'rol': session.get('user_role')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
    
    user_role = user.get('role') or user.get('rol')
    if user_role not in ['ADMIN', 'ASESOR', 'CALIDAD']:
        return jsonify({"status": "error", "message": "No tienes privilegios para ver datos de control de calidad."}), 403
        
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    try:
        cursor = conexion.cursor(dictionary=True)
        # Capturar parámetros de fecha, cliente y es_instalacion, usando hoy por defecto
        hoy = datetime.now().strftime('%Y-%m-%d')
        fecha_inicio = request.args.get('fecha_inicio', hoy)
        fecha_fin = request.args.get('fecha_fin', hoy)
        cliente_filtro = request.args.get('cliente', '').strip()
        es_instalacion = request.args.get('es_instalacion', '').strip()

        # Construir cláusula WHERE común
        base_where = "WHERE calificacion_estrellas IS NOT NULL AND fecha_programada >= %s AND fecha_programada <= %s"
        params = [fecha_inicio, fecha_fin]

        if cliente_filtro:
            base_where += " AND cliente LIKE %s"
            params.append(f"%{cliente_filtro}%")

        if es_instalacion in ['0', '1']:
            base_where += " AND es_instalacion = %s"
            params.append(int(es_instalacion))
        
        # 1. Consulta para los KPIs Generales (incluye promedios de encuesta 1-10)
        cursor.execute(f"""
            SELECT 
                ROUND(AVG(COALESCE((encuesta_rapidez + encuesta_atencion + encuesta_explicacion) / 3.0, calificacion_estrellas * 2.0)), 2) AS promedio_global,
                COUNT(calificacion_estrellas) AS total_calificadas,
                SUM(CASE WHEN COALESCE((encuesta_rapidez + encuesta_atencion + encuesta_explicacion) / 3.0, calificacion_estrellas * 2.0) <= 6.0 THEN 1 ELSE 0 END) AS alertas_criticas,
                ROUND(AVG(encuesta_rapidez), 2) AS promedio_rapidez,
                ROUND(AVG(encuesta_atencion), 2) AS promedio_atencion,
                ROUND(AVG(encuesta_explicacion), 2) AS promedio_explicacion
            FROM visitas_tecnicas
            {base_where}
        """, params)
        kpis = cursor.fetchone()
        
        # Validar nulos en los KPIs
        if kpis:
            if kpis['promedio_global'] is None: kpis['promedio_global'] = 0.0
            if kpis['total_calificadas'] is None: kpis['total_calificadas'] = 0
            if kpis['alertas_criticas'] is None: kpis['alertas_criticas'] = 0
            if kpis['promedio_rapidez'] is None: kpis['promedio_rapidez'] = 0.0
            if kpis['promedio_atencion'] is None: kpis['promedio_atencion'] = 0.0
            if kpis['promedio_explicacion'] is None: kpis['promedio_explicacion'] = 0.0
        else:
            kpis = {
                "promedio_global": 0.0, 
                "total_calificadas": 0, 
                "alertas_criticas": 0,
                "promedio_rapidez": 0.0,
                "promedio_atencion": 0.0,
                "promedio_explicacion": 0.0
            }

        # 2. Consulta para el Ranking de Técnicos (Alimentar Gráfico)
        cursor.execute(f"""
            SELECT 
                tecnico_principal AS nombre,
                ROUND(AVG(COALESCE((encuesta_rapidez + encuesta_atencion + encuesta_explicacion) / 3.0, calificacion_estrellas * 2.0)), 2) AS promedio,
                COUNT(calificacion_estrellas) AS total_visitas,
                CAST(IFNULL(SUM(CASE WHEN COALESCE((encuesta_rapidez + encuesta_atencion + encuesta_explicacion) / 3.0, calificacion_estrellas * 2.0) >= 7.0 THEN 1 ELSE 0 END), 0) AS UNSIGNED) AS buenas,
                CAST(IFNULL(SUM(CASE WHEN COALESCE((encuesta_rapidez + encuesta_atencion + encuesta_explicacion) / 3.0, calificacion_estrellas * 2.0) <= 6.0 THEN 1 ELSE 0 END), 0) AS UNSIGNED) AS malas
            FROM visitas_tecnicas
            {base_where}
            GROUP BY tecnico_principal
            ORDER BY total_visitas DESC, promedio DESC
        """, params)
        ranking_tecnicos = cursor.fetchall()

        # 3. Consulta para la Tabla con Filtros Aplicados
        query_tabla = f"""
            SELECT id_visita, cliente, sector, tecnico_principal, 
                   calificacion_estrellas, calificacion_comentario, hora_fin_visita,
                   encuesta_rapidez, encuesta_atencion, encuesta_explicacion
            FROM visitas_tecnicas
            {base_where}
            ORDER BY hora_fin_visita DESC LIMIT 100
        """
        cursor.execute(query_tabla, params)
        resenas_detalladas = cursor.fetchall()
        
        # Formatear fechas
        for r in resenas_detalladas:
            if r['hora_fin_visita']:
                r['hora_fin_visita'] = r['hora_fin_visita'].isoformat()
        
        # Traer lista de técnicos para el combobox del filtro
        cursor.execute("SELECT nombre FROM tecnicos WHERE activo = 1")
        lista_tecnicos = cursor.fetchall()

        return jsonify({
            "status": "ok",
            "kpis": kpis,
            "ranking": ranking_tecnicos,
            "resenas": resenas_detalladas,
            "tecnicos": lista_tecnicos,
            "filtros": {'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin, 'cliente': cliente_filtro}
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()

@admin_bp.route('/api/admin/auditoria_cliente', methods=['GET'])
def auditoria_cliente():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'id_usuario': session['user_id'], 'rol': session.get('user_role'), 'role': session.get('user_role')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
    
    user_role = user.get('role') or user.get('rol')
    if user_role not in ['ADMIN', 'ASESOR', 'CALIDAD', 'ATC']:
        return jsonify({"status": "error", "message": "No tienes privilegios para realizar auditorías de clientes."}), 403

    contrato = request.args.get('contrato', '').strip()
    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')
    
    # Determinar si es Fibracom o Servicable basándose en el sufijo 'F'
    is_fibracom = contrato.upper().endswith('F')
    
    if is_fibracom:
        contrato_base = contrato[:-1]
        contrato_directorio = contrato
    else:
        contrato_base = contrato
        contrato_directorio = contrato
        
    conexion = get_db_connection()
    if not conexion:
        return jsonify([])
    cursor = conexion.cursor(dictionary=True)
    
    # Obtener el nombre oficial del cliente desde el directorio
    cursor.execute("""
        SELECT nombre_cliente FROM directorio_clientes WHERE contrato = %s
    """, (contrato_directorio,))
    cliente_dir = cursor.fetchone()
    nombre_oficial = cliente_dir['nombre_cliente'] if cliente_dir else None
    
    # 1. Obtener Visitas Técnicas
    if is_fibracom:
        query_visitas = """
            SELECT id_visita, fecha_programada, cliente, problema, solucion_tecnico, 
                   observacion_tecnico, tecnico_principal, tecnico_apoyo, 
                   estado, calificacion_estrellas, calificacion_comentario,
                   'visita_tecnica' AS tipo_registro
            FROM visitas_tecnicas 
            WHERE contrato = %s AND empresa = 'FIBRACOM'
        """
    else:
        query_visitas = """
            SELECT id_visita, fecha_programada, cliente, problema, solucion_tecnico, 
                   observacion_tecnico, tecnico_principal, tecnico_apoyo, 
                   estado, calificacion_estrellas, calificacion_comentario,
                   'visita_tecnica' AS tipo_registro
            FROM visitas_tecnicas 
            WHERE contrato = %s AND (empresa != 'FIBRACOM' OR empresa IS NULL)
        """
    params_visitas = [contrato_base]
    if desde:
        query_visitas += " AND fecha_programada >= %s"
        params_visitas.append(desde)
    if hasta:
        query_visitas += " AND fecha_programada <= %s"
        params_visitas.append(hasta)
        
    cursor.execute(query_visitas, tuple(params_visitas))
    visitas = cursor.fetchall()
    
    # 2. Obtener Atenciones Diarias
    query_atenciones = """
        SELECT id_atencion AS id_visita, fecha AS fecha_programada, cliente, tipo_solicitud AS problema, 
               accion AS solucion_tecnico, observacion AS observacion_tecnico, 
               agente AS tecnico_principal, medio_contacto AS tecnico_apoyo,
               'FINALIZADA' AS estado, NULL AS calificacion_estrellas, NULL AS calificacion_comentario,
               'atencion' AS tipo_registro, hora, olt, ont, router, timer_minutos
        FROM atenciones 
        WHERE contrato = %s
    """
    params_atenciones = [contrato_directorio]
    if desde:
        query_atenciones += " AND fecha >= %s"
        params_atenciones.append(desde)
    if hasta:
        query_atenciones += " AND fecha <= %s"
        params_atenciones.append(hasta)
        
    cursor.execute(query_atenciones, tuple(params_atenciones))
    atenciones = cursor.fetchall()
    
    cursor.close()
    conexion.close()
    
    # Unificar nombre de cliente si se encontró en el directorio
    if nombre_oficial:
        for v in visitas:
            v['cliente'] = nombre_oficial
        for a in atenciones:
            a['cliente'] = nombre_oficial
    
    # Combinar ambas listas
    from datetime import time as dt_time
    resultados = visitas + atenciones
    
    if not resultados and nombre_oficial:
        resultados = [{"cliente": nombre_oficial, "tipo_registro": "meta", "estado": "SIN_REGISTROS", "fecha_programada": ""}]
    
    def get_sort_key(item):
        d_val = item['fecha_programada']
        if isinstance(d_val, (datetime, date)):
            d = d_val
        else:
            try:
                d = datetime.strptime(str(d_val), "%Y-%m-%d").date()
            except:
                d = date.min
        
        t = dt_time.min
        if item['tipo_registro'] == 'atencion' and item.get('hora'):
            h_val = item['hora']
            if isinstance(h_val, dt_time):
                t = h_val
            elif isinstance(h_val, timedelta):
                tot_sec = int(h_val.total_seconds())
                t = dt_time(hour=(tot_sec // 3600) % 24, minute=(tot_sec // 60) % 60, second=tot_sec % 60)
            else:
                try:
                    t = datetime.strptime(str(h_val), "%H:%M:%S").time()
                except:
                    pass
        return (d, t)
        
    resultados.sort(key=get_sort_key, reverse=True)
    
    # Convertir las fechas a cadenas ISO para JSON
    for res in resultados:
        if isinstance(res['fecha_programada'], (datetime, date)):
            res['fecha_programada'] = res['fecha_programada'].isoformat()
        if 'hora' in res:
            if isinstance(res['hora'], dt_time):
                res['hora'] = res['hora'].isoformat()
            elif isinstance(res['hora'], timedelta):
                tot_sec = int(res['hora'].total_seconds())
                res['hora'] = f"{(tot_sec // 3600) % 24:02d}:{(tot_sec // 60) % 60:02d}:{tot_sec % 60:02d}"
            elif res['hora'] is not None:
                res['hora'] = str(res['hora'])
            
    return jsonify(resultados)

@admin_bp.route('/api/admin/tecnicos/ubicaciones', methods=['GET'])
def api_tecnicos_ubicaciones():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'id_usuario': session['user_id'], 'role': session.get('user_role'), 'rol': session.get('user_role')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
    
    user_role = user.get('role') or user.get('rol')
    if user_role not in ['ADMIN', 'ASESOR', 'CALIDAD', 'ATC']:
        return jsonify({"status": "error", "message": "No tienes privilegios para ver la ubicación de los técnicos."}), 403
        
    active_area = request.args.get('area') or session.get('active_area', 'SOPORTE')
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    cursor = conexion.cursor(dictionary=True)
    try:
        # Consulta para traer la ubicación actual global de cada técnico y su estado de conexión
        query = """
            SELECT id_tecnico,
                   nombre AS tecnico, 
                   latitud_actual AS lat, 
                   longitud_actual AS lon, 
                   ultima_conexion AS ultima_actualizacion, 
                   estado_actividad AS estado, 
                   foto_perfil,
                   foto_vehiculo,
                   placa_vehiculo,
                   alerta_panico,
                   mensaje_panico,
                   CASE 
                      WHEN latitud_actual IS NOT NULL 
                       AND longitud_actual IS NOT NULL 
                       AND estado_actividad NOT IN ('Desconectado', 'DESCONECTADO', 'Inactivo', 'INACTIVO') 
                       AND (ultima_conexion >= DATE_SUB(NOW(), INTERVAL 15 MINUTE) OR alerta_panico = 1) 
                        THEN 1 
                      ELSE 0 
                    END AS conectado
            FROM tecnicos
            WHERE activo = 1 
              AND area_trabajo = %s
              AND UPPER(nombre) NOT LIKE '%NO TECNICO%'
              AND UPPER(nombre) NOT LIKE '%TECNOLOGIA%'
              AND UPPER(nombre) NOT LIKE '%TECNOLOGÍA%'
        """
        cursor.execute(query, (active_area,))
        ubicaciones = cursor.fetchall()
        
        # Formatear fecha/hora a ISO para serialización JSON
        for u in ubicaciones:
            if u['ultima_actualizacion']:
                u['ultima_actualizacion'] = u['ultima_actualizacion'].isoformat()
            
        return jsonify({"status": "ok", "ubicaciones": ubicaciones})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()

@admin_bp.route('/api/admin/metricas_globales', methods=['GET'])
def metricas_globales():
    if 'user_id' not in session:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
    if session.get('user_role') not in ['ADMIN', 'ASESOR', 'CALIDAD', 'ATC']:
        return jsonify({"status": "error", "message": "No tienes privilegios para ver métricas globales."}), 403

    active_area = session.get('active_area', 'SOPORTE')
    es_instalacion_val = 1 if active_area == 'INSTALACIONES' else 0

    # Obtener parámetros de filtros (hoy y hace 3 meses por defecto si no se especifican)
    hoy_dt = date.today()
    hace_tres_meses = (hoy_dt - timedelta(days=90)).isoformat()
    hoy_str = hoy_dt.isoformat()

    fecha_inicio = request.args.get('fecha_inicio', hace_tres_meses)
    if not fecha_inicio:
        fecha_inicio = hace_tres_meses
    fecha_fin = request.args.get('fecha_fin', hoy_str)
    if not fecha_fin:
        fecha_fin = hoy_str
    tecnico_filtro = request.args.get('tecnico', '').strip()

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "No se pudo conectar a la base de datos"}), 500
    
    cursor = conexion.cursor(dictionary=True)
    try:
        # Cláusula WHERE común
        where_clause = "WHERE fecha_programada >= %s AND fecha_programada <= %s AND es_instalacion = %s"
        params = [fecha_inicio, fecha_fin, es_instalacion_val]

        if tecnico_filtro and tecnico_filtro != 'TODOS':
            where_clause += " AND (tecnico_principal = %s OR tecnico_apoyo = %s)"
            params.extend([tecnico_filtro, tecnico_filtro])

        # 1. Total visitas/instalaciones y KPIs
        query_kpis = f"""
            SELECT 
                COUNT(*) as total_visitas,
                SUM(CASE WHEN estado IN ('FINALIZADA', 'SOLVENTADA_REMOTA') 
                          AND (solucion_tecnico IS NULL OR (
                              solucion_tecnico NOT LIKE '%GESTIONAR ARREGLO%'
                              AND solucion_tecnico NOT LIKE '%SOLUCIÓN PARCIAL%'
                              AND solucion_tecnico NOT LIKE '%SOLUCION PARCIAL%'
                              AND solucion_tecnico NOT LIKE '%GENERAR CAMBIO DE FO%'
                              AND solucion_tecnico NOT LIKE '%SIN RESPUESTA%'
                              AND solucion_tecnico NOT LIKE '%NO SE PUEDE REALIZAR VISITA%'
                              AND solucion_tecnico NOT LIKE '%NOC%'
                          )) THEN 1 ELSE 0 END) as visitas_efectivas,
                AVG(CASE WHEN estado = 'FINALIZADA' AND hora_inicio_visita IS NOT NULL AND hora_fin_visita IS NOT NULL 
                         THEN TIMESTAMPDIFF(MINUTE, hora_inicio_visita, hora_fin_visita) ELSE NULL END) as tiempo_promedio
            FROM visitas_tecnicas
            {where_clause}
        """
        cursor.execute(query_kpis, params)
        kpis = cursor.fetchone()
        
        total = kpis['total_visitas'] or 0
        efectivas = kpis['visitas_efectivas'] or 0
        tasa_efectividad = round(float(efectivas / total * 100), 1) if total > 0 else 0.0
        tiempo_promedio = round(float(kpis['tiempo_promedio'] or 0), 1)
        
        # 2. Distribución por Estado
        query_estados = f"""
            SELECT estado, COUNT(*) as cantidad
            FROM visitas_tecnicas
            {where_clause}
            GROUP BY estado
        """
        cursor.execute(query_estados, params)
        estados_raw = cursor.fetchall()
        estados = {
            'PENDIENTE': 0,
            'FINALIZADA': 0,
            'REAGENDADA': 0,
            'CANCELADA': 0
        }
        for row in estados_raw:
            est = row['estado']
            cant = row['cantidad']
            if est == 'PENDIENTE':
                estados['PENDIENTE'] += cant
            elif est in ('FINALIZADA', 'SOLVENTADA_REMOTA'):
                estados['FINALIZADA'] += cant
            elif est == 'REAGENDADA':
                estados['REAGENDADA'] += cant
            elif est == 'CANCELADA':
                estados['CANCELADA'] += cant
        
        # 3. Top 3 Clientes con más visitas/instalaciones
        query_top_clientes = f"""
            SELECT cliente, contrato, COUNT(*) as total_visitas
            FROM visitas_tecnicas
            {where_clause}
              AND contrato IS NOT NULL AND contrato != ''
              AND estado IN ('FINALIZADA', 'SOLVENTADA_REMOTA')
            GROUP BY cliente, contrato
            ORDER BY total_visitas DESC
            LIMIT 3
        """
        cursor.execute(query_top_clientes, params)
        top_clientes = cursor.fetchall()
        
        # 4. Top 5 Problemas comunes (o Productos en caso de instalación)
        if es_instalacion_val == 1:
            query_problemas = f"""
                SELECT producto as problema, COUNT(*) as cantidad
                FROM visitas_tecnicas
                {where_clause}
                GROUP BY producto
                ORDER BY cantidad DESC
                LIMIT 5
            """
        else:
            query_problemas = f"""
                SELECT problema, COUNT(*) as cantidad
                FROM visitas_tecnicas
                {where_clause}
                GROUP BY problema
                ORDER BY cantidad DESC
                LIMIT 5
            """
        cursor.execute(query_problemas, params)
        top_problemas = cursor.fetchall()
        
        # 5. Evolución semanal
        query_evolucion = f"""
            SELECT 
                DATE_FORMAT(fecha_programada, '%Y-%u') as semana,
                MIN(fecha_programada) as inicio_semana,
                COUNT(*) as cantidad
            FROM visitas_tecnicas
            {where_clause}
            GROUP BY semana
            ORDER BY inicio_semana ASC
        """
        cursor.execute(query_evolucion, params)
        evolucion_raw = cursor.fetchall()
        
        evolucion = []
        for row in evolucion_raw:
            fecha_dt = row['inicio_semana']
            fecha_str = fecha_dt.strftime('%d/%m') if isinstance(fecha_dt, (datetime, date)) else str(fecha_dt)
            evolucion.append({
                'label': f"Sem {fecha_str}",
                'cantidad': row['cantidad']
            })

        # 6. Obtener lista de técnicos activos para los filtros (excluyendo no técnico y tecnología)
        cursor.execute("SELECT nombre FROM tecnicos WHERE activo = 1 AND nombre NOT IN ('NO TECNICO', 'TECNOLOGIA') ORDER BY nombre ASC")
        tecnicos = [t['nombre'] for t in cursor.fetchall()]
            
        data = {
            'status': 'ok',
            'kpis': {
                'total_visitas': total,
                'tasa_efectividad': tasa_efectividad,
                'tiempo_promedio': tiempo_promedio
            },
            'estados': estados,
            'top_clientes': top_clientes,
            'top_problemas': top_problemas,
            'evolucion': evolucion,
            'tecnicos': tecnicos
        }
        
        return jsonify(data)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super(NumberedCanvas, self).__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor('#64748b'))
        
        # Header (Only on page 2 onwards)
        if self._pageNumber > 1:
            self.setStrokeColor(colors.HexColor('#e2e8f0'))
            self.setLineWidth(0.5)
            self.line(36, 756, 576, 756)
            self.drawString(36, 762, "Futurity - Reporte de Auditoría de Cliente")
        
        # Footer (On all pages)
        self.setStrokeColor(colors.HexColor('#e2e8f0'))
        self.setLineWidth(0.5)
        self.line(36, 45, 576, 45)
        page_text = f"Página {self._pageNumber} de {page_count}"
        self.drawRightString(576, 32, page_text)
        self.drawString(36, 32, "Generado por Sistema Futurity")
        self.restoreState()


@admin_bp.route('/api/admin/reporte_pdf', methods=['GET'])
def reporte_pdf():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'id_usuario': session['user_id'], 'rol': session.get('user_role'), 'role': session.get('user_role')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
    
    user_role = user.get('role') or user.get('rol')
    if user_role not in ['ADMIN', 'ASESOR', 'CALIDAD', 'ATC']:
        return jsonify({"status": "error", "message": "No tienes privilegios para descargar reportes."}), 403

    contrato = request.args.get('contrato', '').strip()
    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')

    if not contrato:
        return jsonify({"status": "error", "message": "Debe especificar un número de contrato"}), 400

    # Determinar si es Fibracom o Servicable basándose en el sufijo 'F'
    is_fibracom = contrato.upper().endswith('F')
    
    if is_fibracom:
        contrato_base = contrato[:-1]
        contrato_directorio = contrato
    else:
        contrato_base = contrato
        contrato_directorio = contrato

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    cursor = conexion.cursor(dictionary=True)
    try:
        # 1. Obtener datos del cliente de directorio_clientes
        cursor.execute("""
            SELECT nombre_cliente AS cliente, zona AS sector, COALESCE(direccion, '-') as direccion FROM directorio_clientes WHERE contrato = %s
        """, (contrato_directorio,))
        cliente_data = cursor.fetchone()
        
        if not cliente_data:
            # Fallback a visitas_tecnicas
            if is_fibracom:
                cursor.execute("""
                    SELECT cliente, sector, direccion FROM visitas_tecnicas WHERE contrato = %s AND empresa = 'FIBRACOM' LIMIT 1
                """, (contrato_base,))
            else:
                cursor.execute("""
                    SELECT cliente, sector, direccion FROM visitas_tecnicas WHERE contrato = %s AND (empresa != 'FIBRACOM' OR empresa IS NULL) LIMIT 1
                """, (contrato_base,))
            cliente_data = cursor.fetchone()
            
            if not cliente_data:
                # Fallback a atenciones
                cursor.execute("""
                    SELECT cliente, sector, '-' as direccion FROM atenciones WHERE contrato = %s LIMIT 1
                """, (contrato_directorio,))
                cliente_data = cursor.fetchone()
        
        cliente_nombre = "Desconocido"
        cliente_sector = "-"
        cliente_direccion = "-"
        if cliente_data:
            cliente_nombre = cliente_data['cliente']
            cliente_sector = cliente_data['sector'] or "-"
            cliente_direccion = cliente_data.get('direccion') or "-"
        
        # 2. Obtener Visitas Técnicas
        if is_fibracom:
            query_visitas = """
                SELECT id_visita, fecha_programada, cliente, problema, solucion_tecnico, 
                       observacion_tecnico, tecnico_principal, tecnico_apoyo, 
                       estado, calificacion_estrellas, calificacion_comentario,
                       'visita_tecnica' AS tipo_registro
                FROM visitas_tecnicas 
                WHERE contrato = %s AND empresa = 'FIBRACOM'
            """
        else:
            query_visitas = """
                SELECT id_visita, fecha_programada, cliente, problema, solucion_tecnico, 
                       observacion_tecnico, tecnico_principal, tecnico_apoyo, 
                       estado, calificacion_estrellas, calificacion_comentario,
                       'visita_tecnica' AS tipo_registro
                FROM visitas_tecnicas 
                WHERE contrato = %s AND (empresa != 'FIBRACOM' OR empresa IS NULL)
            """
        params_visitas = [contrato_base]
        if desde:
            query_visitas += " AND fecha_programada >= %s"
            params_visitas.append(desde)
        if hasta:
            query_visitas += " AND fecha_programada <= %s"
            params_visitas.append(hasta)
        query_visitas += " ORDER BY fecha_programada DESC"
        cursor.execute(query_visitas, tuple(params_visitas))
        visitas = cursor.fetchall()
        
        # 3. Obtener Atenciones Diarias
        query_atenciones = """
            SELECT id_atencion AS id_visita, fecha AS fecha_programada, cliente, tipo_solicitud AS problema, 
                   accion AS solucion_tecnico, observacion AS observacion_tecnico, 
                   agente AS tecnico_principal, medio_contacto AS tecnico_apoyo,
                   'FINALIZADA' AS estado, NULL AS calificacion_estrellas, NULL AS calificacion_comentario,
                   'atencion' AS tipo_registro, hora, olt, ont, router, timer_minutos
            FROM atenciones 
            WHERE contrato = %s
        """
        params_atenciones = [contrato_directorio]
        if desde:
            query_atenciones += " AND fecha >= %s"
            params_atenciones.append(desde)
        if hasta:
            query_atenciones += " AND fecha <= %s"
            params_atenciones.append(hasta)
        query_atenciones += " ORDER BY fecha DESC, hora DESC"
        cursor.execute(query_atenciones, tuple(params_atenciones))
        atenciones = cursor.fetchall()
        
        # KPIs calculations
        total_visitas = len(visitas)
        total_atenciones = len(atenciones)
        
        calificadas = [v for v in visitas if v.get('calificacion_estrellas') is not None]
        if calificadas:
            promedio_estrellas = f"{sum(v['calificacion_estrellas'] for v in calificadas) / len(calificadas):.1f} \u2605"
        else:
            promedio_estrellas = "-"
            
        # Start building PDF
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=letter,
            leftMargin=36,
            rightMargin=36,
            topMargin=54,
            bottomMargin=60
        )
        
        styles = getSampleStyleSheet()
        
        # Define custom colors
        primary_color = colors.HexColor('#b91c1c') # Dark Red
        secondary_color = colors.HexColor('#1e293b') # Slate
        dark_text = colors.HexColor('#0f172a')
        muted_text = colors.HexColor('#475569')
        bg_light = colors.HexColor('#f8fafc')
        border_color = colors.HexColor('#e2e8f0')
        
        # Custom styles
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=20,
            leading=24,
            textColor=primary_color,
            spaceAfter=4
        )
        
        subtitle_style = ParagraphStyle(
            'DocSub',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            leading=12,
            textColor=muted_text,
            spaceAfter=15
        )
        
        section_title_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=12,
            leading=16,
            textColor=secondary_color,
            spaceBefore=15,
            spaceAfter=8,
            keepWithNext=True
        )
        
        meta_label_style = ParagraphStyle(
            'MetaLabel',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            leading=11,
            textColor=dark_text
        )
        
        meta_val_style = ParagraphStyle(
            'MetaVal',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            leading=11,
            textColor=muted_text
        )
        
        kpi_title_style = ParagraphStyle(
            'KPITitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=7,
            leading=9,
            alignment=TA_CENTER,
            textColor=muted_text
        )
        
        kpi_val_style = ParagraphStyle(
            'KPIVal',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=14,
            leading=16,
            alignment=TA_CENTER,
            textColor=dark_text
        )
        
        table_header_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            textColor=colors.white
        )
        
        table_cell_style = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            leading=10,
            textColor=dark_text
        )
        
        table_cell_bold_style = ParagraphStyle(
            'TableCellBold',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            textColor=dark_text
        )
        
        table_cell_italic_style = ParagraphStyle(
            'TableCellItalic',
            parent=styles['Normal'],
            fontName='Helvetica-Oblique',
            fontSize=7.5,
            leading=9,
            textColor=muted_text
        )
        
        story = []
        
        # 1. Header Banner
        story.append(Paragraph("FUTURITY \u2014 REPORTE DE AUDITOR\u00cdA DE CLIENTE", title_style))
        story.append(Paragraph("HISTORIAL CONSOLIDADO DE VISITAS Y SOPORTE DE CALL CENTER", subtitle_style))
        
        # 2. Metadata Block (Customer & Filter details)
        rango_fechas = f"Desde: {desde or 'Inicio'} | Hasta: {hasta or 'Hoy'}"
        fecha_gen = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        meta_data = [
            [
                Paragraph("Contrato:", meta_label_style), Paragraph(f"#{contrato}", meta_val_style),
                Paragraph("Fecha Generaci\u00f3n:", meta_label_style), Paragraph(fecha_gen, meta_val_style)
            ],
            [
                Paragraph("Cliente:", meta_label_style), Paragraph(cliente_nombre, meta_val_style),
                Paragraph("Rango Filtro:", meta_label_style), Paragraph(rango_fechas, meta_val_style)
            ],
            [
                Paragraph("Sector:", meta_label_style), Paragraph(cliente_sector, meta_val_style),
                Paragraph("Direcci\u00f3n:", meta_label_style), Paragraph(cliente_direccion, meta_val_style)
            ]
        ]
        
        meta_table = Table(meta_data, colWidths=[60, 200, 110, 170])
        meta_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BACKGROUND', (0,0), (-1,-1), bg_light),
            ('BOX', (0,0), (-1,-1), 0.5, border_color),
            ('INNERGRID', (0,0), (-1,-1), 0.25, border_color),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 12))
        
        # 3. KPIs Cards
        kpi_data = [
            [
                Paragraph("VISITAS T\u00c9CNICAS REGISTRADAS", kpi_title_style),
                Paragraph("ATENCIONES SOPORTE DIARIO", kpi_title_style),
                Paragraph("CALIFICACI\u00d3N PROM. VISITAS", kpi_title_style)
            ],
            [
                Paragraph(str(total_visitas), kpi_val_style),
                Paragraph(str(total_atenciones), kpi_val_style),
                Paragraph(promedio_estrellas, kpi_val_style)
            ]
        ]
        kpi_table = Table(kpi_data, colWidths=[180, 180, 180])
        kpi_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BACKGROUND', (0,0), (-1,-1), bg_light),
            ('BOX', (0,0), (-1,-1), 1, primary_color),
            ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
            ('TOPPADDING', (0,0), (-1,0), 6),
            ('BOTTOMPADDING', (0,0), (-1,0), 2),
            ('TOPPADDING', (0,1), (-1,-1), 2),
            ('BOTTOMPADDING', (0,1), (-1,-1), 6),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 15))
        
        # 4. Section 1: Visitas Técnicas
        story.append(Paragraph("1. HISTORIAL DE VISITAS T\u00c9CNICAS (VT)", section_title_style))
        
        visit_col_widths = [65, 125, 95, 65, 190]
        visit_table_data = [[
            Paragraph("Fecha", table_header_style),
            Paragraph("Problema Detectado", table_header_style),
            Paragraph("T\u00e9cnico / Apoyo", table_header_style),
            Paragraph("Estado", table_header_style),
            Paragraph("Soluci\u00f3n / Observaci\u00f3n del T\u00e9cnico", table_header_style)
        ]]
        
        if total_visitas > 0:
            for v in visitas:
                f_val = v['fecha_programada']
                f_str = f_val.strftime('%d/%m/%Y') if isinstance(f_val, (datetime, date)) else str(f_val)
                
                sol_raw = (v.get('solucion_tecnico') or '').upper()
                estado_lbl = v['estado']
                if 'NOC' in sol_raw:
                    estado_lbl = 'Ticket NOC'
                elif 'SOLUCIÓN PARCIAL' in sol_raw or 'SOLUCION PARCIAL' in sol_raw or 'GESTIONAR ARREGLO' in sol_raw:
                    estado_lbl = 'Solución Parcial'
                elif estado_lbl == 'FINALIZADA': 
                    estado_lbl = 'Efectiva'
                elif estado_lbl == 'SOLVENTADA_REMOTA': 
                    estado_lbl = 'Solv. Remota'
                
                tech_lbl = v['tecnico_principal'] or 'Sin asignar'
                if v['tecnico_apoyo']:
                    tech_lbl += f" / {v['tecnico_apoyo']}"
                    
                sol_lbl = v['solucion_tecnico'] or 'Sin soluci\u00f3n registrada.'
                if v['observacion_tecnico']:
                    sol_lbl += f"\nObs: {v['observacion_tecnico']}"
                if v['calificacion_estrellas']:
                    sol_lbl += f"\nCalificaci\u00f3n: {'★' * v['calificacion_estrellas']} ({v['calificacion_comentario'] or 'Sin comentarios'})"
                    
                sol_paragraph = Paragraph(sol_lbl.replace('\n', '<br/>'), table_cell_style)
                
                visit_table_data.append([
                    Paragraph(f_str, table_cell_bold_style),
                    Paragraph(v['problema'] or '-', table_cell_style),
                    Paragraph(tech_lbl, table_cell_style),
                    Paragraph(estado_lbl, table_cell_bold_style),
                    sol_paragraph
                ])
        else:
            visit_table_data.append([
                Paragraph("No se registraron visitas t\u00e9cnicas en el per\u00edodo seleccionado.", table_cell_italic_style),
                "", "", "", ""
            ])
            
        visit_table = Table(visit_table_data, colWidths=visit_col_widths, repeatRows=1)
        visit_table_style = [
            ('BACKGROUND', (0,0), (-1,0), primary_color),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOX', (0,0), (-1,-1), 0.5, border_color),
            ('INNERGRID', (0,0), (-1,-1), 0.25, border_color),
        ]
        if total_visitas == 0:
            visit_table_style.append(('SPAN', (0,1), (4,1)))
        else:
            for idx in range(1, len(visit_table_data)):
                if idx % 2 == 0:
                    visit_table_style.append(('BACKGROUND', (0,idx), (-1,idx), bg_light))
                    
        visit_table.setStyle(TableStyle(visit_table_style))
        story.append(visit_table)
        story.append(Spacer(1, 15))
        
        # 5. Section 2: Atenciones de Soporte
        story.append(Paragraph("2. HISTORIAL DE ATENCIONES DE SOPORTE DIARIO", section_title_style))
        
        aten_col_widths = [85, 110, 70, 95, 180]
        aten_table_data = [[
            Paragraph("Fecha / Hora", table_header_style),
            Paragraph("Solicitud / Motivo", table_header_style),
            Paragraph("Medio / Canal", table_header_style),
            Paragraph("Agente Responsable", table_header_style),
            Paragraph("Acci\u00f3n / Detalle de la Atenci\u00f3n", table_header_style)
        ]]
        
        if total_atenciones > 0:
            for a in atenciones:
                f_val = a['fecha_programada']
                f_str = f_val.strftime('%d/%m/%Y') if isinstance(f_val, (datetime, date)) else str(f_val)
                
                h_str = ""
                if a.get('hora'):
                    h_val = a['hora']
                    if hasattr(h_val, 'strftime'):
                        h_str = h_val.strftime(' %H:%M')
                    elif isinstance(h_val, timedelta):
                        tot_sec = int(h_val.total_seconds())
                        h_str = f" {tot_sec // 3600:02d}:{(tot_sec % 3600) // 60:02d}"
                    else:
                        h_str = f" {str(h_val)[:5]}"
                
                via_lbl = a['tecnico_apoyo'] or 'WhatsApp'
                agent_lbl = a['tecnico_principal'] or 'Call Center'
                sol_lbl = a['solucion_tecnico'] or '-'
                if a['observacion_tecnico']:
                    sol_lbl += f"\nObs: {a['observacion_tecnico']}"
                if a.get('olt'):
                    sol_lbl += f" (OLT: {a['olt']})"
                    
                sol_paragraph = Paragraph(sol_lbl.replace('\n', '<br/>'), table_cell_style)
                
                aten_table_data.append([
                    Paragraph(f_str + h_str, table_cell_bold_style),
                    Paragraph(a['problema'] or '-', table_cell_style),
                    Paragraph(via_lbl, table_cell_style),
                    Paragraph(agent_lbl, table_cell_style),
                    sol_paragraph
                ])
        else:
            aten_table_data.append([
                Paragraph("No se registraron atenciones de soporte en el per\u00edodo seleccionado.", table_cell_italic_style),
                "", "", "", ""
            ])
            
        aten_table = Table(aten_table_data, colWidths=aten_col_widths, repeatRows=1)
        aten_table_style = [
            ('BACKGROUND', (0,0), (-1,0), secondary_color),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOX', (0,0), (-1,-1), 0.5, border_color),
            ('INNERGRID', (0,0), (-1,-1), 0.25, border_color),
        ]
        if total_atenciones == 0:
            aten_table_style.append(('SPAN', (0,1), (4,1)))
        else:
            for idx in range(1, len(aten_table_data)):
                if idx % 2 == 0:
                    aten_table_style.append(('BACKGROUND', (0,idx), (-1,idx), bg_light))
                    
        aten_table.setStyle(TableStyle(aten_table_style))
        story.append(aten_table)
        
        # Build Document
        doc.build(story, canvasmaker=NumberedCanvas)
        pdf_buffer.seek(0)
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"Reporte_Auditoria_{contrato}.pdf"
        )
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


def format_datetime_val(dt_val):
    if not dt_val:
        return ""
    if isinstance(dt_val, (datetime, date)):
        return dt_val.strftime('%d/%m/%y %H:%M')
    try:
        dt = datetime.fromisoformat(str(dt_val))
        return dt.strftime('%d/%m/%y %H:%M')
    except:
        return str(dt_val)


def normalizar_servicio_excel_calidad(serv_raw):
    if not serv_raw:
        return "INTERNET_GPON"
    s = str(serv_raw).strip().upper()
    
    # 1. Si es COMBO -> Siempre COMBO_GPON
    if 'COMBO' in s:
        return 'COMBO_GPON'
        
    # 2. Si es CABLE / TELEVISIÓN -> CABLE_HFC o CABLE_GPON
    if 'CABLE' in s or 'TV' in s or 'CANAL' in s or 'TELEVISION' in s:
        if 'HFC' in s or 'COAX' in s or 'ANALOG' in s:
            return 'CABLE_HFC'
        return 'CABLE_GPON'
        
    # 3. Si es INTERNET -> Siempre INTERNET_GPON
    if any(k in s for k in ['INTER', 'MEGA', 'GBPS', 'MBPS', 'GPON', 'NET', 'CONECTIVIDAD', 'PYME', 'HOME', 'HOUSE']):
        return 'INTERNET_GPON'
        
    # 4. Otros casos
    if 'RADIO' in s:
        return 'RADIO'
    if 'SMART HOME' in s:
        return 'SMART HOME'
        
    return 'INTERNET_GPON'


def generar_excel_calidad(visitas, fecha_str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visitas Efectivas"
    
    # Asegurar que se muestre la cuadrícula
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "FECHA Y HORA",
        "CONTRATO",
        "CLIENTE",
        "TELÉFONOS",
        "SECTOR",
        "SERVICIO",
        "SOLUCIÓN DEL TÉCNICO",
        "OBSERVACION TECNICO",
        "TÉCNICO/TECNICOS QUE REALIZAN LA ACTIVIDAD",
        "",  # Espacio para el técnico de apoyo (segunda columna)
        "FINALIZACIÓN"
    ]
    
    # Escribir cabecera
    ws.append(headers)
    
    # Combinar celdas de técnico (I1 y J1)
    ws.merge_cells("I1:J1")
    
    # Estilos cabecera
    header_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_side = Side(style='thin', color='A0A0A0')
    border_style = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    for col_idx in range(1, 12):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border_style
        
    ws.row_dimensions[1].height = 32
    
    # Estilos datos
    data_font = Font(name="Calibri", size=10)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    row_num = 2
    for v in visitas:
        # Formatear fechas de forma segura
        f_reg = format_datetime_val(v.get('fecha_registro'))
        f_fin = format_datetime_val(v.get('hora_fin_visita'))
        serv_formateado = normalizar_servicio_excel_calidad(v.get('servicio'))
        
        row_data = [
            f_reg,
            str(v.get('contrato') or ''),
            str(v.get('cliente') or '').upper(),
            str(v.get('telefonos') or ''),
            str(v.get('sector') or '').upper(),
            serv_formateado,
            str(v.get('solucion_tecnico') or '').upper(),
            str(v.get('observacion_tecnico') or ''),
            str(v.get('tecnico_principal') or '').upper(),
            str(v.get('tecnico_apoyo') or '').upper(),
            f_fin
        ]
        ws.append(row_data)
        
        for col_idx in range(1, 12):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = data_font
            cell.border = border_style
            
            # Alineaciones específicas
            if col_idx in [1, 2, 11]:  # Fecha registro, contrato, finalizacion
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
        ws.row_dimensions[row_num].height = 24
        row_num += 1
        
    # Ajuste automático de anchos de columnas
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            # Omitimos la primera fila para algunas columnas largas en el cálculo para evitar que se ensanchen demasiado
            if cell.row == 1 and col_letter in ['C', 'G', 'H']:
                continue
            if cell.value:
                # Si tiene saltos de línea, consideramos la línea más larga
                lines = str(cell.value).split('\n')
                for line in lines:
                    max_len = max(max_len, len(line))
        
        width = max(max_len + 3, 12)
        # Límites por columna para estética premium
        if col_letter == 'C':  # Cliente
            width = min(width, 35)
        elif col_letter in ['G', 'H']:  # Solución, Observación
            width = min(width, 45)
        elif col_letter == 'D':  # Teléfonos
            width = min(width, 18)
        elif col_letter == 'F':  # Servicio
            width = min(width, 16)
        
        ws.column_dimensions[col_letter].width = width
        
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


@admin_bp.route('/api/admin/reporte_calidad/preview', methods=['GET'])
def preview_reporte_calidad():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'id_usuario': session['user_id'], 'role': session.get('user_role'), 'rol': session.get('user_role')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
        
    fecha = request.args.get('fecha', date.today().isoformat())
    es_instalacion = request.args.get('es_instalacion', '0').strip()
    
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    try:
        cursor = conexion.cursor(dictionary=True)
        
        fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date()
        is_sunday = (fecha_dt.weekday() == 6)
        
        if is_sunday:
            query = """
                SELECT 
                    id_visita, 
                    fecha_registro, 
                    contrato, 
                    cliente, 
                    telefonos, 
                    sector, 
                    servicio, 
                    solucion_tecnico, 
                    observacion_tecnico, 
                    tecnico_principal, 
                    tecnico_apoyo,
                    hora_fin_visita,
                    foto_equipos,
                    foto_equipos_2,
                    firma_cliente,
                    modelo_onu,
                    modelo_router,
                    coordenadas_tecnico,
                    latitud_inicio,
                    longitud_inicio,
                    foto_extra_1,
                    foto_extra_2,
                    foto_extra_3,
                    foto_extra_4,
                    equipos_juntos
                FROM visitas_tecnicas
                WHERE COALESCE(DATE(hora_fin_visita), fecha_programada) BETWEEN DATE_SUB(%s, INTERVAL 2 DAY) AND %s AND estado = 'FINALIZADA'
                  AND tecnico_principal IS NOT NULL 
                  AND tecnico_principal NOT IN ('', 'NO TECNICO', 'SIN ASIGNAR', 'NONE', 'NAN')
                  AND solucion_tecnico IS NOT NULL 
                  AND solucion_tecnico NOT IN (
                      'NO SE PUEDE REALIZAR VISITA - SATURACIÓN DEL DÍA', 
                      'SIN RESPUESTA DEL CLIENTE',
                      'GENERAR CAMBIO DE FO',
                      'GENERAR ARREGLO DE INSTALACIÓN',
                      'GESTIONAR ARREGLO DE INSTALACIÓN'
                  )
                  AND solucion_tecnico NOT LIKE '%NOC%'
                  AND solucion_tecnico NOT LIKE '%PARCIAL%'
                  AND es_instalacion = %s
                ORDER BY COALESCE(DATE(hora_fin_visita), fecha_programada) ASC, hora_fin_visita ASC
            """
            cursor.execute(query, (fecha, fecha, int(es_instalacion)))
        else:
            query = """
                SELECT 
                    id_visita, 
                    fecha_registro, 
                    contrato, 
                    cliente, 
                    telefonos, 
                    sector, 
                    servicio, 
                    solucion_tecnico, 
                    observacion_tecnico, 
                    tecnico_principal, 
                    tecnico_apoyo,
                    hora_fin_visita,
                    foto_equipos,
                    foto_equipos_2,
                    firma_cliente,
                    modelo_onu,
                    modelo_router,
                    coordenadas_tecnico,
                    latitud_inicio,
                    longitud_inicio,
                    foto_extra_1,
                    foto_extra_2,
                    foto_extra_3,
                    foto_extra_4,
                    equipos_juntos
                FROM visitas_tecnicas
                WHERE COALESCE(DATE(hora_fin_visita), fecha_programada) = %s AND estado = 'FINALIZADA'
                  AND tecnico_principal IS NOT NULL 
                  AND tecnico_principal NOT IN ('', 'NO TECNICO', 'SIN ASIGNAR', 'NONE', 'NAN')
                  AND solucion_tecnico IS NOT NULL 
                  AND solucion_tecnico NOT IN (
                      'NO SE PUEDE REALIZAR VISITA - SATURACIÓN DEL DÍA', 
                      'SIN RESPUESTA DEL CLIENTE',
                      'GENERAR CAMBIO DE FO',
                      'GENERAR ARREGLO DE INSTALACIÓN',
                      'GESTIONAR ARREGLO DE INSTALACIÓN'
                  )
                  AND solucion_tecnico NOT LIKE '%NOC%'
                  AND solucion_tecnico NOT LIKE '%PARCIAL%'
                  AND es_instalacion = %s
                ORDER BY hora_fin_visita ASC
            """
            cursor.execute(query, (fecha, int(es_instalacion)))
        visitas = cursor.fetchall()
        
        # Serializar objetos datetime a formato legible/ISO para JSON
        for v in visitas:
            if v['fecha_registro']:
                v['fecha_registro'] = v['fecha_registro'].isoformat()
            if v['hora_fin_visita']:
                v['hora_fin_visita'] = v['hora_fin_visita'].isoformat()
                
        return jsonify({"status": "ok", "visitas": visitas})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/reporte_calidad/excel', methods=['GET'])
def download_excel_reporte_calidad():
    if 'user_id' not in session:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
        
    fecha = request.args.get('fecha', date.today().isoformat())
    es_instalacion = request.args.get('es_instalacion', '0').strip()
    
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    try:
        cursor = conexion.cursor(dictionary=True)
        
        fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date()
        is_sunday = (fecha_dt.weekday() == 6)
        
        if is_sunday:
            query = """
                SELECT 
                    fecha_registro, 
                    contrato, 
                    cliente, 
                    telefonos, 
                    sector, 
                    servicio, 
                    solucion_tecnico, 
                    observacion_tecnico, 
                    tecnico_principal, 
                    tecnico_apoyo,
                    hora_fin_visita
                FROM visitas_tecnicas
                WHERE COALESCE(DATE(hora_fin_visita), fecha_programada) BETWEEN DATE_SUB(%s, INTERVAL 2 DAY) AND %s AND estado = 'FINALIZADA'
                  AND tecnico_principal IS NOT NULL 
                  AND tecnico_principal NOT IN ('', 'NO TECNICO', 'SIN ASIGNAR', 'NONE', 'NAN')
                  AND solucion_tecnico IS NOT NULL 
                  AND solucion_tecnico NOT IN (
                      'NO SE PUEDE REALIZAR VISITA - SATURACIÓN DEL DÍA', 
                      'SIN RESPUESTA DEL CLIENTE',
                      'GENERAR CAMBIO DE FO',
                      'GENERAR ARREGLO DE INSTALACIÓN',
                      'GESTIONAR ARREGLO DE INSTALACIÓN'
                  )
                  AND solucion_tecnico NOT LIKE '%NOC%'
                  AND solucion_tecnico NOT LIKE '%PARCIAL%'
                  AND es_instalacion = %s
                ORDER BY COALESCE(DATE(hora_fin_visita), fecha_programada) ASC, hora_fin_visita ASC
            """
            cursor.execute(query, (fecha, fecha, int(es_instalacion)))
        else:
            query = """
                SELECT 
                    fecha_registro, 
                    contrato, 
                    cliente, 
                    telefonos, 
                    sector, 
                    servicio, 
                    solucion_tecnico, 
                    observacion_tecnico, 
                    tecnico_principal, 
                    tecnico_apoyo,
                    hora_fin_visita
                FROM visitas_tecnicas
                WHERE COALESCE(DATE(hora_fin_visita), fecha_programada) = %s AND estado = 'FINALIZADA'
                  AND tecnico_principal IS NOT NULL 
                  AND tecnico_principal NOT IN ('', 'NO TECNICO', 'SIN ASIGNAR', 'NONE', 'NAN')
                  AND solucion_tecnico IS NOT NULL 
                  AND solucion_tecnico NOT IN (
                      'NO SE PUEDE REALIZAR VISITA - SATURACIÓN DEL DÍA', 
                      'SIN RESPUESTA DEL CLIENTE',
                      'GENERAR CAMBIO DE FO',
                      'GENERAR ARREGLO DE INSTALACIÓN',
                      'GESTIONAR ARREGLO DE INSTALACIÓN'
                  )
                  AND solucion_tecnico NOT LIKE '%NOC%'
                  AND solucion_tecnico NOT LIKE '%PARCIAL%'
                  AND es_instalacion = %s
                ORDER BY hora_fin_visita ASC
            """
            cursor.execute(query, (fecha, int(es_instalacion)))
        visitas = cursor.fetchall()
        cursor.close()
        conexion.close()
        
        excel_buffer = generar_excel_calidad(visitas, fecha)
        filename = f"Reporte_Visitas_Efectivas_{fecha}.xlsx"
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


def format_tec_short(name):
    if not name or name.upper() in ["NO TECNICO", "SIN ASIGNAR", "NONE", "NAN", ""]:
        return "Sin asignar"
    parts = [p.strip() for p in name.split('/')]
    formatted_parts = []
    for p in parts:
        words = p.split()
        if len(words) >= 2:
            formatted_parts.append(f"{words[0].capitalize()} {words[1][0].upper()}.")
        elif len(words) == 1:
            formatted_parts.append(words[0].capitalize())
        else:
            formatted_parts.append(p)
    return " / ".join(formatted_parts)


def generar_excel_actividades(grouped, fecha_str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen de Actividades"
    
    ws.views.sheetView[0].showGridLines = True
    
    headers = ["TÉCNICO", "ACTIVIDAD", "CANTIDAD", "TOTAL"]
    ws.append(headers)
    
    # Peach fill (#FCE4D6)
    header_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_side = Side(style='thin', color='A0A0A0')
    border_style = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    for col_idx in range(1, 5):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border_style
        
    ws.row_dimensions[1].height = 28
    
    data_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    row_num = 2
    for tec, info in grouped.items():
        start_row = row_num
        activities = info['actividades']
        total = info['total']
        
        for act in activities:
            ws.append([
                tec,
                act['actividad'],
                act['cantidad'],
                total
            ])
            
            for col_idx in range(1, 5):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.border = border_style
                
                if col_idx in [1, 3, 4]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                    
            ws.row_dimensions[row_num].height = 22
            row_num += 1
            
        end_row = row_num - 1
        if start_row < end_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)
            
    ws.column_dimensions['A'].width = 25  # Técnico
    ws.column_dimensions['B'].width = 45  # Actividad
    ws.column_dimensions['C'].width = 12  # Cantidad
    ws.column_dimensions['D'].width = 12  # Total
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


@admin_bp.route('/api/admin/reporte_actividades/preview', methods=['GET'])
def preview_reporte_actividades():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'id_usuario': session['user_id'], 'role': session.get('user_role'), 'rol': session.get('user_role')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
        
    fecha = request.args.get('fecha', date.today().isoformat())
    es_instalacion = request.args.get('es_instalacion', '0').strip()
    
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    try:
        cursor = conexion.cursor(dictionary=True)
        query = """
            SELECT 
                tecnico_principal,
                tecnico_apoyo,
                solucion_tecnico,
                COUNT(*) as cantidad
            FROM visitas_tecnicas
            WHERE COALESCE(DATE(hora_fin_visita), fecha_programada) = %s AND estado = 'FINALIZADA'
              AND tecnico_principal IS NOT NULL 
              AND tecnico_principal NOT IN ('', 'NO TECNICO', 'SIN ASIGNAR', 'NONE', 'NAN')
              AND solucion_tecnico IS NOT NULL 
              AND solucion_tecnico NOT IN (
                  'NO SE PUEDE REALIZAR VISITA - SATURACIÓN DEL DÍA', 
                  'SIN RESPUESTA DEL CLIENTE'
              )
              AND es_instalacion = %s
            GROUP BY tecnico_principal, tecnico_apoyo, solucion_tecnico
            ORDER BY tecnico_principal, tecnico_apoyo, cantidad DESC
        """
        cursor.execute(query, (fecha, int(es_instalacion)))
        rows = cursor.fetchall()
        cursor.close()
        conexion.close()
        
        from collections import OrderedDict
        grouped = OrderedDict()
        
        for r in rows:
            tec_p = r['tecnico_principal']
            tec_a = r['tecnico_apoyo']
            actividad = r['solucion_tecnico']
            cantidad = r['cantidad']
            
            disp_name = format_tec_short(tec_p)
            if tec_a and tec_a.upper() not in ["NO TECNICO", "SIN ASIGNAR", "NONE", "NAN", ""]:
                disp_name += " / " + format_tec_short(tec_a)
                
            if disp_name == "Sin asignar":
                continue
                
            if disp_name not in grouped:
                grouped[disp_name] = {
                    "tecnico": disp_name,
                    "actividades": [],
                    "total": 0
                }
                
            grouped[disp_name]["actividades"].append({
                "actividad": str(actividad or '').upper(),
                "cantidad": cantidad
            })
            grouped[disp_name]["total"] += cantidad
            
        return jsonify({"status": "ok", "reporte": list(grouped.values())})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@admin_bp.route('/api/admin/reporte_actividades/excel', methods=['GET'])
def download_excel_reporte_actividades():
    if 'user_id' not in session:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
        
    fecha = request.args.get('fecha', date.today().isoformat())
    es_instalacion = request.args.get('es_instalacion', '0').strip()
    
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    try:
        cursor = conexion.cursor(dictionary=True)
        query = """
            SELECT 
                tecnico_principal,
                tecnico_apoyo,
                solucion_tecnico,
                COUNT(*) as cantidad
            FROM visitas_tecnicas
            WHERE COALESCE(DATE(hora_fin_visita), fecha_programada) = %s AND estado = 'FINALIZADA'
              AND tecnico_principal IS NOT NULL 
              AND tecnico_principal NOT IN ('', 'NO TECNICO', 'SIN ASIGNAR', 'NONE', 'NAN')
              AND solucion_tecnico IS NOT NULL 
              AND solucion_tecnico NOT IN (
                  'NO SE PUEDE REALIZAR VISITA - SATURACIÓN DEL DÍA', 
                  'SIN RESPUESTA DEL CLIENTE'
              )
              AND es_instalacion = %s
            GROUP BY tecnico_principal, tecnico_apoyo, solucion_tecnico
            ORDER BY tecnico_principal, tecnico_apoyo, cantidad DESC
        """
        cursor.execute(query, (fecha, int(es_instalacion)))
        rows = cursor.fetchall()
        cursor.close()
        conexion.close()
        
        from collections import OrderedDict
        grouped = OrderedDict()
        
        for r in rows:
            tec_p = r['tecnico_principal']
            tec_a = r['tecnico_apoyo']
            actividad = r['solucion_tecnico']
            cantidad = r['cantidad']
            
            disp_name = format_tec_short(tec_p)
            if tec_a and tec_a.upper() not in ["NO TECNICO", "SIN ASIGNAR", "NONE", "NAN", ""]:
                disp_name += " / " + format_tec_short(tec_a)
                
            if disp_name == "Sin asignar":
                continue
                
            if disp_name not in grouped:
                grouped[disp_name] = {
                    "actividades": [],
                    "total": 0
                }
                
            grouped[disp_name]["actividades"].append({
                "actividad": str(actividad or '').upper(),
                "cantidad": cantidad
            })
            grouped[disp_name]["total"] += cantidad
            
        excel_buffer = generar_excel_actividades(grouped, fecha)
        filename = f"Reporte_Actividades_Tecnicos_{fecha}.xlsx"
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


def format_datetime_dia_siguiente(dt_val):
    if not dt_val:
        return ""
    if isinstance(dt_val, (datetime, date)):
        dt = dt_val
    else:
        try:
            dt = datetime.fromisoformat(str(dt_val))
        except:
            return str(dt_val)
            
    d = dt.day
    m = dt.month
    y = dt.year
    if isinstance(dt, datetime):
        h = dt.hour
        mi = dt.minute
        s = dt.second
        return f"{d}/{m}/{y} {h:02d}:{mi:02d}:{s:02d}"
    else:
        return f"{d}/{m}/{y}"


def generar_excel_dia_siguiente(visitas, grupo_reagendadas_len, grupo_hoy_len, fecha_str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visitas Siguiente Día"
    
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "HORA DE ASIGNACIÓN",
        "NOMBRE",
        "SECTOR",
        "PROBLEMA",
        "OBSERVACIÓN"
    ]
    
    ws.append(headers)
    
    header_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_side = Side(style='thin', color='A0A0A0')
    border_style = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    for col_idx in range(1, 6):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border_style
        
    ws.row_dimensions[1].height = 28
    
    data_font = Font(name="Calibri", size=10)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    row_num = 2
    for v in visitas:
        f_reg = format_datetime_dia_siguiente(v.get('fecha_registro'))
        row_data = [
            f_reg,
            str(v.get('cliente') or '').upper(),
            str(v.get('sector') or '').upper(),
            str(v.get('problema') or '').upper(),
            ""
        ]
        ws.append(row_data)
        
        for col_idx in range(1, 6):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = data_font
            cell.border = border_style
            if col_idx in [1, 5]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
        
        ws.row_dimensions[row_num].height = 22
        row_num += 1
        
    if grupo_reagendadas_len > 0:
        start_row = 2
        end_row = 2 + grupo_reagendadas_len - 1
        ws.merge_cells(start_row=start_row, start_column=5, end_row=end_row, end_column=5)
        top_cell = ws.cell(row=start_row, column=5)
        top_cell.value = "VISITAS REAGENDADAS Y COORDINADAS DÍAS ANTERIORES"
        top_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        top_cell.font = Font(name="Calibri", size=10, bold=True)
        
    if grupo_hoy_len > 0:
        start_row = 2 + grupo_reagendadas_len
        end_row = 2 + grupo_reagendadas_len + grupo_hoy_len - 1
        ws.merge_cells(start_row=start_row, start_column=5, end_row=end_row, end_column=5)
        top_cell = ws.cell(row=start_row, column=5)
        top_cell.value = "VISITAS GENERADAS EL DÍA DE HOY"
        top_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        top_cell.font = Font(name="Calibri", size=10, bold=True)

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    max_len = max(max_len, len(line))
        width = max(max_len + 3, 12)
        if col_letter == 'A':
            width = 22
        elif col_letter == 'B':
            width = min(width, 35)
        elif col_letter == 'C':
            width = min(width, 22)
        elif col_letter == 'D':
            width = min(width, 28)
        elif col_letter == 'E':
            width = 22
            
        ws.column_dimensions[col_letter].width = width

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


@admin_bp.route('/api/admin/reporte_dia_siguiente/preview', methods=['GET'])
def preview_reporte_dia_siguiente():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'id_usuario': session['user_id'], 'role': session.get('user_role'), 'rol': session.get('user_role')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
        
    fecha = request.args.get('fecha', date.today().isoformat())
    
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    try:
        fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date()
        target_date = fecha_dt + timedelta(days=1)
        
        cursor = conexion.cursor(dictionary=True)
        query = """
            SELECT 
                id_visita,
                fecha_registro,
                cliente,
                sector,
                problema,
                estado
            FROM visitas_tecnicas
            WHERE fecha_programada = %s AND (estado = 'PENDIENTE' OR estado IS NULL)
        """
        cursor.execute(query, (target_date,))
        rows = cursor.fetchall()
        
        grupo_reagendadas = []
        grupo_hoy = []
        
        for v in rows:
            f_reg = v.get('fecha_registro')
            is_hoy = False
            if f_reg:
                if isinstance(f_reg, datetime):
                    reg_date = f_reg.date()
                elif isinstance(f_reg, date):
                    reg_date = f_reg
                else:
                    try:
                        reg_date = datetime.strptime(str(f_reg)[:10], "%Y-%m-%d").date()
                    except:
                        reg_date = None
                
                if reg_date == fecha_dt:
                    is_hoy = True
            
            if is_hoy:
                v['grupo'] = 'HOY'
                grupo_hoy.append(v)
            else:
                v['grupo'] = 'REAGENDADAS'
                grupo_reagendadas.append(v)
                
        def get_reg_time(v):
            f_reg = v.get('fecha_registro')
            if isinstance(f_reg, datetime):
                return f_reg
            if isinstance(f_reg, date):
                return datetime.combine(f_reg, datetime.min.time())
            if f_reg:
                try:
                    return datetime.strptime(str(f_reg), "%Y-%m-%d %H:%M:%S")
                except:
                    pass
            return datetime.min

        grupo_reagendadas.sort(key=get_reg_time)
        grupo_hoy.sort(key=get_reg_time)
        
        visitas_final = grupo_reagendadas + grupo_hoy
        
        for v in visitas_final:
            if v['fecha_registro'] and isinstance(v['fecha_registro'], (datetime, date)):
                v['fecha_registro'] = v['fecha_registro'].isoformat()
                
        return jsonify({"status": "ok", "visitas": visitas_final})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/reporte_dia_siguiente/excel', methods=['GET'])
def download_excel_reporte_dia_siguiente():
    if 'user_id' not in session:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
        
    fecha = request.args.get('fecha', date.today().isoformat())
    
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    try:
        fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date()
        target_date = fecha_dt + timedelta(days=1)
        
        cursor = conexion.cursor(dictionary=True)
        query = """
            SELECT 
                fecha_registro,
                cliente,
                sector,
                problema,
                estado
            FROM visitas_tecnicas
            WHERE fecha_programada = %s AND (estado = 'PENDIENTE' OR estado IS NULL)
        """
        cursor.execute(query, (target_date,))
        rows = cursor.fetchall()
        cursor.close()
        conexion.close()
        
        grupo_reagendadas = []
        grupo_hoy = []
        
        for v in rows:
            f_reg = v.get('fecha_registro')
            is_hoy = False
            if f_reg:
                if isinstance(f_reg, datetime):
                    reg_date = f_reg.date()
                elif isinstance(f_reg, date):
                    reg_date = f_reg
                else:
                    try:
                        reg_date = datetime.strptime(str(f_reg)[:10], "%Y-%m-%d").date()
                    except:
                        reg_date = None
                
                if reg_date == fecha_dt:
                    is_hoy = True
            
            if is_hoy:
                v['grupo'] = 'HOY'
                grupo_hoy.append(v)
            else:
                v['grupo'] = 'REAGENDADAS'
                grupo_reagendadas.append(v)
                
        def get_reg_time(v):
            f_reg = v.get('fecha_registro')
            if isinstance(f_reg, datetime):
                return f_reg
            if isinstance(f_reg, date):
                return datetime.combine(f_reg, datetime.min.time())
            if f_reg:
                try:
                    return datetime.strptime(str(f_reg), "%Y-%m-%d %H:%M:%S")
                except:
                    pass
            return datetime.min

        grupo_reagendadas.sort(key=get_reg_time)
        grupo_hoy.sort(key=get_reg_time)
        
        visitas_final = grupo_reagendadas + grupo_hoy
        
        excel_buffer = generar_excel_dia_siguiente(
            visitas_final, 
            len(grupo_reagendadas), 
            len(grupo_hoy), 
            fecha
        )
        filename = f"Reporte_Visitas_Dia_Siguiente_{fecha}.xlsx"
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


def map_solucion(sol):
    if not sol:
        return "INSPECCIÓN / SOLUCIÓN PARCIAL"
    sol = sol.upper().strip()
    if "CAMBIO DE FIBRA" in sol or "CAMBIO DE FO" in sol:
        return "CAMBIO DE FIBRA REALIZADO"
    if "GESTIONAR ARREGLO" in sol or "COORDINA" in sol or "TICKET AL NOC" in sol:
        return "SE COORDINA CAMBIO DE UTP / FIBRA"
    if "CABLE RG6 / UTP" in sol or "CAMBIO DE CABLE" in sol or "RG6" in sol or "RJ45" in sol:
        return "CAMBIO DE CABLE UTP / RG6"
    if "CONECTORES" in sol or "CONECTOR" in sol:
        return "FISICO / CAMBIO DE CONECTORES APC-UPC O RG6"
    if "CAMBIO DE EQUIPO ONT" in sol or "CAMBIO DE ONU" in sol:
        return "FISICO / CAMBIO DE ONU EN MAL ESTADO"
    if "CONF. DE EQUIPOS" in sol or "CONFIGURACIÓN" in sol or "CONF. DE EQUIPO" in sol:
        return "LÓGICO / CONFIGURACIÓN DE EQUIPOS"
    if "INSPECCIÓN" in sol or "SOLUCIÓN PARCIAL" in sol:
        return "INSPECCIÓN / SOLUCIÓN PARCIAL"
    if "RADIO ENLACE" in sol or "DOM" in sol:
        return "RADIO ENLACE / DOMÓTICA"
    if "ADAPTADOR" in sol or "CONEXIÓN ELÉCTRICA" in sol:
        return "FISICO / CAMBIO DE ADAPTADOR DE CORRIENTE"
    if "ARREGLO DE INSTALACIÓN" in sol or "REUBICACION" in sol or "RETENCIÓN" in sol or "REVISION COMPLETA" in sol or "RETIRO DE EQUIPOS" in sol:
        return "ARREGLO DE INSTALACIÓN / REUBICACIÓN DE EQUIPOS / RETENCIÓN"
    if "EFECTIVA" in sol or "ROUTER" in sol:
        return "INSTALACIÓN EFECTIVA / CAMBIO DE ROUTER"
    return "INSPECCIÓN / SOLUCIÓN PARCIAL"


def map_problema(prob):
    if not prob:
        return "VERIFICAR INSTACION"
    prob = prob.upper().strip()
    if "CAMBIO DE FIBRA" in prob or "CAMBIO DE FO" in prob or "CAMBIOS DE FIBRA" in prob:
        return "CAMBIOS DE FIBRA A REALIZAR"
    if "VERIFICAR INSTACION" in prob or "VERIFICAR INSTALACIÓN" in prob or "VERIFICAR INSTALACION" in prob:
        return "VERIFICAR INSTACION"
    if "ALARMADO" in prob or "LOS" in prob:
        return "EQUIPOS ALARMADOS"
    if "REVISION DE ONT" in prob or "REVISIÓN DE ONT" in prob:
        return "REVISION DE ONT"
    if "LENTITUD" in prob:
        return "LENTITUD EN EL SERVICIO"
    if "REVISION DE SERVICIO" in prob or "COBERTURA" in prob:
        return "REVISION DE SERVICIO/COBERTURA"
    if "ACTUALIZACION" in prob or "ROUTER" in prob or "EQUIPOS" in prob or "COLOCACIÓN ROUTER" in prob or "CONF." in prob:
        return "ACTUALIZACIÓN DE EQUIPO / COLOCACIÓN ROUTER"
    if "VELOCIDAD" in prob:
        return "NO MARCA VELOCIDAD CONTRATADA"
    if "REUBICACION" in prob or "REUBICACIÓN" in prob:
        return "REUBICACION DE EQUIPOS"
    if "COBRADA" in prob or "MANIPULACIÓN" in prob or "MANIPULACION" in prob:
        return "VT COBRADA / MANIPULACION DEL CLI"
    if "STREAMING" in prob or "ZAPPING" in prob or "ACTIVAR" in prob:
        return "ACTIVAR STREAMING"
    if "CANALES" in prob or "BORROSOS" in prob:
        return "CANALES BORROSOS"
    if "POTENCIA" in prob or "GPON" in prob:
        return "POTENCIA DEGRADADA (GPON)"
    if "RETENCION" in prob or "RETENCIÓN" in prob:
        return "RETENCIÓN"
    return "VERIFICAR INSTACION"


@admin_bp.route('/api/admin/cuadro_mando/preview', methods=['GET'])
def preview_cuadro_mando():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'id_usuario': session['user_id'], 'role': session.get('user_role'), 'rol': session.get('user_role')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
        
    fecha = request.args.get('fecha', date.today().isoformat())
    
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    try:
        cursor = conexion.cursor(dictionary=True)
        
        # 1. Obtener la lista de agentes de callcenter activos
        cursor.execute("SELECT nombre FROM callcenter WHERE activo = 1 ORDER BY nombre ASC")
        agentes_list = [row['nombre'] for row in cursor.fetchall()]
        
        # 2. Intentar auto-detectar los 3 agentes más activos en esta fecha
        cursor.execute("""
            SELECT agente, COUNT(*) as c 
            FROM atenciones 
            WHERE fecha = %s AND agente IS NOT NULL AND agente != '' AND agente != 'Importado'
            GROUP BY agente 
            ORDER BY c DESC 
            LIMIT 3
        """, (fecha,))
        detected_rows = cursor.fetchall()
        detected_agentes = [r['agente'] for r in detected_rows]
        
        # Rellenar con valores por defecto si no hay suficientes agentes
        default_agents = ['CC. Luis Saenz', 'CC. Guissella Quezada', 'CC. Mateo Samaniego']
        for default in default_agents:
            if len(detected_agentes) >= 3:
                break
            if default not in detected_agentes:
                detected_agentes.append(default)
        
        # Asegurar longitud 3
        while len(detected_agentes) < 3:
            detected_agentes.append('Sin asignar')
            
        agente_a = request.args.get('agente_a', detected_agentes[0])
        agente_b = request.args.get('agente_b', detected_agentes[1])
        agente_c = request.args.get('agente_c', detected_agentes[2])
        
        agentes = [agente_a, agente_b, agente_c]
        
        # 3. Contar gestiones por agente y categoría
        atenciones_data = {
            'visitas_coordinadas': [0, 0, 0],
            'solventado_llamada': [0, 0, 0],
            'solventado_mensajes': [0, 0, 0],
            'solventado_oficina': [0, 0, 0],
            'otros': [0, 0, 0]
        }
        
        for i, ag in enumerate(agentes):
            if not ag or ag == 'Sin asignar':
                continue
            # Visitas Coordinadas
            cursor.execute("""
                SELECT COUNT(*) as total FROM atenciones 
                WHERE fecha = %s AND agente = %s AND accion IN ('VISITA TECNICA', 'VISITA TECNICA COBRADA')
            """, (fecha, ag))
            atenciones_data['visitas_coordinadas'][i] = cursor.fetchone()['total'] or 0
            
            # Solventado por Llamada
            cursor.execute("""
                SELECT COUNT(*) as total FROM atenciones 
                WHERE fecha = %s AND agente = %s AND accion = 'SOPORTE MEDIANTE LLAMADA'
            """, (fecha, ag))
            atenciones_data['solventado_llamada'][i] = cursor.fetchone()['total'] or 0
            
            # Solventado por Mensajes
            cursor.execute("""
                SELECT COUNT(*) as total FROM atenciones 
                WHERE fecha = %s AND agente = %s AND accion = 'SOPORTE MEDIANTE MENSAJES'
            """, (fecha, ag))
            atenciones_data['solventado_mensajes'][i] = cursor.fetchone()['total'] or 0
            
            # Solventado en Oficina
            cursor.execute("""
                SELECT COUNT(*) as total FROM atenciones 
                WHERE fecha = %s AND agente = %s AND medio_contacto = 'OFICINA'
                  AND (accion NOT IN ('VISITA TECNICA', 'VISITA TECNICA COBRADA') OR accion IS NULL)
            """, (fecha, ag))
            atenciones_data['solventado_oficina'][i] = cursor.fetchone()['total'] or 0
            
            # Info / Transferencia / Otros
            cursor.execute("""
                SELECT COUNT(*) as total FROM atenciones 
                WHERE fecha = %s AND agente = %s 
                  AND (accion NOT IN ('VISITA TECNICA', 'VISITA TECNICA COBRADA', 'SOPORTE MEDIANTE LLAMADA', 'SOPORTE MEDIANTE MENSAJES') OR accion IS NULL)
                  AND (medio_contacto != 'OFICINA' OR medio_contacto IS NULL)
            """, (fecha, ag))
            atenciones_data['otros'][i] = cursor.fetchone()['total'] or 0
            
        # 4. KPIs de Visitas Técnicas de Campo (derecha)
        cursor.execute("""
            SELECT COUNT(*) as total FROM visitas_tecnicas
            WHERE fecha_programada = %s AND DATE(fecha_registro) < %s AND (estado != 'CANCELADA' OR estado IS NULL)
        """, (fecha, fecha))
        kpi_pendientes_anteriores = cursor.fetchone()['total'] or 0
        
        cursor.execute("""
            SELECT COUNT(*) as total FROM visitas_tecnicas
            WHERE COALESCE(DATE(hora_fin_visita), fecha_programada) = %s AND estado = 'FINALIZADA'
              AND tecnico_principal IS NOT NULL 
              AND tecnico_principal NOT IN ('', 'NO TECNICO', 'SIN ASIGNAR', 'NONE', 'NAN')
              AND solucion_tecnico IS NOT NULL 
              AND solucion_tecnico NOT IN (
                  'NO SE PUEDE REALIZAR VISITA - SATURACIÓN DEL DÍA', 
                  'SIN RESPUESTA DEL CLIENTE'
              )
        """, (fecha,))
        kpi_atendidas_hoy = cursor.fetchone()['total'] or 0
        
        fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date()
        manana = (fecha_dt + timedelta(days=1)).isoformat()
        cursor.execute("""
            SELECT COUNT(*) as total FROM visitas_tecnicas
            WHERE fecha_programada = %s AND (estado = 'PENDIENTE' OR estado IS NULL)
        """, (manana,))
        kpi_pendientes_manana = cursor.fetchone()['total'] or 0
        
        kpi_generadas_hoy = max(0, kpi_atendidas_hoy + kpi_pendientes_manana - kpi_pendientes_anteriores)
        kpi_total_carga = kpi_pendientes_anteriores + kpi_generadas_hoy
        
        # 5. Listados de problemas / soluciones
        cursor.execute("""
            SELECT solucion_tecnico, COUNT(*) as cantidad
            FROM visitas_tecnicas
            WHERE COALESCE(DATE(hora_fin_visita), fecha_programada) = %s AND estado = 'FINALIZADA'
              AND tecnico_principal IS NOT NULL 
              AND tecnico_principal NOT IN ('', 'NO TECNICO', 'SIN ASIGNAR', 'NONE', 'NAN')
              AND solucion_tecnico IS NOT NULL 
              AND solucion_tecnico NOT IN (
                  'NO SE PUEDE REALIZAR VISITA - SATURACIÓN DEL DÍA', 
                  'SIN RESPUESTA DEL CLIENTE'
              )
            GROUP BY solucion_tecnico
        """, (fecha,))
        soluciones_rows = cursor.fetchall()
        
        soluciones_dict = {
            "CAMBIO DE FIBRA REALIZADO": 0,
            "SE COORDINA CAMBIO DE UTP / FIBRA": 0,
            "CAMBIO DE CABLE UTP / RG6": 0,
            "FISICO / CAMBIO DE CONECTORES APC-UPC O RG6": 0,
            "FISICO / CAMBIO DE ONU EN MAL ESTADO": 0,
            "LÓGICO / CONFIGURACIÓN DE EQUIPOS": 0,
            "INSPECCIÓN / SOLUCIÓN PARCIAL": 0,
            "RADIO ENLACE / DOMÓTICA": 0,
            "FISICO / CAMBIO DE ADAPTADOR DE CORRIENTE": 0,
            "ARREGLO DE INSTALACIÓN / REUBICACIÓN DE EQUIPOS / RETENCIÓN": 0,
            "INSTALACIÓN EFECTIVA / CAMBIO DE ROUTER": 0,
            "TICKET A TECNOLOGÍA, DAÑO RADIAL": 0,
            "TICKET A TECNOLOGÍA, DAÑO FTTH": 0,
            "TICKET A TECNOLOGÍA, DAÑO HFC": 0
        }
        
        for r in soluciones_rows:
            mapped = map_solucion(r['solucion_tecnico'])
            if mapped in soluciones_dict:
                soluciones_dict[mapped] += r['cantidad']
                
        cursor.execute("""
            SELECT problema, COUNT(*) as cantidad
            FROM visitas_tecnicas
            WHERE fecha_programada = %s AND estado NOT IN ('FINALIZADA', 'CANCELADA', 'SOLVENTADA_REMOTA')
              AND problema IS NOT NULL AND problema != ''
            GROUP BY problema
        """, (manana,))
        problemas_rows = cursor.fetchall()
        
        problemas_dict = {
            "CAMBIOS DE FIBRA A REALIZAR": 0,
            "VERIFICAR INSTACION": 0,
            "EQUIPOS ALARMADOS": 0,
            "REVISION DE ONT": 0,
            "LENTITUD EN EL SERVICIO": 0,
            "REVISION DE SERVICIO/COBERTURA": 0,
            "ACTUALIZACIÓN DE EQUIPO / COLOCACIÓN ROUTER": 0,
            "NO MARCA VELOCIDAD CONTRATADA": 0,
            "REUBICACION DE EQUIPOS": 0,
            "VT COBRADA / MANIPULACION DEL CLI": 0,
            "ACTIVAR STREAMING": 0,
            "CANALES BORROSOS": 0,
            "POTENCIA DEGRADADA (GPON)": 0,
            "RETENCIÓN": 0
        }
        
        for r in problemas_rows:
            mapped = map_problema(r['problema'])
            if mapped in problemas_dict:
                problemas_dict[mapped] += r['cantidad']
                
        return jsonify({
            "status": "ok",
            "fecha": fecha,
            "agentes_list": agentes_list,
            "agente_a": agente_a,
            "agente_b": agente_b,
            "agente_c": agente_c,
            "atenciones": atenciones_data,
            "kpis": {
                "pendientes_anteriores": kpi_pendientes_anteriores,
                "generadas_hoy": kpi_generadas_hoy,
                "total_carga": kpi_total_carga,
                "atendidas_hoy": kpi_atendidas_hoy,
                "pendientes_manana": kpi_pendientes_manana
            },
            "soluciones": soluciones_dict,
            "problemas": problemas_dict
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/cuadro_mando/share_link', methods=['GET'])
def get_cuadro_mando_share_link():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'id_usuario': session['user_id']}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
        
    fecha = request.args.get('fecha')
    if not fecha:
        return jsonify({"status": "error", "message": "Fecha requerida"}), 400
        
    import hashlib
    import os
    import glob
    import re
    from flask import current_app
    
    secret = current_app.secret_key or "fallback_secret_salt_futurity_2026"
    token = hashlib.sha256(f"{fecha}_{secret}".encode('utf-8')).hexdigest()[:16]
    
    # Determinar el dominio base del enlace público
    public_base = os.getenv('PUBLIC_BASE_URL', 'https://atlas.futurity.com.ec').strip()
    if not public_base.endswith("/"):
        public_base += "/"
        
    public_url = f"{public_base}publico/cuadro_mando/{fecha}/{token}"
        
    return jsonify({
        "status": "ok",
        "url": public_url
    })


@admin_bp.route('/api/admin/cuadro_mando/excel', methods=['GET', 'POST'])
def download_excel_cuadro_mando():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'id_usuario': session['user_id']}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
        
    if request.method == 'POST':
        data = request.get_json() or {}
        fecha = data.get('fecha', date.today().isoformat())
        agente_a = data.get('agente_a', 'CC. Luis Saenz')
        agente_b = data.get('agente_b', 'CC. Guissella Quezada')
        agente_c = data.get('agente_c', 'CC. Mateo Samaniego')
        horario_a = data.get('horario_a', '7 AM - 4 PM')
        horario_b = data.get('horario_b', '2 PM - 9 PM')
        horario_c = data.get('horario_c', '10 AM - 8 PM')
    else:
        fecha = request.args.get('fecha', date.today().isoformat())
        agente_a = request.args.get('agente_a', 'CC. Luis Saenz')
        agente_b = request.args.get('agente_b', 'CC. Guissella Quezada')
        agente_c = request.args.get('agente_c', 'CC. Mateo Samaniego')
        horario_a = request.args.get('horario_a', '7 AM - 4 PM')
        horario_b = request.args.get('horario_b', '2 PM - 9 PM')
        horario_c = request.args.get('horario_c', '10 AM - 8 PM')
    
    try:
        soporte_a = int(data.get('soporte_a', 0))
    except:
        soporte_a = 0
        
    try:
        soporte_b = int(data.get('soporte_b', 0))
    except:
        soporte_b = 0
        
    try:
        soporte_c = int(data.get('soporte_c', 0))
    except:
        soporte_c = 0
        
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    try:
        cursor = conexion.cursor(dictionary=True)
        
        # 1. Contar gestiones por agente y categoría
        agentes = [agente_a, agente_b, agente_c]
        atenciones_data = {
            'visitas_coordinadas': [0, 0, 0],
            'solventado_llamada': [0, 0, 0],
            'solventado_mensajes': [0, 0, 0],
            'solventado_oficina': [0, 0, 0],
            'otros': [0, 0, 0]
        }
        
        for i, ag in enumerate(agentes):
            if not ag or ag == 'Sin asignar':
                continue
            # Visitas Coordinadas
            cursor.execute("""
                SELECT COUNT(*) as total FROM atenciones 
                WHERE fecha = %s AND agente = %s AND accion IN ('VISITA TECNICA', 'VISITA TECNICA COBRADA')
            """, (fecha, ag))
            atenciones_data['visitas_coordinadas'][i] = cursor.fetchone()['total'] or 0
            
            # Solventado por Llamada
            cursor.execute("""
                SELECT COUNT(*) as total FROM atenciones 
                WHERE fecha = %s AND agente = %s AND accion = 'SOPORTE MEDIANTE LLAMADA'
            """, (fecha, ag))
            atenciones_data['solventado_llamada'][i] = cursor.fetchone()['total'] or 0
            
            # Solventado por Mensajes
            cursor.execute("""
                SELECT COUNT(*) as total FROM atenciones 
                WHERE fecha = %s AND agente = %s AND accion = 'SOPORTE MEDIANTE MENSAJES'
            """, (fecha, ag))
            atenciones_data['solventado_mensajes'][i] = cursor.fetchone()['total'] or 0
            
            # Solventado en Oficina
            cursor.execute("""
                SELECT COUNT(*) as total FROM atenciones 
                WHERE fecha = %s AND agente = %s AND medio_contacto = 'OFICINA'
                  AND (accion NOT IN ('VISITA TECNICA', 'VISITA TECNICA COBRADA') OR accion IS NULL)
            """, (fecha, ag))
            atenciones_data['solventado_oficina'][i] = cursor.fetchone()['total'] or 0
            
            # Info / Transferencia / Otros
            cursor.execute("""
                SELECT COUNT(*) as total FROM atenciones 
                WHERE fecha = %s AND agente = %s 
                  AND (accion NOT IN ('VISITA TECNICA', 'VISITA TECNICA COBRADA', 'SOPORTE MEDIANTE LLAMADA', 'SOPORTE MEDIANTE MENSAJES') OR accion IS NULL)
                  AND (medio_contacto != 'OFICINA' OR medio_contacto IS NULL)
            """, (fecha, ag))
            atenciones_data['otros'][i] = cursor.fetchone()['total'] or 0
            
        # 2. KPIs de Visitas Técnicas de Campo
        cursor.execute("""
            SELECT COUNT(*) as total FROM visitas_tecnicas
            WHERE fecha_programada = %s AND DATE(fecha_registro) < %s AND (estado != 'CANCELADA' OR estado IS NULL)
        """, (fecha, fecha))
        kpi_pendientes_anteriores = cursor.fetchone()['total'] or 0
        
        cursor.execute("""
            SELECT COUNT(*) as total FROM visitas_tecnicas
            WHERE COALESCE(DATE(hora_fin_visita), fecha_programada) = %s AND estado = 'FINALIZADA'
              AND tecnico_principal IS NOT NULL 
              AND tecnico_principal NOT IN ('', 'NO TECNICO', 'SIN ASIGNAR', 'NONE', 'NAN')
              AND solucion_tecnico IS NOT NULL 
              AND solucion_tecnico NOT IN (
                  'NO SE PUEDE REALIZAR VISITA - SATURACIÓN DEL DÍA', 
                  'SIN RESPUESTA DEL CLIENTE'
              )
        """, (fecha,))
        kpi_atendidas_hoy = cursor.fetchone()['total'] or 0
        
        fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date()
        manana = (fecha_dt + timedelta(days=1)).isoformat()
        cursor.execute("""
            SELECT COUNT(*) as total FROM visitas_tecnicas
            WHERE fecha_programada = %s AND (estado = 'PENDIENTE' OR estado IS NULL)
        """, (manana,))
        kpi_pendientes_manana = cursor.fetchone()['total'] or 0
        
        kpi_generadas_hoy = max(0, kpi_atendidas_hoy + kpi_pendientes_manana - kpi_pendientes_anteriores)
        kpi_total_carga = kpi_pendientes_anteriores + kpi_generadas_hoy
        
        # Listados de problemas / soluciones
        cursor.execute("""
            SELECT solucion_tecnico, COUNT(*) as cantidad
            FROM visitas_tecnicas
            WHERE COALESCE(DATE(hora_fin_visita), fecha_programada) = %s AND estado = 'FINALIZADA'
              AND tecnico_principal IS NOT NULL 
              AND tecnico_principal NOT IN ('', 'NO TECNICO', 'SIN ASIGNAR', 'NONE', 'NAN')
              AND solucion_tecnico IS NOT NULL 
              AND solucion_tecnico NOT IN (
                  'NO SE PUEDE REALIZAR VISITA - SATURACIÓN DEL DÍA', 
                  'SIN RESPUESTA DEL CLIENTE'
              )
            GROUP BY solucion_tecnico
        """, (fecha,))
        soluciones_rows = cursor.fetchall()
        
        soluciones_dict = {
            "CAMBIO DE FIBRA REALIZADO": 0,
            "SE COORDINA CAMBIO DE UTP / FIBRA": 0,
            "CAMBIO DE CABLE UTP / RG6": 0,
            "FISICO / CAMBIO DE CONECTORES APC-UPC O RG6": 0,
            "FISICO / CAMBIO DE ONU EN MAL ESTADO": 0,
            "LÓGICO / CONFIGURACIÓN DE EQUIPOS": 0,
            "INSPECCIÓN / SOLUCIÓN PARCIAL": 0,
            "RADIO ENLACE / DOMÓTICA": 0,
            "FISICO / CAMBIO DE ADAPTADOR DE CORRIENTE": 0,
            "ARREGLO DE INSTALACIÓN / REUBICACIÓN DE EQUIPOS / RETENCIÓN": 0,
            "INSTALACIÓN EFECTIVA / CAMBIO DE ROUTER": 0,
            "TICKET A TECNOLOGÍA, DAÑO RADIAL": 0,
            "TICKET A TECNOLOGÍA, DAÑO FTTH": 0,
            "TICKET A TECNOLOGÍA, DAÑO HFC": 0
        }
        
        for r in soluciones_rows:
            mapped = map_solucion(r['solucion_tecnico'])
            if mapped in soluciones_dict:
                soluciones_dict[mapped] += r['cantidad']
                
        cursor.execute("""
            SELECT problema, COUNT(*) as cantidad
            FROM visitas_tecnicas
            WHERE fecha_programada = %s AND estado NOT IN ('FINALIZADA', 'CANCELADA', 'SOLVENTADA_REMOTA')
              AND problema IS NOT NULL AND problema != ''
            GROUP BY problema
        """, (manana,))
        problemas_rows = cursor.fetchall()
        
        problemas_dict = {
            "CAMBIOS DE FIBRA A REALIZAR": 0,
            "VERIFICAR INSTACION": 0,
            "EQUIPOS ALARMADOS": 0,
            "REVISION DE ONT": 0,
            "LENTITUD EN EL SERVICIO": 0,
            "REVISION DE SERVICIO/COBERTURA": 0,
            "ACTUALIZACIÓN DE EQUIPO / COLOCACIÓN ROUTER": 0,
            "NO MARCA VELOCIDAD CONTRATADA": 0,
            "REUBICACION DE EQUIPOS": 0,
            "VT COBRADA / MANIPULACION DEL CLI": 0,
            "ACTIVAR STREAMING": 0,
            "CANALES BORROSOS": 0,
            "POTENCIA DEGRADADA (GPON)": 0,
            "RETENCIÓN": 0
        }
        
        for r in problemas_rows:
            mapped = map_problema(r['problema'])
            if mapped in problemas_dict:
                problemas_dict[mapped] += r['cantidad']
                
        # 3. GENERAR EL EXCEL CON OPENPYXL
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cuadro de Mando Diario"
        ws.views.sheetView[0].showGridLines = True
        
        # Anchos de columna
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 26
        ws.column_dimensions['F'].width = 28
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 28
        ws.column_dimensions['I'].width = 12
        
        # Fills y Fonts
        fill_hdr_dark = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        fill_lbl_purple = PatternFill(start_color="8064A2", end_color="8064A2", fill_type="solid")
        fill_agent_a = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        fill_agent_b = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_agent_c = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        fill_total_cc = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        fill_total_row = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        
        fill_vis_attended = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        fill_vis_pending = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        fill_sol_hdr = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        fill_prob_hdr = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        font_white_bold = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        font_black_bold = Font(name="Calibri", size=10, bold=True, color="000000")
        font_regular = Font(name="Calibri", size=10, color="000000")
        font_large_bold = Font(name="Calibri", size=16, bold=True, color="000000")
        font_title_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        thin_side = Side(style='thin', color='A0A0A0')
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Fila 1: Títulos generales
        ws.merge_cells("B1:E1")
        ws.cell(row=1, column=2, value="ATENCIONES DIARIAS POR CALL CENTER")
        ws.merge_cells("F1:I1")
        ws.cell(row=1, column=6, value="VISITAS TÉCNICAS REALIZADAS POR EL TÉCNICO OPERADOR DE CAMPO")
        
        for c in range(2, 10):
            cell = ws.cell(row=1, column=c)
            cell.fill = fill_hdr_dark
            cell.font = font_title_white
            cell.alignment = align_center
            cell.border = border_all
        ws.row_dimensions[1].height = 28
        
        # Fila 2: Sub-títulos / Horarios
        ws.cell(row=2, column=1, value="HORARIO").fill = fill_lbl_purple
        ws.cell(row=2, column=1).font = font_white_bold
        ws.cell(row=2, column=1).alignment = align_center
        ws.cell(row=2, column=1).border = border_all
        
        ws.cell(row=2, column=2, value=f"HORARIO A\n{horario_a}").fill = fill_agent_a
        ws.cell(row=2, column=3, value=f"HORARIO B\n{horario_b}").fill = fill_agent_b
        ws.cell(row=2, column=4, value=f"HORARIO C\n{horario_c}").fill = fill_agent_c
        
        ws.cell(row=2, column=5, value="Gestión total de call center").fill = fill_total_cc
        
        ws.cell(row=2, column=6, value="VISITAS PENDIENTES DE AYER Y DE DIAS ANTERIORES")
        ws.cell(row=2, column=7, value="VISITAS GENERADAS HOY")
        ws.merge_cells("H2:I2")
        ws.cell(row=2, column=8, value="TOTAL CARGA DE VISITAS")
        
        for c in range(2, 6):
            cell = ws.cell(row=2, column=c)
            cell.font = font_black_bold
            cell.alignment = align_center
            cell.border = border_all
            
        for c in range(6, 10):
            cell = ws.cell(row=2, column=c)
            cell.font = Font(name="Calibri", size=9, bold=True, color="000000")
            cell.alignment = align_center
            cell.border = border_all
        ws.row_dimensions[2].height = 32
        
        # Fila 3: Agentes de turno / KPI valores
        ws.cell(row=3, column=1, value="AGENTE DE TURNO ESTE DÍA").fill = fill_lbl_purple
        ws.cell(row=3, column=1).font = font_white_bold
        ws.cell(row=3, column=1).alignment = align_center
        ws.cell(row=3, column=1).border = border_all
        
        ws.cell(row=3, column=2, value=agente_a).fill = fill_agent_a
        ws.cell(row=3, column=3, value=agente_b).fill = fill_agent_b
        ws.cell(row=3, column=4, value=agente_c).fill = fill_agent_c
        ws.cell(row=3, column=5, value="=SUM(B7:D7)+SUM(B8:D8)+SUM(B9:D9)+SUM(B10:D10)+SUM(B11:D11)+SUM(B12:D12)").fill = fill_total_cc
        
        ws.cell(row=3, column=6, value=kpi_pendientes_anteriores)
        ws.cell(row=3, column=7, value=kpi_generadas_hoy)
        ws.merge_cells("H3:I3")
        ws.cell(row=3, column=8, value="=F3+G3")
        
        for c in range(2, 5):
            cell = ws.cell(row=3, column=c)
            cell.font = font_black_bold
            cell.alignment = align_center
            cell.border = border_all
            
        ws.cell(row=3, column=5).font = font_large_bold
        ws.cell(row=3, column=5).alignment = align_center
        ws.cell(row=3, column=5).border = border_all
        
        for c in range(6, 10):
            cell = ws.cell(row=3, column=c)
            cell.font = font_large_bold
            cell.alignment = align_center
            cell.border = border_all
        ws.row_dimensions[3].height = 30
        
        # Fila 4: Cabecera Detalle de actividad
        ws.merge_cells("A4:E4")
        ws.cell(row=4, column=1, value="DETALLE DE ACTIVIDAD POR CALL CENTER").fill = fill_lbl_purple
        ws.cell(row=4, column=1).font = font_white_bold
        ws.cell(row=4, column=1).alignment = align_center
        ws.cell(row=4, column=1).border = border_all
        
        ws.merge_cells("F4:I4")
        ws.cell(row=4, column=6, value="DETALLE DE VISITAS POR DAÑOS").fill = fill_hdr_dark
        ws.cell(row=4, column=6).font = font_title_white
        ws.cell(row=4, column=6).alignment = align_center
        ws.cell(row=4, column=6).border = border_all
        ws.row_dimensions[4].height = 28
        
        # Fila 5: Agentes cabecera / Atendidos vs Pendientes mañana
        ws.cell(row=5, column=1, value="AGENTE").fill = fill_lbl_purple
        ws.cell(row=5, column=1).font = font_white_bold
        ws.cell(row=5, column=1).alignment = align_center
        ws.cell(row=5, column=1).border = border_all
        
        ws.cell(row=5, column=2, value=agente_a).fill = fill_agent_a
        ws.cell(row=5, column=3, value=agente_b).fill = fill_agent_b
        ws.cell(row=5, column=4, value=agente_c).fill = fill_agent_c
        ws.cell(row=5, column=5, value="TOTAL GESTIONES").fill = fill_hdr_dark
        
        for c in range(2, 5):
            cell = ws.cell(row=5, column=c)
            cell.font = font_black_bold
            cell.alignment = align_center
            cell.border = border_all
            
        ws.cell(row=5, column=5).font = font_white_bold
        ws.cell(row=5, column=5).alignment = align_center
        ws.cell(row=5, column=5).border = border_all
        
        ws.merge_cells("F5:G5")
        ws.cell(row=5, column=6, value="VISITAS ATENDIDAS Y SOLUCIONADAS +\nCAMBIOS DE FO").fill = fill_vis_attended
        ws.cell(row=5, column=6).font = Font(name="Calibri", size=9, bold=True, color="000000")
        ws.cell(row=5, column=6).alignment = align_center
        ws.cell(row=5, column=6).border = border_all
        
        ws.merge_cells("H5:I5")
        ws.cell(row=5, column=8, value="VISITAS PENDIENTES MAÑANA").fill = fill_vis_pending
        ws.cell(row=5, column=8).font = Font(name="Calibri", size=9, bold=True, color="000000")
        ws.cell(row=5, column=8).alignment = align_center
        ws.cell(row=5, column=8).border = border_all
        ws.row_dimensions[5].height = 32
        
        # Fila 6: TOTAL DE GESTIONES AL DÍA / Conteo de visitas de hoy y mañana
        ws.cell(row=6, column=1, value="TOTAL DE GESTIONES AL DÍA").fill = fill_total_row
        ws.cell(row=6, column=1).font = font_white_bold
        ws.cell(row=6, column=1).alignment = align_left
        ws.cell(row=6, column=1).border = border_all
        
        ws.cell(row=6, column=2, value="=SUM(B7:B12)").fill = fill_total_row
        ws.cell(row=6, column=3, value="=SUM(C7:C12)").fill = fill_total_row
        ws.cell(row=6, column=4, value="=SUM(D7:D12)").fill = fill_total_row
        ws.cell(row=6, column=5, value="=SUM(E7:E12)").fill = fill_total_row
        
        for c in range(2, 6):
            cell = ws.cell(row=6, column=c)
            cell.font = font_white_bold
            cell.alignment = align_center
            cell.border = border_all
            
        ws.merge_cells("F6:G6")
        ws.cell(row=6, column=6, value=kpi_atendidas_hoy)
        ws.cell(row=6, column=6).font = font_large_bold
        ws.cell(row=6, column=6).alignment = align_center
        ws.cell(row=6, column=6).border = border_all
        
        ws.merge_cells("H6:I6")
        ws.cell(row=6, column=8, value=kpi_pendientes_manana)
        ws.cell(row=6, column=8).font = font_large_bold
        ws.cell(row=6, column=8).alignment = align_center
        ws.cell(row=6, column=8).border = border_all
        ws.row_dimensions[6].height = 28
        
        # Filas 7 a 12: Categorías de atenciones / Subtítulos de daños
        categorias_rows = [
            ("VISITAS COORDINADAS", atenciones_data['visitas_coordinadas']),
            ("SOLVENTADO POR LLAMADA", atenciones_data['solventado_llamada']),
            ("SOLVENTADO POR MENSAJES", atenciones_data['solventado_mensajes']),
            ("SOLVENTADO EN OFICINA", atenciones_data['solventado_oficina']),
            ("SOPORTE A TÉCNICOS VT / INST", [soporte_a, soporte_b, soporte_c]),
            ("INFO / TRANSFERENCIAS - OTROS", atenciones_data['otros'])
        ]
        
        ws.merge_cells("F7:G7")
        ws.cell(row=7, column=6, value="PROBLEMA / SOLUCION DE VISITAS DE HOY").fill = fill_sol_hdr
        ws.cell(row=7, column=6).font = Font(name="Calibri", size=9, bold=True, color="000000")
        ws.cell(row=7, column=6).alignment = align_center
        ws.cell(row=7, column=6).border = border_all
        
        ws.merge_cells("H7:I7")
        ws.cell(row=7, column=8, value="PROBLEMAS DE VISITAS PARA MAÑANA").fill = fill_prob_hdr
        ws.cell(row=7, column=8).font = Font(name="Calibri", size=9, bold=True, color="000000")
        ws.cell(row=7, column=8).alignment = align_center
        ws.cell(row=7, column=8).border = border_all
        
        for idx, (cat_name, vals) in enumerate(categorias_rows, start=7):
            ws.cell(row=idx, column=1, value=cat_name).fill = fill_lbl_purple
            ws.cell(row=idx, column=1).font = font_white_bold
            ws.cell(row=idx, column=1).alignment = align_left
            ws.cell(row=idx, column=1).border = border_all
            
            ws.cell(row=idx, column=2, value=vals[0]).fill = fill_agent_a
            ws.cell(row=idx, column=3, value=vals[1]).fill = fill_agent_b
            ws.cell(row=idx, column=4, value=vals[2]).fill = fill_agent_c
            ws.cell(row=idx, column=5, value=f"=SUM(B{idx}:D{idx})").fill = fill_total_cc
            
            for c in range(2, 6):
                cell = ws.cell(row=idx, column=c)
                cell.font = font_black_bold
                cell.alignment = align_center
                cell.border = border_all
            ws.row_dimensions[idx].height = 24
            
        # Filas 8 a 21: Lista de problemas y soluciones detallados
        soluciones_lista = [
            "CAMBIO DE FIBRA REALIZADO",
            "SE COORDINA CAMBIO DE UTP / FIBRA",
            "CAMBIO DE CABLE UTP / RG6",
            "FISICO / CAMBIO DE CONECTORES APC-UPC O RG6",
            "FISICO / CAMBIO DE ONU EN MAL ESTADO",
            "LÓGICO / CONFIGURACIÓN DE EQUIPOS",
            "INSPECCIÓN / SOLUCIÓN PARCIAL",
            "RADIO ENLACE / DOMÓTICA",
            "FISICO / CAMBIO DE ADAPTADOR DE CORRIENTE",
            "ARREGLO DE INSTALACIÓN / REUBICACIÓN DE EQUIPOS / RETENCIÓN",
            "INSTALACIÓN EFECTIVA / CAMBIO DE ROUTER",
            "TICKET A TECNOLOGÍA, DAÑO RADIAL",
            "TICKET A TECNOLOGÍA, DAÑO FTTH",
            "TICKET A TECNOLOGÍA, DAÑO HFC"
        ]
        
        problemas_lista = [
            "CAMBIOS DE FIBRA A REALIZAR",
            "VERIFICAR INSTACION",
            "EQUIPOS ALARMADOS",
            "REVISION DE ONT",
            "LENTITUD EN EL SERVICIO",
            "REVISION DE SERVICIO/COBERTURA",
            "ACTUALIZACIÓN DE EQUIPO / COLOCACIÓN ROUTER",
            "NO MARCA VELOCIDAD CONTRATADA",
            "REUBICACION DE EQUIPOS",
            "VT COBRADA / MANIPULACION DEL CLI",
            "ACTIVAR STREAMING",
            "CANALES BORROSOS",
            "POTENCIA DEGRADADA (GPON)",
            "RETENCIÓN"
        ]
        
        for idx in range(14):
            row_idx = 8 + idx
            sol_name = soluciones_lista[idx]
            sol_val = soluciones_dict.get(sol_name, 0)
            
            prob_name = problemas_lista[idx]
            prob_dict_name = "VERIFICAR INSTACION" if prob_name == "VERIFICAR INSTACION" else prob_name
            prob_val = problemas_dict.get(prob_dict_name, 0)
            
            # Lado izquierdo (soluciones de hoy)
            ws.cell(row=row_idx, column=6, value=sol_name).font = Font(name="Calibri", size=8, color="000000")
            ws.cell(row=row_idx, column=6).alignment = align_left
            ws.cell(row=row_idx, column=6).border = border_all
            
            ws.cell(row=row_idx, column=7, value=sol_val).font = font_black_bold
            ws.cell(row=row_idx, column=7).alignment = align_center
            ws.cell(row=row_idx, column=7).border = border_all
            
            # Lado derecho (problemas de mañana)
            ws.cell(row=row_idx, column=8, value=prob_name).font = Font(name="Calibri", size=8, color="000000")
            ws.cell(row=row_idx, column=8).alignment = align_left
            ws.cell(row=row_idx, column=8).border = border_all
            
            ws.cell(row=row_idx, column=9, value=prob_val).font = font_black_bold
            ws.cell(row=row_idx, column=9).alignment = align_center
            ws.cell(row=row_idx, column=9).border = border_all
            
            ws.row_dimensions[row_idx].height = 22
            
        # Escribir tabla vertical auxiliar para el Gráfico 1 (Gestion por Agente) para compatibilidad con Google Sheets
        ws.cell(row=1, column=11, value="Agente")
        ws.cell(row=1, column=12, value="Gestión")
        ws.cell(row=2, column=11, value="=B5")
        ws.cell(row=2, column=12, value="=B6")
        ws.cell(row=3, column=11, value="=C5")
        ws.cell(row=3, column=12, value="=C6")
        ws.cell(row=4, column=11, value="=D5")
        ws.cell(row=4, column=12, value="=D6")

        # 4. AÑADIR LOS GRÁFICOS PIE NATIVOS EN EL EXCEL (Filas 23+)
        try:
            from openpyxl.chart import PieChart, Reference
            from openpyxl.chart.label import DataLabelList
            
            # Gráfico 1: GESTIÓN POR AGENTE (usando la tabla vertical de columnas K y L)
            chart_ag = PieChart()
            labels_ref_ag = Reference(ws, min_col=11, min_row=2, max_row=4)
            data_ref_ag = Reference(ws, min_col=12, min_row=1, max_row=4)
            chart_ag.add_data(data_ref_ag, titles_from_data=True)
            chart_ag.set_categories(labels_ref_ag)
            chart_ag.title = "GESTIÓN POR AGENTE"
            chart_ag.dataLabels = DataLabelList()
            chart_ag.dataLabels.showPercent = True
            chart_ag.dataLabels.showVal = True
            chart_ag.dataLabels.showCatName = False
            chart_ag.dataLabels.showSerName = False
            chart_ag.width = 11
            chart_ag.height = 7
            ws.add_chart(chart_ag, "A23")
            
            # Gráfico 2: TIPO DE ATENCIÓN
            chart_tp = PieChart()
            labels_ref_tp = Reference(ws, min_col=1, min_row=7, max_row=12)
            data_ref_tp = Reference(ws, min_col=5, min_row=7, max_row=12)
            chart_tp.add_data(data_ref_tp, titles_from_data=False)
            chart_tp.set_categories(labels_ref_tp)
            chart_tp.title = "TIPO DE ATENCIÓN"
            chart_tp.dataLabels = DataLabelList()
            chart_tp.dataLabels.showPercent = True
            chart_tp.dataLabels.showVal = True
            chart_tp.dataLabels.showCatName = False
            chart_tp.dataLabels.showSerName = False
            chart_tp.width = 11
            chart_tp.height = 7
            ws.add_chart(chart_tp, "D23")
        except Exception as chart_err:
            print(f"Error al generar gráficos nativos de openpyxl: {chart_err}")
            
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"Reporte_General_{fecha}.xlsx"
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@admin_bp.route('/api/admin/inventario', methods=['GET'])
def api_obtener_inventario():
    token = request.headers.get('Authorization')
    user = None
    role = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
        if user:
            role = user.get('role')
    elif 'user_id' in session:
        user = {'sub': session['user_id']}
        role = session.get('user_role')

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
    if role not in ['ADMIN', 'BODEGA', 'ASESOR']:
        return jsonify({"status": "error", "message": "No tienes privilegios para ver inventario."}), 403
        
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    try:
        cursor = conexion.cursor(dictionary=True)
        
        # 1. Obtener catálogo de materiales con stock en bodega
        cursor.execute("""
            SELECT id_material, codigo_material, nombre_material, unidad_medida, categoria, stock_bodega, stock_minimo, activo 
            FROM materiales 
            WHERE activo = 1 
            ORDER BY codigo_material ASC, nombre_material ASC
        """)
        materiales = cursor.fetchall()
        
        # 2. Obtener lista detallada de técnicos y sus placas de vehículos
        cursor.execute("SELECT id_tecnico, nombre, placa_vehiculo, placa_asignada_hoy FROM tecnicos WHERE activo = 1 ORDER BY nombre ASC")
        tecnicos_vehiculos = cursor.fetchall()
        
        # Obtener lista única de placas para la vista clásica de tabla de custodias
        placas_set = set()
        for tv in tecnicos_vehiculos:
            if tv.get('placa_vehiculo'):
                placas_set.add(tv['placa_vehiculo'])
            if tv.get('placa_asignada_hoy'):
                placas_set.add(tv['placa_asignada_hoy'])
        placas = sorted(list(placas_set))
        
        # 3. Obtener stock disponible en custodia por placa
        cursor.execute("SELECT placa_vehiculo, id_material, cantidad_disponible FROM inventario_tecnicos")
        custodia_raw = cursor.fetchall()
        
        # 4. Obtener consumo histórico (usado) por placa
        cursor.execute("""
            SELECT t.placa_vehiculo, vm.id_material, SUM(vm.cantidad_usada) as total_usado
            FROM visitas_materiales vm
            JOIN visitas_tecnicas vt ON vm.id_visita = vt.id_visita
            JOIN tecnicos t ON vt.tecnico_principal = t.nombre
            WHERE vt.estado = 'FINALIZADA' AND t.placa_vehiculo IS NOT NULL
            GROUP BY t.placa_vehiculo, vm.id_material
        """)
        consumo_raw = cursor.fetchall()
        
        # Estructurar la respuesta por placa y material
        inventario_tecnicos = {}
        for placa in placas:
            inventario_tecnicos[placa] = {}
            for mat in materiales:
                inventario_tecnicos[placa][str(mat['id_material'])] = {
                    "cantidad_disponible": 0,
                    "total_usado": 0
                }
                
        for row in custodia_raw:
            placa = row['placa_vehiculo']
            id_mat = str(row['id_material'])
            if placa in inventario_tecnicos and id_mat in inventario_tecnicos[placa]:
                inventario_tecnicos[placa][id_mat]['cantidad_disponible'] = row['cantidad_disponible']
                
        for row in consumo_raw:
            placa = row['placa_vehiculo']
            id_mat = str(row['id_material'])
            if placa in inventario_tecnicos and id_mat in inventario_tecnicos[placa]:
                inventario_tecnicos[placa][id_mat]['total_usado'] = int(row['total_usado'] or 0)
                
        return jsonify({
            "status": "ok",
            "materiales": materiales,
            "placas": placas,
            "tecnicos": placas,  # Retorna placas bajo la clave 'tecnicos' para compatibilidad con el JS
            "tecnicos_vehiculos": tecnicos_vehiculos,
            "inventario_tecnicos": inventario_tecnicos
        })
        
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/inventario/bodega/ingreso', methods=['POST'])
def api_bodega_ingreso():
    token = request.headers.get('Authorization')
    user = None
    role = None
    user_name = 'ADMIN'
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
        if user:
            role = user.get('role')
            user_name = user.get('nombre') or user.get('sub') or 'BODEGA'
    elif 'user_id' in session:
        user = {'sub': session['user_id']}
        role = session.get('user_role')
        user_name = session.get('user_name') or 'BODEGA'

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
    if role not in ['ADMIN', 'BODEGA']:
        return jsonify({"status": "error", "message": "No tienes privilegios para ingresar insumos a bodega."}), 403
        
    datos = request.get_json() or {}
    
    # Check if multiple items array is provided or single item
    items = datos.get('items')
    if not items or not isinstance(items, list):
        id_material = datos.get('id_material')
        cantidad = datos.get('cantidad')
        if id_material and cantidad and int(cantidad) > 0:
            items = [{'id_material': int(id_material), 'cantidad': int(cantidad)}]
        else:
            return jsonify({"status": "error", "message": "Debe proporcionar al menos un material válido y cantidad mayor a cero."}), 400

    # Filter valid items
    valid_items = []
    for it in items:
        m_id = it.get('id_material')
        cant = it.get('cantidad')
        if m_id and cant and int(cant) > 0:
            valid_items.append({
                'id_material': int(m_id),
                'cantidad': int(cant)
            })

    if not valid_items:
        return jsonify({"status": "error", "message": "No hay ítems válidos para procesar."}), 400

    fecha_ingreso_str = datos.get('fecha') or datos.get('fecha_ingreso') or date.today().isoformat()
    documento = datos.get('documento') or datos.get('factura') or None
    proveedor = datos.get('proveedor') or None
    comentario = datos.get('comentario') or datos.get('observacion') or None

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión con la base de datos"}), 500
        
    try:
        cursor = conexion.cursor(dictionary=True)
        total_items_count = len(valid_items)
        total_unidades_count = sum(it['cantidad'] for it in valid_items)
        
        # 1. Insert header in compras_ingresos_bodega
        cursor.execute("""
            INSERT INTO compras_ingresos_bodega 
            (fecha_ingreso, documento, proveedor, comentario, registrado_por, total_items, total_unidades)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (fecha_ingreso_str, documento, proveedor, comentario, user_name, total_items_count, total_unidades_count))
        id_ingreso = cursor.lastrowid

        # 2. Process each item
        for it in valid_items:
            m_id = it['id_material']
            cant = it['cantidad']
            
            # Get current stock
            cursor.execute("SELECT stock_bodega FROM materiales WHERE id_material = %s FOR UPDATE", (m_id,))
            mat_row = cursor.fetchone()
            stock_ant = mat_row['stock_bodega'] if mat_row and mat_row['stock_bodega'] is not None else 0
            stock_nuevo = stock_ant + cant

            # Update material stock
            cursor.execute("""
                UPDATE materiales 
                SET stock_bodega = stock_bodega + %s 
                WHERE id_material = %s
            """, (cant, m_id))

            # Insert detail item in compras_ingresos_items
            cursor.execute("""
                INSERT INTO compras_ingresos_items 
                (id_ingreso, id_material, cantidad, stock_anterior, stock_nuevo)
                VALUES (%s, %s, %s, %s, %s)
            """, (id_ingreso, m_id, cant, stock_ant, stock_nuevo))

        conexion.commit()
        return jsonify({
            "status": "ok", 
            "message": f"Se registraron exitosamente {total_items_count} producto(s) ({total_unidades_count} unidades en total) a Bodega Central.",
            "id_ingreso": id_ingreso
        })
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": f"Error al procesar ingreso: {str(e)}"}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/inventario/tecnico/entrega', methods=['POST'])
def api_tecnico_entrega():
    token = request.headers.get('Authorization')
    user = None
    role = None
    user_name = 'ADMIN'
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
        if user:
            role = user.get('role')
            user_name = user.get('nombre') or user.get('sub') or 'BODEGA'
    elif 'user_id' in session:
        user = {'sub': session['user_id']}
        role = session.get('user_role')
        user_name = session.get('user_name') or 'BODEGA'

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
    if role not in ['ADMIN', 'BODEGA']:
        return jsonify({"status": "error", "message": "No tienes privilegios para entregar insumos a técnicos."}), 403
        
    datos = request.get_json() or {}
    placa_vehiculo = datos.get('placa_vehiculo') or datos.get('tecnico_nombre')
    tecnico_nombre = datos.get('tecnico_responsable') or datos.get('tecnico_nombre') or placa_vehiculo

    if not placa_vehiculo:
        return jsonify({"status": "error", "message": "Debe especificar la placa del vehículo destino."}), 400

    # Multiple items vs single item
    items = datos.get('items')
    if not items or not isinstance(items, list):
        id_material = datos.get('id_material')
        cantidad = datos.get('cantidad')
        if id_material and cantidad and int(cantidad) > 0:
            items = [{'id_material': int(id_material), 'cantidad': int(cantidad)}]
        else:
            return jsonify({"status": "error", "message": "Debe especificar al menos un material válido con cantidad mayor a cero."}), 400

    valid_items = []
    for it in items:
        m_id = it.get('id_material')
        cant = it.get('cantidad')
        if m_id and cant and int(cant) > 0:
            valid_items.append({
                'id_material': int(m_id),
                'cantidad': int(cant)
            })

    if not valid_items:
        return jsonify({"status": "error", "message": "No hay ítems válidos para entregar."}), 400
        
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión con la base de datos"}), 500
        
    try:
        cursor = conexion.cursor(dictionary=True)
        
        # 1. Validar que haya stock suficiente para TODOS los ítems antes de proceder
        for it in valid_items:
            m_id = it['id_material']
            cant = it['cantidad']
            cursor.execute("SELECT nombre_material, stock_bodega, unidad_medida FROM materiales WHERE id_material = %s FOR UPDATE", (m_id,))
            mat_row = cursor.fetchone()
            if not mat_row:
                return jsonify({"status": "error", "message": f"El material ID {m_id} no existe en el catálogo."}), 400
            stock_disp = mat_row['stock_bodega'] or 0
            if stock_disp < cant:
                return jsonify({
                    "status": "error", 
                    "message": f"Stock insuficiente en Bodega Central para '{mat_row['nombre_material']}'. Disponible: {stock_disp} {mat_row['unidad_medida']}, Solicitado: {cant} {mat_row['unidad_medida']}."
                }), 400

        # 2. Registrar cabecera en requisiciones_entregas_placas
        documento_req = datos.get('documento') or None
        fecha_entrega_str = datos.get('fecha') or date.today().isoformat()
        comentario_req = datos.get('comentario') or None
        
        cursor.execute("""
            INSERT INTO requisiciones_entregas_placas
            (documento_req, placa_vehiculo, tecnico_responsable, fecha_entrega, comentario, registrado_por, total_items, total_unidades)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (documento_req, placa_vehiculo, tecnico_nombre, fecha_entrega_str, comentario_req, user_name, len(valid_items), sum(it['cantidad'] for it in valid_items)))
        id_requisicion = cursor.lastrowid

        # 3. Descontar de bodega, sumar al vehículo y registrar detalle
        total_unidades = 0
        for it in valid_items:
            m_id = it['id_material']
            cant = it['cantidad']
            total_unidades += cant

            # Stock actual antes del descuento
            cursor.execute("SELECT stock_bodega FROM materiales WHERE id_material = %s FOR UPDATE", (m_id,))
            mat_cur = cursor.fetchone()
            stock_disp = mat_cur['stock_bodega'] or 0

            # Descontar de bodega
            cursor.execute("""
                UPDATE materiales 
                SET stock_bodega = stock_bodega - %s 
                WHERE id_material = %s
            """, (cant, m_id))
            
            # Sumar a inventario del vehículo
            cursor.execute("""
                INSERT INTO inventario_tecnicos (placa_vehiculo, id_material, cantidad_disponible)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE cantidad_disponible = cantidad_disponible + VALUES(cantidad_disponible)
            """, (placa_vehiculo, m_id, cant))

            # Registrar detalle en requisiciones_entregas_items
            cursor.execute("""
                INSERT INTO requisiciones_entregas_items
                (id_requisicion, id_material, cantidad, stock_bodega_anterior, stock_bodega_nuevo)
                VALUES (%s, %s, %s, %s, %s)
            """, (id_requisicion, m_id, cant, stock_disp, stock_disp - cant))

            # Registrar en historial de traspasos
            cursor.execute("""
                INSERT INTO traspasos_tecnicos 
                (tecnico_origen, placa_origen, tecnico_destino, placa_destino, id_material, cantidad, agente_registro)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, ('BODEGA CENTRAL', 'BODEGA', tecnico_nombre, placa_vehiculo, m_id, cant, user_name))
        
        conexion.commit()
        return jsonify({
            "status": "ok", 
            "message": f"Se entregaron exitosamente {len(valid_items)} producto(s) ({total_unidades} unidades en total) a la placa {placa_vehiculo}.",
            "id_requisicion": id_requisicion
        })
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": f"Error al entregar materiales: {str(e)}"}), 500
    finally:
        cursor.close()
        conexion.close()


# ==========================================
# 🏢 ENDPOINTS DE PROVEEDORES
# ==========================================

@admin_bp.route('/api/admin/proveedores', methods=['GET'])
def api_get_proveedores():
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión"}), 500
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("SELECT * FROM proveedores WHERE activo = 1 ORDER BY nombre_empresa ASC")
        rows = cursor.fetchall()
        return jsonify({"status": "ok", "proveedores": rows})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/proveedores', methods=['POST'])
def api_create_proveedor():
    datos = request.get_json() or {}
    nombre_empresa = datos.get('nombre_empresa', '').strip()
    if not nombre_empresa:
        return jsonify({"status": "error", "message": "El nombre o razón social de la empresa es obligatorio."}), 400

    ruc = datos.get('ruc', '').strip() or None
    contacto_nombre = datos.get('contacto_nombre', '').strip() or None
    telefono = datos.get('telefono', '').strip() or None
    email = datos.get('email', '').strip() or None
    direccion = datos.get('direccion', '').strip() or None
    observaciones = datos.get('observaciones', '').strip() or None

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión"}), 500
    try:
        cursor = conexion.cursor()
        cursor.execute("""
            INSERT INTO proveedores 
            (ruc, nombre_empresa, contacto_nombre, telefono, email, direccion, observaciones, activo)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 1)
        """, (ruc, nombre_empresa, contacto_nombre, telefono, email, direccion, observaciones))
        conexion.commit()
        return jsonify({"status": "ok", "message": "Proveedor registrado exitosamente."})
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/proveedores/<int:id_proveedor>', methods=['PUT'])
def api_update_proveedor(id_proveedor):
    datos = request.get_json() or {}
    nombre_empresa = datos.get('nombre_empresa', '').strip()
    if not nombre_empresa:
        return jsonify({"status": "error", "message": "El nombre o razón social de la empresa es obligatorio."}), 400

    ruc = datos.get('ruc', '').strip() or None
    contacto_nombre = datos.get('contacto_nombre', '').strip() or None
    telefono = datos.get('telefono', '').strip() or None
    email = datos.get('email', '').strip() or None
    direccion = datos.get('direccion', '').strip() or None
    observaciones = datos.get('observaciones', '').strip() or None

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión"}), 500
    try:
        cursor = conexion.cursor()
        cursor.execute("""
            UPDATE proveedores 
            SET ruc = %s, nombre_empresa = %s, contacto_nombre = %s, telefono = %s, email = %s, direccion = %s, observaciones = %s
            WHERE id_proveedor = %s
        """, (ruc, nombre_empresa, contacto_nombre, telefono, email, direccion, observaciones, id_proveedor))
        conexion.commit()
        return jsonify({"status": "ok", "message": "Proveedor actualizado exitosamente."})
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/proveedores/<int:id_proveedor>', methods=['DELETE'])
def api_delete_proveedor(id_proveedor):
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión"}), 500
    try:
        cursor = conexion.cursor()
        cursor.execute("UPDATE proveedores SET activo = 0 WHERE id_proveedor = %s", (id_proveedor,))
        conexion.commit()
        return jsonify({"status": "ok", "message": "Proveedor eliminado exitosamente."})
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


# ==========================================
# 📦 HISTORIAL DE COMPRAS / INGRESOS BODEGA
# ==========================================

@admin_bp.route('/api/admin/inventario/compras', methods=['GET'])
def api_get_historial_compras():
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión"}), 500
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("""
            SELECT id_ingreso, fecha_ingreso, documento, proveedor, comentario, registrado_por, total_items, total_unidades, fecha_registro 
            FROM compras_ingresos_bodega 
            ORDER BY fecha_registro DESC 
            LIMIT 300
        """)
        compras = cursor.fetchall()
        return jsonify({"status": "ok", "compras": compras})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/inventario/compras/<int:id_ingreso>', methods=['GET'])
def api_get_detalle_compra(id_ingreso):
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión"}), 500
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("""
            SELECT ci.id_item, ci.id_material, m.codigo_material, m.nombre_material, m.unidad_medida, 
                   ci.cantidad, ci.stock_anterior, ci.stock_nuevo
            FROM compras_ingresos_items ci
            JOIN materiales m ON ci.id_material = m.id_material
            WHERE ci.id_ingreso = %s
            ORDER BY m.nombre_material ASC
        """, (id_ingreso,))
        items = cursor.fetchall()
        return jsonify({"status": "ok", "items": items})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


# ==========================================
# 🚚 HISTORIAL DE REQUISICIONES / ENTREGAS PLACAS
# ==========================================

@admin_bp.route('/api/admin/inventario/requisiciones', methods=['GET'])
def api_get_historial_requisiciones():
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión"}), 500
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("""
            SELECT id_requisicion, documento_req, placa_vehiculo, tecnico_responsable, fecha_entrega, comentario, registrado_por, total_items, total_unidades, fecha_registro
            FROM requisiciones_entregas_placas
            ORDER BY fecha_registro DESC
            LIMIT 300
        """)
        reqs = cursor.fetchall()
        return jsonify({"status": "ok", "requisiciones": reqs})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/inventario/requisiciones/<int:id_requisicion>', methods=['GET'])
def api_get_detalle_requisicion(id_requisicion):
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión"}), 500
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("""
            SELECT ri.id_item, ri.id_material, m.codigo_material, m.nombre_material, m.unidad_medida, 
                   ri.cantidad, ri.stock_bodega_anterior, ri.stock_bodega_nuevo
            FROM requisiciones_entregas_items ri
            JOIN materiales m ON ri.id_material = m.id_material
            WHERE ri.id_requisicion = %s
            ORDER BY m.nombre_material ASC
        """, (id_requisicion,))
        items = cursor.fetchall()
        return jsonify({"status": "ok", "items": items})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()



@admin_bp.route('/api/admin/inventario/tecnico/devolucion', methods=['POST'])
def api_tecnico_devolucion():
    token = request.headers.get('Authorization')
    user = None
    role = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
        if user:
            role = user.get('role')
    elif 'user_id' in session:
        user = {'sub': session['user_id']}
        role = session.get('user_role')

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
    if role not in ['ADMIN', 'BODEGA']:
        return jsonify({"status": "error", "message": "No tienes privilegios para registrar devoluciones."}), 403
        
    datos = request.get_json() or {}
    placa_vehiculo = datos.get('tecnico_nombre') or datos.get('placa_vehiculo')
    id_material = datos.get('id_material')
    cantidad = datos.get('cantidad')
    
    if not placa_vehiculo or not id_material or not cantidad or int(cantidad) <= 0:
        return jsonify({"status": "error", "message": "Parámetros inválidos"}), 400
        
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión"}), 500
        
    try:
        cursor = conexion.cursor()
        
        # 1. Validar que la placa tenga suficiente cantidad para devolver
        cursor.execute("""
            SELECT cantidad_disponible FROM inventario_tecnicos 
            WHERE placa_vehiculo = %s AND id_material = %s
        """, (placa_vehiculo, int(id_material)))
        row = cursor.fetchone()
        if not row or row[0] < int(cantidad):
            return jsonify({"status": "error", "message": "La placa no dispone de esa cantidad en custodia"}), 400
            
        # 2. Descontar a la placa
        cursor.execute("""
            UPDATE inventario_tecnicos 
            SET cantidad_disponible = cantidad_disponible - %s 
            WHERE placa_vehiculo = %s AND id_material = %s
        """, (int(cantidad), placa_vehiculo, int(id_material)))
        
        # 3. Sumar a bodega
        cursor.execute("""
            UPDATE materiales 
            SET stock_bodega = stock_bodega + %s 
            WHERE id_material = %s
        """, (int(cantidad), int(id_material)))
        
        conexion.commit()
        return jsonify({"status": "ok", "message": "Devolución registrada con éxito"})
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/materiales', methods=['POST'])
def api_crear_material():
    token = request.headers.get('Authorization')
    user = None
    role = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
        if user:
            role = user.get('role')
    elif 'user_id' in session:
        user = {'sub': session['user_id']}
        role = session.get('user_role')

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
    if role not in ['ADMIN', 'BODEGA']:
        return jsonify({"status": "error", "message": "No tienes privilegios para crear productos."}), 403

    datos = request.get_json() or {}
    codigo = (datos.get('codigo_material') or '').strip().upper()
    nombre = (datos.get('nombre_material') or '').strip().upper()
    unidad = (datos.get('unidad_medida') or 'UNIDADES').strip().upper()
    categoria = (datos.get('categoria') or 'GENERAL').strip().upper()
    stock_bodega = int(datos.get('stock_bodega') or 0)
    stock_minimo = int(datos.get('stock_minimo') or 0)

    if not codigo or not nombre:
        return jsonify({"status": "error", "message": "El código y el nombre del producto son requeridos."}), 400

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión"}), 500

    try:
        cursor = conexion.cursor()
        # Verificar duplicados por codigo
        cursor.execute("SELECT id_material FROM materiales WHERE codigo_material = %s", (codigo,))
        if cursor.fetchone():
            return jsonify({"status": "error", "message": f"Ya existe un producto con el código '{codigo}'."}), 400

        cursor.execute("""
            INSERT INTO materiales (codigo_material, nombre_material, unidad_medida, categoria, stock_bodega, stock_minimo, activo)
            VALUES (%s, %s, %s, %s, %s, %s, 1)
        """, (codigo, nombre, unidad, categoria, stock_bodega, stock_minimo))
        conexion.commit()
        return jsonify({"status": "ok", "message": "Producto creado con éxito", "id_material": cursor.lastrowid})
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/materiales/<int:id_material>', methods=['PUT'])
def api_actualizar_material(id_material):
    token = request.headers.get('Authorization')
    user = None
    role = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
        if user:
            role = user.get('role')
    elif 'user_id' in session:
        user = {'sub': session['user_id']}
        role = session.get('user_role')

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
    if role not in ['ADMIN', 'BODEGA']:
        return jsonify({"status": "error", "message": "No tienes privilegios para editar productos."}), 403

    datos = request.get_json() or {}
    codigo = (datos.get('codigo_material') or '').strip().upper()
    nombre = (datos.get('nombre_material') or '').strip().upper()
    unidad = (datos.get('unidad_medida') or 'UNIDADES').strip().upper()
    categoria = (datos.get('categoria') or 'GENERAL').strip().upper()
    stock_bodega = int(datos.get('stock_bodega') if datos.get('stock_bodega') is not None else 0)
    stock_minimo = int(datos.get('stock_minimo') if datos.get('stock_minimo') is not None else 0)

    if not codigo or not nombre:
        return jsonify({"status": "error", "message": "El código y nombre del producto no pueden estar vacíos."}), 400

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión"}), 500

    try:
        cursor = conexion.cursor()
        cursor.execute("""
            UPDATE materiales 
            SET codigo_material = %s, nombre_material = %s, unidad_medida = %s, categoria = %s, stock_bodega = %s, stock_minimo = %s
            WHERE id_material = %s
        """, (codigo, nombre, unidad, categoria, stock_bodega, stock_minimo, id_material))
        conexion.commit()
        return jsonify({"status": "ok", "message": "Producto actualizado con éxito"})
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/materiales/<int:id_material>', methods=['DELETE'])
def api_eliminar_material(id_material):
    token = request.headers.get('Authorization')
    user = None
    role = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
        if user:
            role = user.get('role')
    elif 'user_id' in session:
        user = {'sub': session['user_id']}
        role = session.get('user_role')

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
    if role not in ['ADMIN', 'BODEGA']:
        return jsonify({"status": "error", "message": "No tienes privilegios para desactivar productos."}), 403

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión"}), 500

    try:
        cursor = conexion.cursor()
        # Soft delete: set activo = 0
        cursor.execute("UPDATE materiales SET activo = 0 WHERE id_material = %s", (id_material,))
        conexion.commit()
        return jsonify({"status": "ok", "message": "Producto desactivado del catálogo con éxito"})
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()



@admin_bp.route('/api/admin/tecnicos/mas_cercano', methods=['GET'])
def obtener_tecnico_mas_cercano():
    if 'user_id' not in session or session.get('user_role') not in ['ADMIN', 'ASESOR', 'CALIDAD', 'ATC']:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
        
    lat_str = request.args.get('lat')
    lon_str = request.args.get('lon')
    
    if not lat_str or not lon_str:
        return jsonify({"status": "error", "message": "Parámetros lat y lon son requeridos"}), 400
        
    try:
        lat_visita = float(lat_str)
        lon_visita = float(lon_str)
    except ValueError:
        return jsonify({"status": "error", "message": "Coordenadas no válidas"}), 400
        
    active_area = session.get('active_area', 'SOPORTE')
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión"}), 500
        
    cursor = conexion.cursor(dictionary=True)
    try:
        # Consultamos técnicos activos, cuyo área de trabajo sea la activa de la sesión,
        # y que tengan coordenadas GPS reales (no nulas) registradas.
        cursor.execute("""
            SELECT id_tecnico, nombre, latitud_actual, longitud_actual, estado_actividad, placa_vehiculo
            FROM tecnicos
            WHERE activo = 1 
              AND area_trabajo = %s
              AND latitud_actual IS NOT NULL 
              AND longitud_actual IS NOT NULL
              AND ultima_conexion >= DATE_SUB(NOW(), INTERVAL 10 MINUTE)
        """, (active_area,))
        tecnicos = cursor.fetchall()
        
        import math
        recomendados = []
        for tec in tecnicos:
            lat_tec = float(tec['latitud_actual'])
            lon_tec = float(tec['longitud_actual'])
            
            # Haversine distance
            R = 6371.0
            phi1 = math.radians(lat_visita)
            phi2 = math.radians(lat_tec)
            delta_phi = math.radians(lat_tec - lat_visita)
            delta_lambda = math.radians(lon_tec - lon_visita)
            
            a = math.sin(delta_phi / 2.0)**2 + \
                math.cos(phi1) * math.cos(phi2) * \
                math.sin(delta_lambda / 2.0)**2
            c = 2.0 * math.atan2(math.sqrt(a), math.sqrt(1.0 - a))
            distancia = R * c
            
            recomendados.append({
                "id_tecnico": tec['id_tecnico'],
                "nombre": tec['nombre'],
                "estado": tec['estado_actividad'] or 'Disponible',
                "placa": tec['placa_vehiculo'] or 'S/P',
                "distancia_km": round(distancia, 2)
            })
            
        # Ordenar de menor a mayor distancia
        recomendados.sort(key=lambda x: x['distancia_km'])
        
        # Devolver el top 3
        return jsonify({"status": "ok", "recomendados": recomendados[:3]})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/metricas_tiempos', methods=['GET'])
def metricas_tiempos():
    if 'user_id' not in session:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
    if session.get('user_role') not in ['ADMIN', 'ASESOR', 'CALIDAD', 'ATC']:
        return jsonify({"status": "error", "message": "No tienes privilegios para ver métricas de tiempos."}), 403

    active_area = session.get('active_area', 'SOPORTE')
    es_instalacion_val = 1 if active_area == 'INSTALACIONES' else 0

    # Obtener parámetros de filtros (hoy y hace 3 meses por defecto si no se especifican)
    hoy_dt = date.today()
    hace_tres_meses = (hoy_dt - timedelta(days=90)).isoformat()
    hoy_str = hoy_dt.isoformat()

    fecha_inicio = request.args.get('fecha_inicio', hace_tres_meses)
    if not fecha_inicio:
        fecha_inicio = hace_tres_meses
    fecha_fin = request.args.get('fecha_fin', hoy_str)
    if not fecha_fin:
        fecha_fin = hoy_str
    tecnico_filtro = request.args.get('tecnico', '').strip()

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "No se pudo conectar a la base de datos"}), 500
    
    cursor = conexion.cursor(dictionary=True)
    try:
        # Cláusula WHERE común
        where_clause = "WHERE fecha_programada >= %s AND fecha_programada <= %s AND es_instalacion = %s"
        params = [fecha_inicio, fecha_fin, es_instalacion_val]

        if tecnico_filtro and tecnico_filtro != 'TODOS':
            where_clause += " AND (tecnico_principal = %s OR tecnico_apoyo = %s)"
            params.extend([tecnico_filtro, tecnico_filtro])

        # 1. KPIs Generales de Tiempos
        query_kpis = f"""
            SELECT 
                COUNT(*) as total_asignadas,
                SUM(CASE WHEN estado IN ('FINALIZADA', 'SOLVENTADA_REMOTA') 
                          AND (solucion_tecnico IS NULL OR (
                              solucion_tecnico NOT LIKE '%GESTIONAR ARREGLO%'
                              AND solucion_tecnico NOT LIKE '%SOLUCIÓN PARCIAL%'
                              AND solucion_tecnico NOT LIKE '%SOLUCION PARCIAL%'
                              AND solucion_tecnico NOT LIKE '%GENERAR CAMBIO DE FO%'
                              AND solucion_tecnico NOT LIKE '%SIN RESPUESTA%'
                              AND solucion_tecnico NOT LIKE '%NO SE PUEDE REALIZAR VISITA%'
                              AND solucion_tecnico NOT LIKE '%NOC%'
                          )) THEN 1 ELSE 0 END) as total_finalizadas,
                AVG(CASE WHEN estado = 'FINALIZADA' AND hora_en_ruta IS NOT NULL AND hora_inicio_visita IS NOT NULL 
                         THEN TIMESTAMPDIFF(MINUTE, hora_en_ruta, hora_inicio_visita) ELSE NULL END) as avg_traslado,
                AVG(CASE WHEN estado = 'FINALIZADA' AND hora_inicio_visita IS NOT NULL AND hora_fin_visita IS NOT NULL 
                         THEN TIMESTAMPDIFF(MINUTE, hora_inicio_visita, hora_fin_visita) ELSE NULL END) as avg_resolucion,
                AVG(CASE WHEN estado = 'FINALIZADA' AND hora_en_ruta IS NOT NULL AND hora_fin_visita IS NOT NULL 
                         THEN TIMESTAMPDIFF(MINUTE, hora_en_ruta, hora_fin_visita) ELSE NULL END) as avg_total
            FROM visitas_tecnicas
            {where_clause}
        """
        cursor.execute(query_kpis, params)
        kpis_row = cursor.fetchone()
        
        total_asignadas = kpis_row['total_asignadas'] or 0
        total_finalizadas = kpis_row['total_finalizadas'] or 0
        tasa_efectividad = round(float(total_finalizadas / total_asignadas * 100), 1) if total_asignadas > 0 else 0.0
        
        kpis = {
            'total_asignadas': total_asignadas,
            'total_finalizadas': total_finalizadas,
            'tasa_efectividad': tasa_efectividad,
            'avg_traslado': round(float(kpis_row['avg_traslado'] or 0), 1),
            'avg_resolucion': round(float(kpis_row['avg_resolucion'] or 0), 1),
            'avg_total': round(float(kpis_row['avg_total'] or 0), 1)
        }

        # 2. Comparativa por Técnico (solo si no se filtra por un técnico específico)
        comparativa_tecnicos = []
        if not tecnico_filtro or tecnico_filtro == 'TODOS':
            query_comp = f"""
                SELECT 
                    tecnico_principal AS tecnico,
                    COUNT(*) as asignadas,
                    SUM(CASE WHEN estado IN ('FINALIZADA', 'SOLVENTADA_REMOTA') 
                              AND (solucion_tecnico IS NULL OR (
                                  solucion_tecnico NOT LIKE '%GESTIONAR ARREGLO%'
                                  AND solucion_tecnico NOT LIKE '%SOLUCIÓN PARCIAL%'
                                  AND solucion_tecnico NOT LIKE '%SOLUCION PARCIAL%'
                                  AND solucion_tecnico NOT LIKE '%GENERAR CAMBIO DE FO%'
                                  AND solucion_tecnico NOT LIKE '%SIN RESPUESTA%'
                                  AND solucion_tecnico NOT LIKE '%NO SE PUEDE REALIZAR VISITA%'
                                  AND solucion_tecnico NOT LIKE '%NOC%'
                              )) THEN 1 ELSE 0 END) as finalizadas,
                    AVG(CASE WHEN estado = 'FINALIZADA' AND hora_en_ruta IS NOT NULL AND hora_inicio_visita IS NOT NULL 
                             THEN TIMESTAMPDIFF(MINUTE, hora_en_ruta, hora_inicio_visita) ELSE NULL END) as avg_traslado,
                    AVG(CASE WHEN estado = 'FINALIZADA' AND hora_inicio_visita IS NOT NULL AND hora_fin_visita IS NOT NULL 
                             THEN TIMESTAMPDIFF(MINUTE, hora_inicio_visita, hora_fin_visita) ELSE NULL END) as avg_resolucion
                FROM visitas_tecnicas
                {where_clause}
                  AND tecnico_principal IS NOT NULL AND tecnico_principal != '' AND tecnico_principal NOT IN ('NO TECNICO', 'TECNOLOGIA')
                GROUP BY tecnico_principal
                ORDER BY finalizadas DESC
            """
            cursor.execute(query_comp, params)
            rows_comp = cursor.fetchall()
            for r in rows_comp:
                comparativa_tecnicos.append({
                    'tecnico': r['tecnico'],
                    'asignadas': r['asignadas'],
                    'finalizadas': r['finalizadas'],
                    'avg_traslado': round(float(r['avg_traslado'] or 0), 1),
                    'avg_resolucion': round(float(r['avg_resolucion'] or 0), 1)
                })

        # 3. Tiempos por Tipo de Problema / Producto
        categoria_col = 'producto' if es_instalacion_val == 1 else 'problema'
        query_problemas = f"""
            SELECT 
                {categoria_col} as categoria,
                COUNT(*) as cantidad,
                AVG(CASE WHEN estado = 'FINALIZADA' AND hora_inicio_visita IS NOT NULL AND hora_fin_visita IS NOT NULL 
                         THEN TIMESTAMPDIFF(MINUTE, hora_inicio_visita, hora_fin_visita) ELSE NULL END) as avg_resolucion
            FROM visitas_tecnicas
            {where_clause}
              AND {categoria_col} IS NOT NULL AND {categoria_col} != ''
            GROUP BY {categoria_col}
            ORDER BY cantidad DESC
            LIMIT 8
        """
        cursor.execute(query_problemas, params)
        rows_prob = cursor.fetchall()
        tiempos_problemas = []
        for r in rows_prob:
            tiempos_problemas.append({
                'categoria': r['categoria'],
                'cantidad': r['cantidad'],
                'avg_resolucion': round(float(r['avg_resolucion'] or 0), 1)
            })

        # 4. Evolución semanal de tiempos
        query_evolucion = f"""
            SELECT 
                DATE_FORMAT(fecha_programada, '%Y-%u') as semana,
                MIN(fecha_programada) as inicio_semana,
                AVG(CASE WHEN estado = 'FINALIZADA' AND hora_en_ruta IS NOT NULL AND hora_inicio_visita IS NOT NULL 
                         THEN TIMESTAMPDIFF(MINUTE, hora_en_ruta, hora_inicio_visita) ELSE NULL END) as avg_traslado,
                AVG(CASE WHEN estado = 'FINALIZADA' AND hora_inicio_visita IS NOT NULL AND hora_fin_visita IS NOT NULL 
                         THEN TIMESTAMPDIFF(MINUTE, hora_inicio_visita, hora_fin_visita) ELSE NULL END) as avg_resolucion
            FROM visitas_tecnicas
            {where_clause}
            GROUP BY semana
            ORDER BY inicio_semana ASC
        """
        cursor.execute(query_evolucion, params)
        rows_evol = cursor.fetchall()
        evolucion = []
        for r in rows_evol:
            ini_dt = r['inicio_semana']
            fecha_lbl = ini_dt.strftime('%d/%m') if isinstance(ini_dt, (datetime, date)) else str(ini_dt)
            evolucion.append({
                'label': f"Sem {fecha_lbl}",
                'avg_traslado': round(float(r['avg_traslado'] or 0), 1),
                'avg_resolucion': round(float(r['avg_resolucion'] or 0), 1)
            })

        # 5. Detalle de Bitácora SLA (últimas 50 visitas con tiempos calculados)
        query_bitacora = f"""
            SELECT 
                id_visita,
                cliente,
                contrato,
                tecnico_principal,
                sector,
                {categoria_col} as categoria,
                estado,
                hora_en_ruta,
                hora_inicio_visita,
                hora_fin_visita,
                CASE WHEN hora_en_ruta IS NOT NULL AND hora_inicio_visita IS NOT NULL 
                     THEN TIMESTAMPDIFF(MINUTE, hora_en_ruta, hora_inicio_visita) ELSE NULL END as tiempo_traslado,
                CASE WHEN hora_inicio_visita IS NOT NULL AND hora_fin_visita IS NOT NULL 
                     THEN TIMESTAMPDIFF(MINUTE, hora_inicio_visita, hora_fin_visita) ELSE NULL END as tiempo_resolucion
            FROM visitas_tecnicas
            {where_clause}
            ORDER BY fecha_programada DESC, id_visita DESC
            LIMIT 50
        """
        cursor.execute(query_bitacora, params)
        rows_bit = cursor.fetchall()
        bitacora = []
        for r in rows_bit:
            bitacora.append({
                'id_visita': r['id_visita'],
                'cliente': r['cliente'],
                'contrato': r['contrato'] or '',
                'tecnico': r['tecnico_principal'] or 'Sin asignar',
                'sector': r['sector'] or '-',
                'categoria': r['categoria'] or 'General',
                'estado': r['estado'],
                'hora_en_ruta': r['hora_en_ruta'].isoformat() if r['hora_en_ruta'] else None,
                'hora_inicio': r['hora_inicio_visita'].isoformat() if r['hora_inicio_visita'] else None,
                'hora_fin': r['hora_fin_visita'].isoformat() if r['hora_fin_visita'] else None,
                'tiempo_traslado': r['tiempo_traslado'],
                'tiempo_resolucion': r['tiempo_resolucion']
            })

        return jsonify({
            'status': 'ok',
            'kpis': kpis,
            'comparativa_tecnicos': comparativa_tecnicos,
            'tiempos_problemas': tiempos_problemas,
            'evolucion': evolucion,
            'bitacora': bitacora
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


# ==========================================
# RUTAS DE GESTIÓN DE RECORDATORIOS Y BLOQUEOS
# ==========================================

@admin_bp.route('/api/admin/recordatorios', methods=['GET'])
def obtener_recordatorios():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'id_usuario': session['user_id'], 'rol': session.get('user_role')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
        
    fecha_filtro = request.args.get('fecha')
    
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    try:
        cursor = conexion.cursor(dictionary=True)
        if fecha_filtro:
            query = """
                SELECT r.*, t.nombre as tecnico_nombre 
                FROM recordatorios_bloqueos r
                LEFT JOIN tecnicos t ON r.tecnico_id = t.id_tecnico
                WHERE r.fecha = %s AND r.activo = 1
                ORDER BY r.hora_inicio ASC
            """
            cursor.execute(query, (fecha_filtro,))
        else:
            query = """
                SELECT r.*, t.nombre as tecnico_nombre 
                FROM recordatorios_bloqueos r
                LEFT JOIN tecnicos t ON r.tecnico_id = t.id_tecnico
                WHERE r.activo = 1
                ORDER BY r.fecha DESC, r.hora_inicio ASC
            """
            cursor.execute(query)
            
        rows = cursor.fetchall()
        
        # Formatear timedelta / date a string para JSON
        for r in rows:
            r['fecha'] = r['fecha'].isoformat()
            if r['hora_inicio']:
                tot_sec = int(r['hora_inicio'].total_seconds())
                r['hora_inicio'] = f"{tot_sec // 3600:02d}:{(tot_sec % 3600) // 60:02d}"
            if r['hora_fin']:
                tot_sec = int(r['hora_fin'].total_seconds())
                r['hora_fin'] = f"{tot_sec // 3600:02d}:{(tot_sec % 3600) // 60:02d}"
            if r['fecha_creacion']:
                r['fecha_creacion'] = r['fecha_creacion'].isoformat()
                
        return jsonify({"status": "ok", "recordatorios": rows})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()

@admin_bp.route('/api/admin/recordatorios', methods=['POST'])
def crear_recordatorio():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'id_usuario': session['user_id'], 'rol': session.get('user_role'), 'nombre': session.get('user_name')}

    user_role = user.get('role') or user.get('rol') if user else None
    if not user or user_role not in ['ADMIN', 'ASESOR', 'CALIDAD', 'ATC']:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
        
    datos = request.get_json() or {}
    titulo = datos.get('titulo', '').strip()
    descripcion = datos.get('descripcion', '').strip()
    tipo = datos.get('tipo')
    fecha = datos.get('fecha')
    hora_inicio = datos.get('hora_inicio') or None
    hora_fin = datos.get('hora_fin') or None
    tecnico_id = datos.get('tecnico_id') or None
    
    if not titulo or not tipo or not fecha:
        return jsonify({"status": "error", "message": "Título, Tipo y Fecha son obligatorios."}), 400
        
    creado_por = user.get('nombre') or user.get('username') or 'Sistema'
    
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    try:
        cursor = conexion.cursor()
        query = """
            INSERT INTO recordatorios_bloqueos 
            (titulo, descripcion, tipo, fecha, hora_inicio, hora_fin, tecnico_id, creado_por, activo)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 1)
        """
        cursor.execute(query, (titulo, descripcion, tipo, fecha, hora_inicio, hora_fin, tecnico_id, creado_por))
        conexion.commit()
        return jsonify({"status": "ok", "message": "Recordatorio/Bloqueo creado con éxito."})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()

@admin_bp.route('/api/admin/recordatorios/<int:id_recordatorio>', methods=['DELETE'])
@admin_bp.route('/api/admin/recordatorios/<int:id_recordatorio>/atender', methods=['POST'])
def eliminar_recordatorio(id_recordatorio):
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'id_usuario': session['user_id'], 'rol': session.get('user_role')}

    user_role = user.get('role') or user.get('rol') if user else None
    if not user or user_role not in ['ADMIN', 'ASESOR', 'CALIDAD', 'ATC']:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
        
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    try:
        cursor = conexion.cursor()
        cursor.execute("UPDATE recordatorios_bloqueos SET activo = 0 WHERE id_recordatorio = %s", (id_recordatorio,))
        conexion.commit()
        return jsonify({"status": "ok", "message": "Recordatorio/Bloqueo eliminado con éxito."})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()

@admin_bp.route('/api/admin/recordatorios/verificar', methods=['GET'])
def verificar_conflictos_agenda():
    if 'user_id' not in session:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
        
    fecha = request.args.get('fecha')
    tecnico = request.args.get('tecnico', '').strip()
    preferencia = request.args.get('preferencia', '').strip().lower()
    
    if not fecha:
        return jsonify({"status": "error", "message": "Falta la fecha de consulta."}), 400
        
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de conexión a la base de datos"}), 500
        
    try:
        cursor = conexion.cursor(dictionary=True)
        # Traer todos los recordatorios activos para esa fecha
        cursor.execute("""
            SELECT r.*, t.nombre as tecnico_nombre 
            FROM recordatorios_bloqueos r
            LEFT JOIN tecnicos t ON r.tecnico_id = t.id_tecnico
            WHERE r.fecha = %s AND r.activo = 1
        """, (fecha,))
        recordatorios = cursor.fetchall()
        
        conflictos = []
        
        # Función auxiliar para convertir preferencia horaria a minutos del día
        from utils import normalizar_horario_texto
        pref_inicio, pref_fin = normalizar_horario_texto(preferencia)
        
        for r in recordatorios:
            # Convertir hora_inicio y hora_fin a minutos
            rec_inicio_min = None
            rec_fin_min = None
            
            if r['hora_inicio']:
                tot_sec = int(r['hora_inicio'].total_seconds())
                rec_inicio_min = tot_sec // 60
            if r['hora_fin']:
                tot_sec = int(r['hora_fin'].total_seconds())
                rec_fin_min = tot_sec // 60
                
            # Determinar si hay solapamiento horario si el recordatorio tiene horas
            solapamiento = True
            if rec_inicio_min is not None and rec_fin_min is not None and pref_inicio is not None and pref_fin is not None:
                # Caso de no solapamiento: (fin_rec <= inicio_pref) o (inicio_rec >= fin_pref)
                if rec_fin_min <= pref_inicio or rec_inicio_min >= pref_fin:
                    solapamiento = False
                    
            if not solapamiento:
                continue
                
            # Reglas de filtrado
            tipo = r['tipo']
            tipo_label = {
                'INVENTARIO': 'Inventario General',
                'REUNION': 'Reunión General',
                'BLOQUEO_GENERAL': 'Bloqueo General',
                'RECORDATORIO_TECNICO': 'Recordatorio Técnico'
            }.get(tipo, tipo)
            
            hora_str = ""
            if r['hora_inicio'] and r['hora_fin']:
                ts_ini = int(r['hora_inicio'].total_seconds())
                ts_fin = int(r['hora_fin'].total_seconds())
                hora_str = f" ({ts_ini//3600:02d}:{(ts_ini%3600)//60:02d} - {ts_fin//3600:02d}:{(ts_fin%3600)//60:02d})"
            
            desc_str = f": {r['descripcion']}" if r['descripcion'] else ""
            
            # A. Si es un bloqueo general (Inventario / Reunión / Bloqueo General) -> Aplica a todos
            if tipo in ['INVENTARIO', 'REUNION', 'BLOQUEO_GENERAL']:
                conflictos.append(f"<strong>[{tipo_label}]</strong> {r['titulo']}{hora_str}{desc_str}")
            
            # B. Si es recordatorio de técnico -> Solo aplica a ese técnico
            elif tipo == 'RECORDATORIO_TECNICO' and r['tecnico_nombre'] and tecnico:
                if r['tecnico_nombre'].upper() == tecnico.upper():
                    conflictos.append(f"<strong>[{tipo_label}]</strong> Técnico {r['tecnico_nombre']} tiene: {r['titulo']}{hora_str}{desc_str}")
                    
        return jsonify({"status": "ok", "conflictos": conflictos})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/visitas/<int:id_visita>/editar', methods=['POST'])
def editar_visita(id_visita):
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'id_usuario': session['user_id'], 'rol': session.get('user_role')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401
    
    user_role = user.get('role') or user.get('rol')
    if user_role not in ['ADMIN', 'ASESOR', 'CALIDAD', 'ATC']:
        return jsonify({"status": "error", "message": "No tienes permiso para editar visitas."}), 403

    # Validar que la visita no esté cerrada o en progreso antes de editar
    conexion_check = get_db_connection()
    cursor_check = conexion_check.cursor(dictionary=True)
    try:
        cursor_check.execute("SELECT estado FROM visitas_tecnicas WHERE id_visita = %s", (id_visita,))
        visita_actual = cursor_check.fetchone()
        if not visita_actual:
            return jsonify({"status": "error", "message": "Visita no encontrada."}), 404
        if visita_actual['estado'] not in ['PENDIENTE', 'EN_RUTA']:
            return jsonify({"status": "error", "message": "No se puede editar una visita que ya se encuentra cerrada o en progreso."}), 400
    finally:
        cursor_check.close()
        conexion_check.close()

    cliente = request.form.get('cliente', '').strip()
    contrato = request.form.get('contrato', '').strip() or None
    telefonos = request.form.get('telefonos', '').strip() or None
    sector = request.form.get('sector', '').strip()
    direccion = request.form.get('direccion', '').strip()
    lat = request.form.get('latitud', '').strip() or None
    lon = request.form.get('longitud', '').strip() or None
    preferencia = request.form.get('preferencia_horaria', '').strip()
    servicio = request.form.get('servicio', '').strip()
    
    velocidad_mbps_raw = request.form.get('velocidad_mbps', '').strip()
    velocidad_mbps = int(velocidad_mbps_raw) if velocidad_mbps_raw else None
    
    problema = request.form.get('problema', '').strip()
    obs_call = request.form.get('observacion_callcenter', '').strip() or None
    fecha_prog = request.form.get('fecha_programada', '').strip()
    estado = request.form.get('estado', '').strip()
    tecnico_principal = request.form.get('tecnico_principal')
    tecnico_apoyo = request.form.get('tecnico_apoyo')

    # Reconstruir información técnica
    info_parts = []
    caja = request.form.get('info_caja', '').strip()
    hilo = request.form.get('info_hilo', '').strip()
    ip = request.form.get('info_ip', '').strip()
    vlan = request.form.get('info_vlan', '').strip()
    usr = request.form.get('info_usr', '').strip()
    pas = request.form.get('info_pas', '').strip()

    if caja: info_parts.append(f"CAJA: {caja}")
    if hilo: info_parts.append(f"HILO: {hilo}")
    if ip: info_parts.append(f"IP: {ip}")
    if vlan: info_parts.append(f"VLAN: {vlan}")
    if usr: info_parts.append(f"USR: {usr}")
    if pas: info_parts.append(f"PAS: {pas}")
    informacion_tecnico = "\n".join(info_parts) if info_parts else None

    # Normalizar horario texto a minutos para el optimizador
    from utils import normalizar_horario_texto
    ventana_inicio, ventana_fin = normalizar_horario_texto(preferencia)

    conexion = get_db_connection()
    cursor = conexion.cursor()
    try:
        # Si no viene especificado en la petición, preservar los técnicos asignados actuales
        if tecnico_principal is None:
            cursor.execute("SELECT tecnico_principal, tecnico_apoyo FROM visitas_tecnicas WHERE id_visita = %s", (id_visita,))
            row_tec = cursor.fetchone()
            if row_tec:
                tecnico_principal = row_tec[0]
                tecnico_apoyo = row_tec[1]

        # Si se restablece a un estado activo, limpiar datos de cierre para permitir que se vuelva a ejecutar
        if estado in ['PENDIENTE', 'REAGENDADA', 'EN_RUTA', 'EN_PROGRESO']:
            query = """
                UPDATE visitas_tecnicas 
                SET cliente = %s, contrato = %s, telefonos = %s, sector = %s, direccion = %s,
                    latitud = %s, longitud = %s, preferencia_horaria = %s, servicio = %s,
                    velocidad_mbps = %s, problema = %s, observacion_callcenter = %s,
                    informacion_tecnico = %s, fecha_programada = %s, estado = %s,
                    tecnico_principal = %s, tecnico_apoyo = %s, ventana_inicio_min = %s, ventana_fin_min = %s,
                    hora_fin_visita = NULL, solucion_tecnico = NULL, observacion_tecnico = NULL,
                    modelo_onu = NULL, modelo_router = NULL, coordenadas_tecnico = NULL,
                    foto_equipos = NULL, foto_equipos_2 = NULL, firma_cliente = NULL
                WHERE id_visita = %s
            """
            cursor.execute(query, (
                cliente, contrato, telefonos, sector, direccion, lat, lon, preferencia, servicio,
                velocidad_mbps, problema, obs_call, informacion_tecnico, fecha_prog, estado,
                tecnico_principal, tecnico_apoyo, ventana_inicio, ventana_fin, id_visita
            ))
        else:
            query = """
                UPDATE visitas_tecnicas 
                SET cliente = %s, contrato = %s, telefonos = %s, sector = %s, direccion = %s,
                    latitud = %s, longitud = %s, preferencia_horaria = %s, servicio = %s,
                    velocidad_mbps = %s, problema = %s, observacion_callcenter = %s,
                    informacion_tecnico = %s, fecha_programada = %s, estado = %s,
                    tecnico_principal = %s, tecnico_apoyo = %s, ventana_inicio_min = %s, ventana_fin_min = %s
                WHERE id_visita = %s
            """
            cursor.execute(query, (
                cliente, contrato, telefonos, sector, direccion, lat, lon, preferencia, servicio,
                velocidad_mbps, problema, obs_call, informacion_tecnico, fecha_prog, estado,
                tecnico_principal, tecnico_apoyo, ventana_inicio, ventana_fin, id_visita
            ))
        conexion.commit()
        return jsonify({"status": "ok", "message": "Visita actualizada correctamente."})
    except Exception as e:
        print(f"Error al editar visita: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/inventario/rastreo_sn/<sn>', methods=['GET'])
def api_rastreo_sn(sn):
    """
    Rastreador global de número de serie (ONU / Router)
    Busca en directorio de clientes (activos), visitas técnicas y equipos retirados.
    """
    sn_clean = sn.strip()
    if not sn_clean:
        return jsonify({"status": "error", "message": "Número de serie no especificado"}), 400

    conexion = get_db_connection()
    cursor = conexion.cursor(dictionary=True)
    try:
        like_pattern = f"%{sn_clean}%"
        
        # 1. Buscar en directorio_clientes
        sql_clientes = """
            SELECT contrato, nombre_cliente, cedula, direccion,
                   numero_serie AS onu_sn, router_principal, numero_serie_router AS router_sn,
                   router_secundario, numero_serie_router_secundario AS router2_sn
            FROM directorio_clientes
            WHERE numero_serie LIKE %s 
               OR numero_serie_router LIKE %s 
               OR numero_serie_router_secundario LIKE %s
            LIMIT 5
        """
        cursor.execute(sql_clientes, (like_pattern, like_pattern, like_pattern))
        clientes_matches = cursor.fetchall()

        # 2. Buscar en equipos_retirados_visitas
        sql_retirados = """
            SELECT id_retiro, id_visita, tipo_equipo, numero_serie, modelo,
                   motivo_retiro, observacion_retiro, tecnico, placa_vehiculo,
                   estado_custodia, fecha_retiro, fecha_devolucion_bodega, recibido_por
            FROM equipos_retirados_visitas
            WHERE numero_serie LIKE %s
            ORDER BY fecha_retiro DESC
            LIMIT 10
        """
        cursor.execute(sql_retirados, (like_pattern,))
        retirados_matches = cursor.fetchall()
        for r in retirados_matches:
            if r.get('fecha_retiro'):
                r['fecha_retiro'] = str(r['fecha_retiro'])
            if r.get('fecha_devolucion_bodega'):
                r['fecha_devolucion_bodega'] = str(r['fecha_devolucion_bodega'])

        # 3. Buscar en visitas_tecnicas
        sql_visitas = """
            SELECT id_visita, contrato, cliente, direccion, tecnico_principal, tecnico_apoyo,
                   fecha_programada, estado, numero_serie_onu, numero_serie_router,
                   numero_serie_router_secundario, modelo_onu, modelo_router
            FROM visitas_tecnicas
            WHERE numero_serie_onu LIKE %s 
               OR numero_serie_router LIKE %s 
               OR numero_serie_router_secundario LIKE %s
            ORDER BY fecha_programada DESC
            LIMIT 10
        """
        cursor.execute(sql_visitas, (like_pattern, like_pattern, like_pattern))
        visitas_matches = cursor.fetchall()
        for v in visitas_matches:
            if v.get('fecha_programada'):
                v['fecha_programada'] = str(v['fecha_programada'])

        # Determinar estado principal
        estado_global = "NO_ENCONTRADO"
        resumen = {}

        if clientes_matches:
            c = clientes_matches[0]
            tipo_coincide = "ONU" if sn_clean.lower() in (c.get('onu_sn') or '').lower() else "ROUTER"
            estado_global = "INSTALADO_EN_CLIENTE"
            resumen = {
                "tipo": "CLIENTE_ACTIVO",
                "contrato": c.get('contrato'),
                "cliente": c.get('nombre_cliente'),
                "cedula": c.get('cedula'),
                "direccion": c.get('direccion'),
                "equipo_coincide": tipo_coincide,
                "onu_sn": c.get('onu_sn'),
                "router_sn": c.get('router_sn')
            }
        elif retirados_matches:
            r = retirados_matches[0]
            if r.get('estado_custodia') == 'EN_VEHICULO':
                estado_global = "RETIRADO_EN_CUSTODIA"
            else:
                estado_global = "RETIRADO_DEVUELTO_BODEGA"
            resumen = {
                "tipo": "EQUIPO_RETIRADO",
                "tecnico": r.get('tecnico'),
                "placa": r.get('placa_vehiculo'),
                "estado_custodia": r.get('estado_custodia'),
                "motivo": r.get('motivo_retiro'),
                "observacion": r.get('observacion_retiro'),
                "fecha_retiro": r.get('fecha_retiro'),
                "fecha_devolucion": r.get('fecha_devolucion_bodega'),
                "recibido_por": r.get('recibido_por')
            }
        elif visitas_matches:
            v = visitas_matches[0]
            estado_global = "REGISTRADO_EN_VISITA"
            resumen = {
                "tipo": "VISITA_TECNICA",
                "id_visita": v.get('id_visita'),
                "contrato": v.get('contrato'),
                "cliente": v.get('cliente'),
                "tecnico": v.get('tecnico_principal'),
                "fecha": v.get('fecha_programada'),
                "estado_visita": v.get('estado')
            }

        return jsonify({
            "status": "ok",
            "sn_buscado": sn_clean,
            "estado_global": estado_global,
            "resumen": resumen,
            "clientes": clientes_matches,
            "retirados": retirados_matches,
            "visitas": visitas_matches
        })
    except Exception as e:
        print(f"Error en api_rastreo_sn: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


# ==========================================================
# GESTIÓN Y TRAZABILIDAD DE EQUIPOS (ONUs Y ROUTERS)
# ==========================================================

@admin_bp.route('/api/admin/catalogo_ont', methods=['GET'])
def api_admin_catalogo_ont():
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("SELECT id_ont, nombre, activo FROM catalogo_modelos_ont WHERE activo = 1 ORDER BY nombre ASC")
        return jsonify({"status": "ok", "catalogos": cursor.fetchall()})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/catalogo_router', methods=['GET'])
def api_admin_catalogo_router():
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("SELECT id_router, nombre, activo FROM catalogo_modelos_router WHERE activo = 1 ORDER BY nombre ASC")
        return jsonify({"status": "ok", "catalogos": cursor.fetchall()})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/equipos/resumen', methods=['GET'])
def api_equipos_resumen():
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cursor = conexion.cursor(dictionary=True)
        # Resumen general por estado en trazabilidad
        cursor.execute("""
            SELECT estado, COUNT(*) as cantidad
            FROM trazabilidad_equipos
            GROUP BY estado
        """)
        por_estado = {r['estado']: r['cantidad'] for r in cursor.fetchall()}

        # Total real de equipos retirados en visitas de soporte
        cursor.execute("SELECT COUNT(*) as c FROM equipos_retirados_visitas")
        total_retirados_visitas = cursor.fetchone().get('c', 0)
        total_retirados = max(por_estado.get('RETIRADO_AVERIA', 0), total_retirados_visitas)

        # Resumen por modelo en bodega
        cursor.execute("""
            SELECT tipo_equipo, modelo, marca,
                   SUM(CASE WHEN estado = 'EN_BODEGA' THEN 1 ELSE 0 END) as en_bodega,
                   SUM(CASE WHEN estado = 'EN_VEHICULO' THEN 1 ELSE 0 END) as en_vehiculo,
                   SUM(CASE WHEN estado = 'INSTALADO_CLIENTE' THEN 1 ELSE 0 END) as instalados,
                   SUM(CASE WHEN estado = 'RETIRADO_AVERIA' THEN 1 ELSE 0 END) as retirados,
                   COUNT(*) as total
            FROM trazabilidad_equipos
            GROUP BY tipo_equipo, modelo, marca
            ORDER BY tipo_equipo ASC, modelo ASC
        """)
        por_modelo = cursor.fetchall()

        # Resumen por placa
        cursor.execute("""
            SELECT ubicacion_placa as placa, COUNT(*) as total_equipos
            FROM trazabilidad_equipos
            WHERE estado = 'EN_VEHICULO' AND ubicacion_placa IS NOT NULL AND ubicacion_placa != ''
            GROUP BY ubicacion_placa
            ORDER BY ubicacion_placa ASC
        """)
        por_placa = cursor.fetchall()

        return jsonify({
            "status": "ok",
            "totales": {
                "en_bodega": por_estado.get('EN_BODEGA', 0),
                "en_vehiculo": por_estado.get('EN_VEHICULO', 0),
                "instalados": por_estado.get('INSTALADO_CLIENTE', 0),
                "retirados": total_retirados,
                "total": por_estado.get('EN_BODEGA', 0) + por_estado.get('EN_VEHICULO', 0) + por_estado.get('INSTALADO_CLIENTE', 0) + total_retirados
            },
            "modelos": por_modelo,
            "placas": por_placa
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/equipos/lista', methods=['GET'])
def api_equipos_lista():
    estado = request.args.get('estado', '').strip()
    modelo = request.args.get('modelo', '').strip()
    tipo = request.args.get('tipo', '').strip()
    placa = request.args.get('placa', '').strip()
    search = request.args.get('search', '').strip()

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cursor = conexion.cursor(dictionary=True)

        # Si el filtro es RETIRADOS, consultamos directamente la tabla de retiros de visitas
        if estado == 'RETIRADO_AVERIA':
            query = """
                SELECT er.id_retiro as id_equipo, er.id_retiro, er.tipo_equipo,
                       CASE 
                           WHEN er.modelo LIKE '%HUAWEI%' THEN 'HUAWEI'
                           WHEN er.modelo LIKE '%TP-LINK%' OR er.modelo LIKE '%TPLINK%' OR er.modelo LIKE '%EX511%' OR er.modelo LIKE '%XX231%' OR er.modelo LIKE '%XX530%' THEN 'TP-LINK'
                           WHEN er.modelo LIKE '%MERCUSYS%' OR er.modelo LIKE '%MR70%' THEN 'MERCUSYS'
                           WHEN er.modelo LIKE '%FIBERHOME%' THEN 'FIBERHOME'
                           WHEN er.modelo LIKE '%ZTE%' THEN 'ZTE'
                           ELSE ''
                       END as marca,
                       er.modelo, er.numero_serie,
                       'RETIRADO_AVERIA' as estado,
                       er.placa_vehiculo as ubicacion_placa,
                       er.tecnico as tecnico_entrega,
                       v.contrato as contrato_cliente,
                       v.cliente as nombre_cliente,
                       er.motivo_retiro, er.observacion_retiro,
                       er.estado_custodia, er.fecha_devolucion_bodega, er.recibido_por,
                       er.fecha_retiro as fecha_ingreso_bodega
                FROM equipos_retirados_visitas er
                LEFT JOIN visitas_tecnicas v ON er.id_visita = v.id_visita
                WHERE 1=1
            """
            params = []
            if modelo:
                query += " AND er.modelo = %s"
                params.append(modelo)
            if tipo:
                query += " AND er.tipo_equipo = %s"
                params.append(tipo)
            if placa:
                query += " AND er.placa_vehiculo = %s"
                params.append(placa)
            if search:
                query += " AND (er.numero_serie LIKE %s OR er.modelo LIKE %s OR er.tecnico LIKE %s OR v.cliente LIKE %s OR v.contrato LIKE %s)"
                search_param = f"%{search}%"
                params.extend([search_param, search_param, search_param, search_param, search_param])

            query += " ORDER BY er.fecha_retiro DESC LIMIT 500"
            cursor.execute(query, tuple(params))
            equipos = cursor.fetchall()
            for eq in equipos:
                if eq.get('fecha_ingreso_bodega') and hasattr(eq['fecha_ingreso_bodega'], 'isoformat'):
                    eq['fecha_ingreso_bodega'] = eq['fecha_ingreso_bodega'].isoformat()
                if eq.get('fecha_devolucion_bodega') and hasattr(eq['fecha_devolucion_bodega'], 'isoformat'):
                    eq['fecha_devolucion_bodega'] = eq['fecha_devolucion_bodega'].isoformat()

            return jsonify({"status": "ok", "equipos": equipos, "total": len(equipos), "es_retirados": True})

        # Para los demás estados (EN_BODEGA, EN_VEHICULO, INSTALADO_CLIENTE)
        query = "SELECT * FROM trazabilidad_equipos WHERE 1=1"
        params = []

        if estado:
            query += " AND estado = %s"
            params.append(estado)
        if modelo:
            query += " AND modelo = %s"
            params.append(modelo)
        if tipo:
            query += " AND tipo_equipo = %s"
            params.append(tipo)
        if placa:
            query += " AND ubicacion_placa = %s"
            params.append(placa)
        if search:
            query += " AND (numero_serie LIKE %s OR modelo LIKE %s OR contrato_cliente LIKE %s OR nombre_cliente LIKE %s)"
            search_param = f"%{search}%"
            params.extend([search_param, search_param, search_param, search_param])

        query += " ORDER BY id_equipo DESC LIMIT 500"
        cursor.execute(query, tuple(params))
        equipos = cursor.fetchall()

        for eq in equipos:
            for f in ['fecha_ingreso_bodega', 'fecha_entrega_vehiculo', 'fecha_instalacion']:
                if eq.get(f) and hasattr(eq[f], 'isoformat'):
                    eq[f] = eq[f].isoformat()

        return jsonify({"status": "ok", "equipos": equipos, "total": len(equipos), "es_retirados": False})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/equipos/ingreso_masivo', methods=['POST'])
def api_equipos_ingreso_masivo():
    data = request.get_json() or {}
    tipo_equipo = (data.get('tipo_equipo') or 'ONT').strip().upper()
    modelo = (data.get('modelo') or '').strip()
    marca = (data.get('marca') or 'GENERAL').strip().upper()
    seriales_raw = data.get('seriales') or []
    observacion = (data.get('observacion') or '').strip()
    user_name = session.get('user_name', 'Administrador')

    if not modelo:
        return jsonify({"status": "error", "message": "Debe especificar el modelo"}), 400

    # Limpiar y deduplicar seriales enviados en el lote
    seriales = []
    for s in seriales_raw:
        s_clean = str(s).strip().upper()
        if s_clean and s_clean not in seriales:
            seriales.append(s_clean)

    if not seriales:
        return jsonify({"status": "error", "message": "No se enviaron números de serie válidos"}), 400

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cursor = conexion.cursor(dictionary=True)
        # Verificar seriales ya existentes
        format_strings = ','.join(['%s'] * len(seriales))
        cursor.execute(f"SELECT numero_serie, estado, ubicacion_placa FROM trazabilidad_equipos WHERE numero_serie IN ({format_strings})", tuple(seriales))
        existentes_map = {r['numero_serie']: r for r in cursor.fetchall()}

        ingresados = 0
        actualizados = 0
        duplicados = []

        for sn in seriales:
            if sn in existentes_map:
                r = existentes_map[sn]
                if r['estado'] == 'EN_BODEGA':
                    duplicados.append(sn)
                else:
                    # Si estaba en vehículo o retirado y se vuelve a ingresar a bodega
                    cursor.execute("""
                        UPDATE trazabilidad_equipos
                        SET estado = 'EN_BODEGA',
                            ubicacion_placa = NULL,
                            fecha_ingreso_bodega = NOW(),
                            modelo = %s,
                            tipo_equipo = %s,
                            marca = %s,
                            observacion = CONCAT(COALESCE(observacion, ''), ' | Reingresado a bodega')
                        WHERE numero_serie = %s
                    """, (modelo, tipo_equipo, marca, sn))
                    actualizados += 1
            else:
                cursor.execute("""
                    INSERT INTO trazabilidad_equipos (tipo_equipo, modelo, marca, numero_serie, estado, registrado_por, observacion)
                    VALUES (%s, %s, %s, %s, 'EN_BODEGA', %s, %s)
                """, (tipo_equipo, modelo, marca, sn, user_name, observacion if observacion else None))
                ingresados += 1

        conexion.commit()
        return jsonify({
            "status": "ok",
            "message": f"Ingreso completado: {ingresados} nuevos, {actualizados} reingresados.",
            "ingresados": ingresados,
            "actualizados": actualizados,
            "duplicados": duplicados
        })
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/equipos/despacho_buseta', methods=['POST'])
def api_equipos_despacho_buseta():
    data = request.get_json() or {}
    placa_vehiculo = (data.get('placa_vehiculo') or '').strip().upper()
    seriales_raw = data.get('seriales') or []
    observacion = (data.get('observacion') or '').strip()

    if not placa_vehiculo:
        return jsonify({"status": "error", "message": "Debe seleccionar una placa de destino"}), 400

    seriales = [str(s).strip().upper() for s in seriales_raw if str(s).strip()]
    if not seriales:
        return jsonify({"status": "error", "message": "No se enviaron seriales a despachar"}), 400

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cursor = conexion.cursor(dictionary=True)
        format_strings = ','.join(['%s'] * len(seriales))
        cursor.execute(f"SELECT numero_serie, estado, modelo FROM trazabilidad_equipos WHERE numero_serie IN ({format_strings})", tuple(seriales))
        encontrados = {r['numero_serie']: r for r in cursor.fetchall()}

        despachados = 0
        no_encontrados = []
        ya_en_vehiculo = []

        for sn in seriales:
            if sn not in encontrados:
                no_encontrados.append(sn)
            else:
                eq = encontrados[sn]
                cursor.execute("""
                    UPDATE trazabilidad_equipos
                    SET estado = 'EN_VEHICULO',
                        ubicacion_placa = %s,
                        fecha_entrega_vehiculo = NOW(),
                        observacion = CONCAT(COALESCE(observacion, ''), %s)
                    WHERE numero_serie = %s
                """, (placa_vehiculo, f" | Despachado a {placa_vehiculo}", sn))
                despachados += 1

        conexion.commit()
        return jsonify({
            "status": "ok",
            "message": f"Despacho exitoso: {despachados} equipos asignados a {placa_vehiculo}.",
            "despachados": despachados,
            "no_encontrados": no_encontrados
        })
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/equipos/reingreso_bodega', methods=['POST'])
def api_equipos_reingreso_bodega():
    data = request.get_json() or {}
    seriales_raw = data.get('seriales') or []
    seriales = [str(s).strip().upper() for s in seriales_raw if str(s).strip()]
    if not seriales:
        return jsonify({"status": "error", "message": "No se enviaron seriales"}), 400

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cursor = conexion.cursor()
        format_strings = ','.join(['%s'] * len(seriales))
        cursor.execute(f"""
            UPDATE trazabilidad_equipos
            SET estado = 'EN_BODEGA',
                ubicacion_placa = NULL,
                fecha_ingreso_bodega = NOW()
            WHERE numero_serie IN ({format_strings})
        """, tuple(seriales))
        conexion.commit()
        return jsonify({"status": "ok", "message": f"{cursor.rowcount} equipos reingresados a bodega."})
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/equipos/<int:id_equipo>', methods=['DELETE'])
def api_equipos_eliminar(id_equipo):
    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cursor = conexion.cursor()
        cursor.execute("DELETE FROM trazabilidad_equipos WHERE id_equipo = %s", (id_equipo,))
        conexion.commit()
        return jsonify({"status": "ok", "message": "Equipo eliminado correctamente."})
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/equipos_retirados', methods=['GET'])
def api_admin_equipos_retirados_lista():
    estado_custodia = request.args.get('estado_custodia', 'TODOS').strip()
    placa = request.args.get('placa', 'TODAS').strip()
    search = request.args.get('search', '').strip()

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cursor = conexion.cursor(dictionary=True)
        query = """
            SELECT er.id_retiro, er.id_visita, er.tipo_equipo, er.numero_serie, er.modelo,
                   er.motivo_retiro, er.observacion_retiro, er.tecnico, er.placa_vehiculo,
                   er.estado_custodia, er.fecha_retiro, er.fecha_devolucion_bodega, er.recibido_por,
                   v.cliente, v.contrato, v.direccion
            FROM equipos_retirados_visitas er
            LEFT JOIN visitas_tecnicas v ON er.id_visita = v.id_visita
            WHERE 1=1
        """
        params = []
        if estado_custodia and estado_custodia != 'TODOS':
            query += " AND er.estado_custodia = %s"
            params.append(estado_custodia)
        if placa and placa != 'TODAS':
            query += " AND er.placa_vehiculo = %s"
            params.append(placa)
        if search:
            query += " AND (er.numero_serie LIKE %s OR er.modelo LIKE %s OR er.tecnico LIKE %s OR v.cliente LIKE %s)"
            search_param = f"%{search}%"
            params.extend([search_param, search_param, search_param, search_param])

        query += " ORDER BY er.fecha_retiro DESC LIMIT 500"
        cursor.execute(query, tuple(params))
        retirados = cursor.fetchall()

        for r in retirados:
            if r.get('fecha_retiro') and hasattr(r['fecha_retiro'], 'strftime'):
                r['fecha_retiro'] = r['fecha_retiro'].strftime('%Y-%m-%d %H:%M:%S')
            elif r.get('fecha_retiro'):
                r['fecha_retiro'] = str(r['fecha_retiro'])

            if r.get('fecha_devolucion_bodega') and hasattr(r['fecha_devolucion_bodega'], 'strftime'):
                r['fecha_devolucion_bodega'] = r['fecha_devolucion_bodega'].strftime('%Y-%m-%d %H:%M:%S')
            elif r.get('fecha_devolucion_bodega'):
                r['fecha_devolucion_bodega'] = str(r['fecha_devolucion_bodega'])

        return jsonify({
            "status": "ok",
            "equipos_retirados": retirados,
            "total": len(retirados)
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/equipos_retirados/<int:id_retiro>/recibir_bodega', methods=['POST'])
def api_admin_recibir_equipo_retirado(id_retiro):
    token = request.headers.get('Authorization')
    usuario = "BODEGA"
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        u = verify_token(token)
        if u:
            usuario = u.get('username') or 'BODEGA'

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("""
            UPDATE equipos_retirados_visitas
            SET estado_custodia = 'DEVUELTO_BODEGA',
                fecha_devolucion_bodega = NOW(),
                recibido_por = %s
            WHERE id_retiro = %s
        """, (usuario, id_retiro))
        conexion.commit()
        return jsonify({"status": "ok", "message": "Equipo recibido en Bodega Central exitosamente."})
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


@admin_bp.route('/api/admin/equipos_retirados/recibir_masivo', methods=['POST'])
def api_admin_recibir_equipos_masivo():
    token = request.headers.get('Authorization')
    usuario = "BODEGA"
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        u = verify_token(token)
        if u:
            usuario = u.get('username') or 'BODEGA'

    data = request.json or {}
    ids_retiro = data.get('ids_retiro', [])
    if not ids_retiro:
        return jsonify({"status": "error", "message": "No se enviaron IDs de retiro."}), 400

    conexion = get_db_connection()
    if not conexion:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cursor = conexion.cursor(dictionary=True)
        format_strings = ','.join(['%s'] * len(ids_retiro))
        query = f"""
            UPDATE equipos_retirados_visitas
            SET estado_custodia = 'DEVUELTO_BODEGA',
                fecha_devolucion_bodega = NOW(),
                recibido_por = %s
            WHERE id_retiro IN ({format_strings})
        """
        params = [usuario] + [int(x) for x in ids_retiro]
        cursor.execute(query, tuple(params))
        conexion.commit()
        return jsonify({"status": "ok", "message": f"{cursor.rowcount} equipo(s) recibidos en Bodega."})
    except Exception as e:
        conexion.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conexion.close()


# =========================================================
# SECCIÓN: REQUISICIONES DIGITALES CON FIRMA DEL TÉCNICO
# =========================================================

@admin_bp.route('/api/admin/requisiciones', methods=['GET'])
def api_admin_requisiciones_lista():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'sub': session['user_id'], 'role': session.get('user_role'), 'nombre': session.get('user_name')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401

    estado = request.args.get('estado')
    placa = request.args.get('placa')

    conn = get_db_connection()
    if not conn:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cur = conn.cursor(dictionary=True)
        query = "SELECT * FROM requisiciones_materiales WHERE 1=1"
        params = []
        if estado and estado != 'TODOS':
            query += " AND estado = %s"
            params.append(estado)
        if placa and placa != 'TODAS':
            query += " AND placa_vehiculo = %s"
            params.append(placa)
        query += " ORDER BY fecha_solicitud DESC LIMIT 100"
        
        cur.execute(query, tuple(params))
        reqs = cur.fetchall()

        for r in reqs:
            cur.execute("""
                SELECT ri.*, m.stock_bodega
                FROM requisiciones_materiales_items ri
                LEFT JOIN materiales m ON ri.id_material = m.id_material
                WHERE ri.id_requisicion = %s
            """, (r['id_requisicion'],))
            r['items'] = cur.fetchall()
            if r.get('fecha_solicitud'):
                r['fecha_solicitud_fmt'] = r['fecha_solicitud'].strftime('%Y-%m-%d %H:%M')
            if r.get('fecha_aprobacion'):
                r['fecha_aprobacion_fmt'] = r['fecha_aprobacion'].strftime('%Y-%m-%d %H:%M')
            if r.get('fecha_entrega'):
                r['fecha_entrega_fmt'] = r['fecha_entrega'].strftime('%Y-%m-%d %H:%M')

        return jsonify({"status": "ok", "requisiciones": reqs})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@admin_bp.route('/api/admin/requisiciones/crear', methods=['POST'])
def api_admin_requisiciones_crear():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'sub': session['user_id'], 'role': session.get('user_role'), 'nombre': session.get('user_name')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401

    data = request.json or {}
    placa = str(data.get('placa_vehiculo') or '').strip().upper()
    nombre_tecnico = str(data.get('nombre_tecnico') or user.get('nombre') or 'Técnico').strip()
    id_tecnico = data.get('id_tecnico')
    items = data.get('items') or []
    observaciones = data.get('observaciones')

    if not placa:
        return jsonify({"status": "error", "message": "La placa del vehículo es obligatoria."}), 400
    if not items:
        return jsonify({"status": "error", "message": "Debe agregar al menos un material a la solicitud."}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cur = conn.cursor(dictionary=True)

        year = datetime.now().year
        cur.execute("SELECT COUNT(*) as total FROM requisiciones_materiales WHERE YEAR(fecha_solicitud) = %s", (year,))
        count = (cur.fetchone()['total'] or 0) + 1
        numero_solicitud = f"REQ-{year}-{count:04d}"

        creado_por = str(user.get('nombre') or user.get('sub') or 'Usuario').strip()

        cur.execute("""
            INSERT INTO requisiciones_materiales (
                numero_solicitud, placa_vehiculo, nombre_tecnico, id_tecnico,
                fecha_solicitud, estado, observaciones, creado_por
            ) VALUES (%s, %s, %s, %s, NOW(), 'PENDIENTE', %s, %s)
        """, (numero_solicitud, placa, nombre_tecnico, id_tecnico, observaciones, creado_por))
        id_requisicion = cur.lastrowid

        for it in items:
            id_mat = it.get('id_material')
            cant = int(it.get('cantidad_solicitada') or 0)
            if cant <= 0:
                continue

            cur.execute("SELECT codigo_material, nombre_material, unidad_medida FROM materiales WHERE id_material = %s", (id_mat,))
            mat_info = cur.fetchone()
            if mat_info:
                cur.execute("""
                    INSERT INTO requisiciones_materiales_items (
                        id_requisicion, id_material, codigo_material, nombre_material,
                        cantidad_solicitada, cantidad_aprobada, unidad_medida
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (
                    id_requisicion, id_mat, mat_info['codigo_material'], mat_info['nombre_material'],
                    cant, cant, mat_info['unidad_medida']
                ))

        conn.commit()
        return jsonify({"status": "ok", "message": f"Solicitud {numero_solicitud} enviada con éxito.", "id_requisicion": id_requisicion, "numero_solicitud": numero_solicitud})
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@admin_bp.route('/api/admin/requisiciones/<int:id_requisicion>/aprobar', methods=['POST'])
def api_admin_requisiciones_aprobar(id_requisicion):
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'sub': session['user_id'], 'role': session.get('user_role'), 'nombre': session.get('user_name')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401

    data = request.json or {}
    items_aprobados = data.get('items') or []
    aprobado_por = str(user.get('nombre') or 'Bodeguero').strip()

    conn = get_db_connection()
    if not conn:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cur = conn.cursor(dictionary=True)

        for it in items_aprobados:
            id_item = it.get('id_item')
            cant_apr = int(it.get('cantidad_aprobada') or 0)
            cur.execute("UPDATE requisiciones_materiales_items SET cantidad_aprobada = %s WHERE id_item = %s AND id_requisicion = %s", (cant_apr, id_item, id_requisicion))

        cur.execute("""
            UPDATE requisiciones_materiales
            SET estado = 'LISTO_ENTREGA',
                fecha_aprobacion = NOW(),
                aprobado_por = %s
            WHERE id_requisicion = %s
        """, (aprobado_por, id_requisicion))

        conn.commit()
        return jsonify({"status": "ok", "message": "Requisición aprobada y lista para entrega en mostrador."})
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@admin_bp.route('/api/admin/requisiciones/<int:id_requisicion>/firmar_y_entregar', methods=['POST'])
def api_admin_requisiciones_firmar_y_entregar(id_requisicion):
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'sub': session['user_id'], 'role': session.get('user_role'), 'nombre': session.get('user_name')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401

    data = request.json or {}
    firma_base64 = data.get('firma_tecnico')
    entregado_por = str(user.get('nombre') or 'Bodega').strip()

    if not firma_base64:
        return jsonify({"status": "error", "message": "La firma digital del técnico es obligatoria."}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cur = conn.cursor(dictionary=True)

        cur.execute("SELECT * FROM requisiciones_materiales WHERE id_requisicion = %s", (id_requisicion,))
        req = cur.fetchone()
        if not req:
            return jsonify({"status": "error", "message": "Requisición no encontrada"}), 404
        if req['estado'] == 'ENTREGADA':
            return jsonify({"status": "error", "message": "Esta requisición ya fue entregada anteriormente."}), 400

        placa = req['placa_vehiculo']

        cur.execute("SELECT * FROM requisiciones_materiales_items WHERE id_requisicion = %s", (id_requisicion,))
        items = cur.fetchall()

        for it in items:
            id_mat = it['id_material']
            cant = int(it['cantidad_aprobada'] or 0)
            if cant <= 0:
                continue

            # Restar de materiales (bodega)
            cur.execute("UPDATE materiales SET stock_bodega = GREATEST(0, stock_bodega - %s) WHERE id_material = %s", (cant, id_mat))

            # Sumar a inventario_tecnicos de la placa
            cur.execute("SELECT id_inventario, cantidad_disponible FROM inventario_tecnicos WHERE placa_vehiculo = %s AND id_material = %s", (placa, id_mat))
            inv_row = cur.fetchone()
            if inv_row:
                cur.execute("UPDATE inventario_tecnicos SET cantidad_disponible = cantidad_disponible + %s WHERE id_inventario = %s", (cant, inv_row['id_inventario']))
            else:
                cur.execute("INSERT INTO inventario_tecnicos (placa_vehiculo, id_material, cantidad_disponible) VALUES (%s, %s, %s)", (placa, id_mat, cant))

        cur.execute("""
            UPDATE requisiciones_materiales
            SET estado = 'ENTREGADA',
                fecha_entrega = NOW(),
                firma_tecnico = %s,
                entregado_por = %s
            WHERE id_requisicion = %s
        """, (firma_base64, entregado_por, id_requisicion))

        conn.commit()
        return jsonify({
            "status": "ok",
            "message": f"¡Materiales entregados y transferidos exitosamente a la placa {placa}! Comprobante digital firmado guardado.",
            "numero_solicitud": req['numero_solicitud']
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@admin_bp.route('/api/admin/requisiciones/<int:id_requisicion>/rechazar', methods=['POST'])
def api_admin_requisiciones_rechazar(id_requisicion):
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'sub': session['user_id'], 'role': session.get('user_role'), 'nombre': session.get('user_name')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401

    data = request.json or {}
    motivo = data.get('motivo_rechazo') or 'Rechazado por bodega'
    aprobado_por = str(user.get('nombre') or 'Bodeguero').strip()

    conn = get_db_connection()
    if not conn:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cur = conn.cursor()
        cur.execute("""
            UPDATE requisiciones_materiales
            SET estado = 'RECHAZADA',
                fecha_aprobacion = NOW(),
                aprobado_por = %s,
                observaciones = CONCAT(COALESCE(observaciones, ''), ' [RECHAZO: ', %s, ']')
            WHERE id_requisicion = %s
        """, (aprobado_por, motivo, id_requisicion))
        conn.commit()
        return jsonify({"status": "ok", "message": "Requisición rechazada."})
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# =========================================================
# SECCIÓN: CIERRE Y LIQUIDACIÓN MENSUAL DE PLACAS (CORTE)
# =========================================================

@admin_bp.route('/api/admin/inventario/liquidacion_mensual', methods=['GET'])
def api_admin_inventario_liquidacion_mensual():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'sub': session['user_id'], 'role': session.get('user_role'), 'nombre': session.get('user_name')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401

    placa = request.args.get('placa')
    f_ini_str = request.args.get('fecha_inicio')
    f_fin_str = request.args.get('fecha_fin')

    hoy = date.today()
    if not f_fin_str:
        f_fin = date(hoy.year, hoy.month, 26)
    else:
        try:
            f_fin = datetime.strptime(f_fin_str, '%Y-%m-%d').date()
        except:
            f_fin = hoy

    if not f_ini_str:
        if f_fin.month == 1:
            f_ini = date(f_fin.year - 1, 12, 27)
        else:
            f_ini = date(f_fin.year, f_fin.month - 1, 27)
    else:
        try:
            f_ini = datetime.strptime(f_ini_str, '%Y-%m-%d').date()
        except:
            f_ini = date(f_fin.year, f_fin.month - 1 if f_fin.month > 1 else 12, 27)

    if not placa:
        return jsonify({"status": "error", "message": "Debe seleccionar una placa de vehículo."}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cur = conn.cursor(dictionary=True)

        cur.execute("SELECT id_material, codigo_material, nombre_material, unidad_medida, stock_bodega FROM materiales WHERE activo = 1 ORDER BY nombre_material ASC")
        materiales = cur.fetchall()

        cur.execute("SELECT nombre FROM tecnicos WHERE activo = 1 AND (placa_vehiculo = %s OR placa_asignada_hoy = %s) LIMIT 1", (placa, placa))
        tec_row = cur.fetchone()
        tecnico_responsable = tec_row['nombre'] if tec_row else "Técnico Responsable"

        cur.execute("SELECT id_material, cantidad_disponible FROM inventario_tecnicos WHERE placa_vehiculo = %s", (placa,))
        stock_actual_map = {r['id_material']: r['cantidad_disponible'] for r in cur.fetchall()}

        cur.execute("""
            SELECT ri.id_material, SUM(ri.cantidad_aprobada) as total_entregado
            FROM requisiciones_materiales_items ri
            JOIN requisiciones_materiales r ON ri.id_requisicion = r.id_requisicion
            WHERE r.placa_vehiculo = %s
              AND r.estado = 'ENTREGADA'
              AND DATE(r.fecha_entrega) BETWEEN %s AND %s
            GROUP BY ri.id_material
        """, (placa, f_ini, f_fin))
        entregas_map = {r['id_material']: int(r['total_entregado'] or 0) for r in cur.fetchall()}

        cur.execute("""
            SELECT vm.id_material, SUM(vm.cantidad_usada) as total_consumido
            FROM visitas_materiales vm
            JOIN visitas_tecnicas vt ON vm.id_visita = vt.id_visita
            JOIN tecnicos t ON vt.tecnico_principal = t.nombre
            WHERE (t.placa_vehiculo = %s OR t.placa_asignada_hoy = %s)
              AND vt.estado = 'FINALIZADA'
              AND vt.fecha BETWEEN %s AND %s
            GROUP BY vm.id_material
        """, (placa, placa, f_ini, f_fin))
        consumo_map = {r['id_material']: int(r['total_consumido'] or 0) for r in cur.fetchall()}

        cur.execute("""
            SELECT c.id_cierre
            FROM cierres_inventario_mensual c
            WHERE c.placa_vehiculo = %s AND c.fecha_fin < %s
            ORDER BY c.fecha_fin DESC LIMIT 1
        """, (placa, f_ini))
        ultimo_cierre = cur.fetchone()
        stock_inicial_map = {}
        if ultimo_cierre:
            cur.execute("SELECT id_material, conteo_fisico FROM cierres_inventario_mensual_items WHERE id_cierre = %s", (ultimo_cierre['id_cierre'],))
            stock_inicial_map = {r['id_material']: r['conteo_fisico'] for r in cur.fetchall()}

        items_liquidacion = []
        for m in materiales:
            id_m = m['id_material']
            stock_actual = stock_actual_map.get(id_m, 0)
            entregas = entregas_map.get(id_m, 0)
            consumo = consumo_map.get(id_m, 0)
            devoluciones = 0

            if id_m in stock_inicial_map:
                stk_ini = stock_inicial_map[id_m]
            else:
                stk_ini = max(0, stock_actual + consumo - entregas)

            stk_teorico = max(0, stk_ini + entregas - consumo - devoluciones)

            if stk_ini > 0 or entregas > 0 or consumo > 0 or stock_actual > 0:
                items_liquidacion.append({
                    "id_material": id_m,
                    "codigo_material": m['codigo_material'],
                    "nombre_material": m['nombre_material'],
                    "unidad_medida": m['unidad_medida'],
                    "stock_inicial": stk_ini,
                    "entregas_bodega": entregas,
                    "consumo_visitas": consumo,
                    "devoluciones": devoluciones,
                    "stock_teorico": stk_teorico,
                    "stock_actual_sistema": stock_actual,
                    "conteo_fisico": stk_teorico,
                    "diferencia": 0
                })

        periodo_nombre = f"{f_ini.strftime('%d/%b')} al {f_fin.strftime('%d/%b/%Y')}".upper()

        return jsonify({
            "status": "ok",
            "placa": placa,
            "tecnico_responsable": tecnico_responsable,
            "fecha_inicio": f_ini.isoformat(),
            "fecha_fin": f_fin.isoformat(),
            "periodo_nombre": periodo_nombre,
            "items": items_liquidacion
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@admin_bp.route('/api/admin/inventario/cierre_mensual/guardar', methods=['POST'])
def api_admin_inventario_cierre_guardar():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'sub': session['user_id'], 'role': session.get('user_role'), 'nombre': session.get('user_name')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401

    data = request.json or {}
    placa = str(data.get('placa_vehiculo') or '').strip().upper()
    periodo_mes = data.get('periodo_mes') or 'MENSUAL'
    fecha_inicio = data.get('fecha_inicio')
    fecha_fin = data.get('fecha_fin')
    tecnico_responsable = data.get('tecnico_responsable') or 'Técnico'
    items = data.get('items') or []
    observaciones = data.get('observaciones')
    cerrado_por = str(user.get('nombre') or 'Bodega').strip()

    if not placa or not items:
        return jsonify({"status": "error", "message": "Datos incompletos para ejecutar el cierre mensual."}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cur = conn.cursor(dictionary=True)

        total_items = len(items)
        total_faltantes = sum(1 for it in items if int(it.get('diferencia') or 0) < 0)

        cur.execute("""
            INSERT INTO cierres_inventario_mensual (
                periodo_mes, fecha_inicio, fecha_fin, placa_vehiculo, tecnico_responsable,
                total_items, total_faltantes, cerrado_por, fecha_cierre, observaciones
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW(), %s)
        """, (periodo_mes, fecha_inicio, fecha_fin, placa, tecnico_responsable, total_items, total_faltantes, cerrado_por, observaciones))
        id_cierre = cur.lastrowid

        for it in items:
            id_m = it.get('id_material')
            stk_ini = int(it.get('stock_inicial') or 0)
            entregas = int(it.get('entregas_bodega') or 0)
            consumo = int(it.get('consumo_visitas') or 0)
            devoluciones = int(it.get('devoluciones') or 0)
            stk_teorico = int(it.get('stock_teorico') or 0)
            conteo = int(it.get('conteo_fisico') or 0)
            dif = conteo - stk_teorico

            cur.execute("""
                INSERT INTO cierres_inventario_mensual_items (
                    id_cierre, id_material, codigo_material, nombre_material, unidad_medida,
                    stock_inicial, entregas_bodega, consumo_visitas, devoluciones,
                    stock_teorico, conteo_fisico, diferencia
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                id_cierre, id_m, it.get('codigo_material'), it.get('nombre_material'), it.get('unidad_medida') or 'UNIDADES',
                stk_ini, entregas, consumo, devoluciones, stk_teorico, conteo, dif
            ))

            cur.execute("SELECT id_inventario FROM inventario_tecnicos WHERE placa_vehiculo = %s AND id_material = %s", (placa, id_m))
            inv_exist = cur.fetchone()
            if inv_exist:
                cur.execute("UPDATE inventario_tecnicos SET cantidad_disponible = %s WHERE id_inventario = %s", (conteo, inv_exist['id_inventario']))
            else:
                cur.execute("INSERT INTO inventario_tecnicos (placa_vehiculo, id_material, cantidad_disponible) VALUES (%s, %s, %s)", (placa, id_m, conteo))

        conn.commit()
        return jsonify({
            "status": "ok",
            "message": f"¡Cierre mensual de la placa {placa} completado exitosamente! El nuevo stock inicial ha sido fijado.",
            "id_cierre": id_cierre
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@admin_bp.route('/api/admin/inventario/cierres_historico', methods=['GET'])
def api_admin_inventario_cierres_historico():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'sub': session['user_id'], 'role': session.get('user_role'), 'nombre': session.get('user_name')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401

    placa = request.args.get('placa')

    conn = get_db_connection()
    if not conn:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cur = conn.cursor(dictionary=True)
        query = "SELECT * FROM cierres_inventario_mensual WHERE 1=1"
        params = []
        if placa and placa != 'TODAS':
            query += " AND placa_vehiculo = %s"
            params.append(placa)
        query += " ORDER BY fecha_cierre DESC LIMIT 50"
        cur.execute(query, tuple(params))
        cierres = cur.fetchall()

        for c in cierres:
            cur.execute("SELECT * FROM cierres_inventario_mensual_items WHERE id_cierre = %s ORDER BY nombre_material ASC", (c['id_cierre'],))
            c['items'] = cur.fetchall()
            if c.get('fecha_cierre'):
                c['fecha_cierre_fmt'] = c['fecha_cierre'].strftime('%Y-%m-%d %H:%M')
            if c.get('fecha_inicio'):
                c['fecha_inicio_fmt'] = c['fecha_inicio'].strftime('%Y-%m-%d')
            if c.get('fecha_fin'):
                c['fecha_fin_fmt'] = c['fecha_fin'].strftime('%Y-%m-%d')

        return jsonify({"status": "ok", "cierres": cierres})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@admin_bp.route('/api/admin/visitas_materiales_reporte', methods=['GET'])
def api_admin_visitas_materiales_reporte():
    token = request.headers.get('Authorization')
    user = None
    if token and token.startswith("Bearer "):
        from utils_jwt import verify_token
        user = verify_token(token)
    elif 'user_id' in session:
        user = {'sub': session['user_id'], 'role': session.get('user_role'), 'nombre': session.get('user_name')}

    if not user:
        return jsonify({"status": "error", "message": "No autorizado"}), 401

    fecha = request.args.get('fecha', datetime.now().strftime('%Y-%m-%d')).strip()
    tecnico = request.args.get('tecnico', '').strip()
    placa = request.args.get('placa', '').strip()
    estado = request.args.get('estado', '').strip()
    search = request.args.get('search', '').strip()

    conn = get_db_connection()
    if not conn:
        return jsonify({"status": "error", "message": "Error de base de datos"}), 500
    try:
        cur = conn.cursor(dictionary=True)

        query = """
            SELECT v.id_visita, v.fecha_programada, v.hora_inicio_visita, v.hora_fin_visita,
                   v.tecnico_principal, v.tecnico_apoyo, v.estado, v.prioridad,
                   v.cliente, v.contrato, v.telefonos, v.direccion, v.sector, v.servicio,
                   v.problema, v.solucion_tecnico, v.observacion_tecnico,
                   v.modelo_onu, v.numero_serie_onu, v.modelo_router, v.numero_serie_router,
                   v.router_secundario, v.numero_serie_router_secundario,
                   COALESCE(t.placa_asignada_hoy, t.placa_vehiculo, 'S/P') as placa_vehiculo
            FROM visitas_tecnicas v
            LEFT JOIN tecnicos t ON v.tecnico_principal = t.nombre
            WHERE DATE(v.fecha_programada) = %s
        """
        params = [fecha]

        if tecnico and tecnico != 'TODOS':
            query += " AND v.tecnico_principal = %s"
            params.append(tecnico)
        if placa and placa != 'TODAS':
            query += " AND (t.placa_asignada_hoy = %s OR t.placa_vehiculo = %s)"
            params.extend([placa, placa])
        if estado and estado != 'TODOS':
            query += " AND v.estado = %s"
            params.append(estado)
        if search:
            query += " AND (v.cliente LIKE %s OR v.contrato LIKE %s OR v.tecnico_principal LIKE %s OR v.problema LIKE %s OR v.solucion_tecnico LIKE %s)"
            sp = f"%{search}%"
            params.extend([sp, sp, sp, sp, sp])

        query += " ORDER BY v.hora_fin_visita DESC, v.id_visita DESC"
        cur.execute(query, tuple(params))
        visitas = cur.fetchall()

        id_visitas = [v['id_visita'] for v in visitas]

        # 1. Obtener materiales usados
        materiales_por_visita = {}
        resumen_materiales = {}
        total_unidades_insumos = 0

        if id_visitas:
            placeholders = ','.join(['%s'] * len(id_visitas))
            q_mat = f"""
                SELECT vm.id_visita, vm.id_material, vm.cantidad_usada,
                       m.codigo_material, m.nombre_material, m.unidad_medida, m.categoria
                FROM visitas_materiales vm
                JOIN materiales m ON vm.id_material = m.id_material
                WHERE vm.id_visita IN ({placeholders})
                ORDER BY m.nombre_material ASC
            """
            cur.execute(q_mat, tuple(id_visitas))
            rows_mat = cur.fetchall()

            for rm in rows_mat:
                v_id = rm['id_visita']
                if v_id not in materiales_por_visita:
                    materiales_por_visita[v_id] = []
                materiales_por_visita[v_id].append(rm)

                # Acumulador resumen
                m_id = rm['id_material']
                if m_id not in resumen_materiales:
                    resumen_materiales[m_id] = {
                        "id_material": m_id,
                        "codigo_material": rm['codigo_material'],
                        "nombre_material": rm['nombre_material'],
                        "unidad_medida": rm['unidad_medida'],
                        "categoria": rm.get('categoria', 'GENERAL'),
                        "cantidad_total": 0,
                        "visitas_count": 0
                    }
                resumen_materiales[m_id]["cantidad_total"] += rm['cantidad_usada']
                resumen_materiales[m_id]["visitas_count"] += 1
                total_unidades_insumos += rm['cantidad_usada']

        # 2. Obtener equipos retirados
        retirados_por_visita = {}
        total_equipos_retirados = 0
        if id_visitas:
            placeholders = ','.join(['%s'] * len(id_visitas))
            q_ret = f"""
                SELECT er.id_retiro, er.id_visita, er.tipo_equipo, er.modelo, er.numero_serie,
                       er.motivo_retiro, er.observacion_retiro, er.estado_custodia, er.fecha_retiro
                FROM equipos_retirados_visitas er
                WHERE er.id_visita IN ({placeholders})
            """
            cur.execute(q_ret, tuple(id_visitas))
            rows_ret = cur.fetchall()
            for rr in rows_ret:
                v_id = rr['id_visita']
                if v_id not in retirados_por_visita:
                    retirados_por_visita[v_id] = []
                retirados_por_visita[v_id].append(rr)
                total_equipos_retirados += 1

        # 3. Equipos instalados y formateo de visitas
        total_onus_instaladas = 0
        total_routers_instalados = 0

        for v in visitas:
            v_id = v['id_visita']
            v['materiales'] = materiales_por_visita.get(v_id, [])
            v['equipos_retirados'] = retirados_por_visita.get(v_id, [])

            if v.get('fecha_programada'):
                v['fecha_programada_fmt'] = v['fecha_programada'].strftime('%Y-%m-%d')
            if v.get('hora_inicio_visita') and hasattr(v['hora_inicio_visita'], 'isoformat'):
                v['hora_inicio_visita'] = v['hora_inicio_visita'].isoformat()
            if v.get('hora_fin_visita') and hasattr(v['hora_fin_visita'], 'isoformat'):
                v['hora_fin_visita'] = v['hora_fin_visita'].isoformat()

            # Conteo de equipos instalados
            if v.get('numero_serie_onu') and v.get('numero_serie_onu') not in ['None', 'S/N', '']:
                total_onus_instaladas += 1
            if v.get('numero_serie_router') and v.get('numero_serie_router') not in ['None', 'S/N', '']:
                total_routers_instalados += 1
            if v.get('numero_serie_router_secundario') and v.get('numero_serie_router_secundario') not in ['None', 'S/N', '']:
                total_routers_instalados += 1

        # Metricas
        visitas_finalizadas = sum(1 for v in visitas if v['estado'] == 'FINALIZADA')
        visitas_en_progreso = sum(1 for v in visitas if v['estado'] in ['EN_PROGRESO', 'EN_RUTA'])

        resumen_list = sorted(list(resumen_materiales.values()), key=lambda x: x['nombre_material'])

        # Lista de técnicos activos con su placa
        cur.execute("SELECT id_tecnico, nombre, COALESCE(placa_asignada_hoy, placa_vehiculo, 'S/P') as placa FROM tecnicos WHERE activo = 1 ORDER BY nombre ASC")
        tecnicos_activos = cur.fetchall()

        return jsonify({
            "status": "ok",
            "fecha": fecha,
            "tecnicos": tecnicos_activos,
            "totales": {
                "visitas_total": len(visitas),
                "visitas_finalizadas": visitas_finalizadas,
                "visitas_en_progreso": visitas_en_progreso,
                "total_insumos_consumidos": total_unidades_insumos,
                "onus_instaladas": total_onus_instaladas,
                "routers_instalados": total_routers_instalados,
                "equipos_instalados_total": total_onus_instaladas + total_routers_instalados,
                "equipos_retirados_total": total_equipos_retirados,
                "materiales_resumen": resumen_list
            },
            "visitas": visitas
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cur.close()
        conn.close()
